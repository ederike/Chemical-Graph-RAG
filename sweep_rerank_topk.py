#!/usr/bin/env python3
"""
rerank_top_k 敏感性分析（走 benchmark 三步）。

  1) generate   只在 TDS 文档池（chunk_max_vectors 对应的 doc）里随机抽样出题
  2) evaluate   同一问题集，依次改 retrieve.rerank_top_k 做评测
  3) summarize  汇总准确率 / 召回 / 延迟，制表 + 曲线

结果根目录 RESULTS_ROOT：

  {RESULTS_ROOT}/questions_complex.json
  {RESULTS_ROOT}/benchmark_result_topk={k}/evals_complex.json
  {RESULTS_ROOT}/rerank_topk_sweep_summary.json|.csv|.xlsx
  {RESULTS_ROOT}/rerank_topk_acc_recall.png
  {RESULTS_ROOT}/rerank_topk_latency.png

改下面「可调参数」，或::

    python sweep_rerank_topk.py              # 三步全跑
    python sweep_rerank_topk.py generate
    python sweep_rerank_topk.py evaluate
    python sweep_rerank_topk.py summarize
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# 可调参数
# ---------------------------------------------------------------------------

# generate | evaluate | summarize | all（可被命令行覆盖）
STEP: str = "summarize"

# RERANK_TOP_KS: List[int] = [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
RERANK_TOP_KS: List[int] = [23,24,25,26,27,28,29,30]

BENCHMARK_CONFIG: str = "benchmark/config.yaml"
RESULTS_ROOT: str = "benchmark_results_sen-analysis-topk"

# 出题量；None 则用 benchmark/config.yaml 的 generate.hop_counts
HOP_COUNTS: Optional[Dict[int, int]] = None
# 抽样随机种子；None 则用 yaml generate.seed
SEED: Optional[int] = None

# True：已有 questions / 某 k 的 evals 则跳过该步
SKIP_EXISTING: bool = True

# 评测前 pin TDS 范围内的 FAISS 分片；全部 k 跑完再 unpin
PIN_INDEXES: bool = True

# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from useless.benchmark.config import BenchmarkConfig
from useless.benchmark.evaluator import QueryEvaluator
from useless.benchmark.utils import resolve_path
from useless.benchmark.workflow import TestQueryWorkflow
from src.utils.config import parse_vector_id_range

_TOPK_DIR_RE = re.compile(r"^benchmark_result_topk=(\d+)$")


def _as_path(p: str | Path) -> Path:
    path = Path(p)
    return path if path.is_absolute() else resolve_path(path)


def _uniq_ks(raw: List[int]) -> List[int]:
    out: List[int] = []
    seen = set()
    for k in raw:
        k = int(k)
        if k in seen:
            continue
        seen.add(k)
        out.append(k)
    return out


def _find_evals_file(folder: Path, preferred: str) -> Optional[Path]:
    cand = folder / preferred
    if cand.is_file():
        return cand
    hits = sorted(folder.glob("evals*.json"))
    return hits[0] if hits else None


def _tds_max_doc_id(
    db_path: Path,
    chunk_max_vectors: int,
    chunk_lo: int = 0,
) -> int:
    """范围内 chunk 的 max(doc_id)。lo/hi 为 0 表示该端不限制。"""
    try:
        cap = int(chunk_max_vectors or 0)
    except (TypeError, ValueError):
        cap = 0
    try:
        lo = int(chunk_lo or 0)
    except (TypeError, ValueError):
        lo = 0
    if cap <= 0 and lo <= 0:
        return 0
    sql = "SELECT MAX(doc_id) FROM chunk WHERE 1=1"
    params = []
    if lo > 0:
        sql += " AND id >= ?"
        params.append(lo)
    if cap > 0:
        sql += " AND id <= ?"
        params.append(cap)
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        row = conn.execute(sql, tuple(params)).fetchone()
    finally:
        conn.close()
    if not row or row[0] is None:
        return cap
    return int(row[0])


def _questions_path(results_root: Path, cfg: BenchmarkConfig) -> Path:
    name = cfg.questions_filename or "questions_complex.json"
    return results_root / name


def _run_generate(
    *,
    config_path: str,
    results_root: Path,
    hop_counts: Optional[Dict[int, int]],
    seed: Optional[int],
) -> Path:
    results_root.mkdir(parents=True, exist_ok=True)
    cfg = BenchmarkConfig.from_yaml(
        config_path,
        overrides={
            "run": {"mode": "generate"},
            "paths": {
                "output_dir": str(results_root),
            },
        },
    )
    if hop_counts:
        cfg.hop_counts = {int(h): int(n) for h, n in hop_counts.items()}
    if seed is not None:
        cfg.seed = int(seed)

    wf = TestQueryWorkflow(cfg)
    dcfg = wf._load_dhmf_config()
    chunk_lo, chunk_hi = parse_vector_id_range(
        getattr(dcfg.retrieve, "chunk_max_vectors", 0)
    )
    db_path = Path(cfg.db_path)
    max_doc_id = _tds_max_doc_id(db_path, chunk_hi, chunk_lo)
    out = wf.cfg.questions_file()

    print(
        f"[sweep] generate TDS pool chunk_max_vectors="
        f"{getattr(dcfg.retrieve, 'chunk_max_vectors', 0)!r} "
        f"id_range={chunk_lo}..{chunk_hi} max_doc_id={max_doc_id} "
        f"hop_counts={wf.cfg.hop_counts} seed={wf.cfg.seed} → {out}",
        file=sys.stderr,
    )
    dataset = wf.generate_questions(save=True, max_doc_id=max_doc_id)
    meta = dataset.setdefault("meta", {})
    meta["scope"] = "tds"
    meta["chunk_max_vectors"] = getattr(dcfg.retrieve, "chunk_max_vectors", 0)
    meta["max_doc_id"] = max_doc_id
    wf.save_json(dataset, out)
    n = len(dataset.get("questions") or [])
    ok = sum(1 for q in (dataset.get("questions") or []) if q.get("gen_status") == 1)
    print(
        f"[sweep] generate done questions={n} ok={ok} "
        f"source_docs={meta.get('n_source_docs')}",
        file=sys.stderr,
    )
    return out


def _run_one_eval(
    *,
    config_path: str,
    out_dir: Path,
    questions_path: Path,
    k: int,
    dcfg,
    dhmf,
) -> Path:
    cfg = BenchmarkConfig.from_yaml(
        config_path,
        overrides={
            "run": {"mode": "evaluate"},
            "paths": {
                "output_dir": str(out_dir),
                "eval_questions_path": str(questions_path),
            },
        },
    )
    wf = TestQueryWorkflow(cfg)
    wf.dhmf_config = dcfg
    wf.dhmf = dhmf
    dcfg.retrieve.rerank_top_k = int(k)
    dcfg.retrieve.enable_rerank = True

    print(
        f"[sweep] evaluate k={k}  rerank_top_k={dcfg.retrieve.rerank_top_k}  → {out_dir}",
        file=sys.stderr,
    )
    eval_data = wf.evaluate(save=True)
    eval_data.setdefault("meta", {})["rerank_top_k"] = int(k)
    eval_data.setdefault("meta", {})["scope"] = "tds"
    out = wf.cfg.eval_results_file()
    wf.save_json(eval_data, out)
    return out


def _run_evaluate(
    *,
    config_path: str,
    results_root: Path,
    questions_path: Path,
    ks: List[int],
    evals_name: str,
    skip_existing: bool,
    pin_indexes: bool,
) -> None:
    if not questions_path.is_file():
        raise FileNotFoundError(
            f"问题集不存在，请先 generate：{questions_path}"
        )
    planned = [(results_root / f"benchmark_result_topk={k}", k) for k in ks]
    print(f"[sweep] questions  = {questions_path}", file=sys.stderr)
    print(f"[sweep] rerank_top_k = {ks}", file=sys.stderr)

    print("[sweep] loading DHMF (once) …", file=sys.stderr)
    bootstrap = TestQueryWorkflow.from_config(config_path)
    dcfg = bootstrap._load_dhmf_config()
    dhmf = bootstrap.setup_dhmf()

    pinned = False
    if pin_indexes and hasattr(dhmf, "pin_retrieve_indexes"):
        print("[sweep] pin_retrieve_indexes …", file=sys.stderr)
        stats = dhmf.pin_retrieve_indexes()
        pinned = True
        for name, st in (stats or {}).items():
            print(
                f"[sweep]   pinned {name}: shards={len(st.get('shards') or [])} "
                f"ntotal={st.get('ntotal')} dt={st.get('seconds')}s",
                file=sys.stderr,
            )

    failed: List[Tuple[int, str]] = []
    try:
        for out_dir, k in planned:
            evals_path = out_dir / evals_name
            if skip_existing and evals_path.is_file():
                print(f"[sweep] skip k={k}（已有 {evals_path}）", file=sys.stderr)
                continue
            try:
                _run_one_eval(
                    config_path=config_path,
                    out_dir=out_dir,
                    questions_path=questions_path,
                    k=k,
                    dcfg=dcfg,
                    dhmf=dhmf,
                )
                print(f"[sweep] done k={k}", file=sys.stderr)
            except KeyboardInterrupt:
                print(f"[sweep] interrupted at k={k}", file=sys.stderr)
                raise
            except Exception as e:
                print(f"[sweep] failed k={k}: {e}", file=sys.stderr)
                failed.append((k, str(e)))
    finally:
        if pinned and hasattr(dhmf, "unpin_retrieve_indexes"):
            print("[sweep] unpin_retrieve_indexes …", file=sys.stderr)
            dhmf.unpin_retrieve_indexes()

    if failed:
        print("[sweep] unfinished:", file=sys.stderr)
        for k, msg in failed:
            print(f"  k={k}: {msg}", file=sys.stderr)


def _metrics_from_eval(eval_data: Dict[str, Any]) -> Dict[str, Any]:
    """用现有 evaluator 汇总，取超图一路的准确率 / 召回 / 延迟。"""
    report = QueryEvaluator.build_report_document(eval_data)
    summary = report.get("summary") or {}
    hg = summary.get("hypergraph") if isinstance(summary.get("hypergraph"), dict) else summary
    acc = (hg.get("llm_acc") or {}) if isinstance(hg.get("llm_acc"), dict) else {}
    rec = (hg.get("doc_recall") or summary.get("doc_recall") or {})
    if not isinstance(rec, dict):
        rec = {}
    lat = hg.get("latency") if isinstance(hg.get("latency"), dict) else {}
    meta = report.get("meta") or {}
    return {
        "n_results": meta.get("n_results"),
        "eval_done": meta.get("eval_done"),
        "accuracy": acc.get("accuracy"),
        "n_judged": acc.get("n_judged"),
        "n_correct": acc.get("n_correct"),
        "n_wrong": acc.get("n_wrong"),
        "mean_recall": rec.get("mean_recall"),
        "mean_query_s": lat.get("mean_query_s"),
        "mean_retrieve_s": lat.get("mean_retrieve_s"),
        "mean_rerank_s": lat.get("mean_rerank_s"),
        "mean_wall_s": lat.get("mean_wall_s"),
        "by_hop": hg.get("by_hop") or {},
        "report": report,
    }


def _discover_runs(results_root: Path, evals_name: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not results_root.is_dir():
        return rows
    for folder in sorted(results_root.iterdir()):
        if not folder.is_dir():
            continue
        m = _TOPK_DIR_RE.match(folder.name)
        if not m:
            continue
        k = int(m.group(1))
        ev_path = _find_evals_file(folder, evals_name)
        if ev_path is None:
            print(f"[sweep] skip {folder.name}：无 evals JSON", file=sys.stderr)
            continue
        try:
            with open(ev_path, "r", encoding="utf-8") as f:
                eval_data = json.load(f)
        except Exception as e:
            print(f"[sweep] skip {folder.name}：读 {ev_path.name} 失败: {e}", file=sys.stderr)
            continue
        if not isinstance(eval_data, dict) or "results" not in eval_data:
            print(f"[sweep] skip {folder.name}：不是评测结果 JSON", file=sys.stderr)
            continue
        try:
            metrics = _metrics_from_eval(eval_data)
        except Exception as e:
            print(f"[sweep] skip {folder.name}：汇总失败: {e}", file=sys.stderr)
            continue
        rows.append({
            "rerank_top_k": k,
            "dir": folder.name,
            "evals_path": str(ev_path),
            **metrics,
        })
    rows.sort(key=lambda r: r["rerank_top_k"])
    return rows


def _configure_chinese_font() -> None:
    try:
        from matplotlib import font_manager
        import matplotlib.pyplot as plt
    except ImportError:
        return
    font_dirs = [
        ROOT / "scripts" / "assets" / "fonts",
        ROOT / "assets" / "fonts",
    ]
    for font_dir in font_dirs:
        if not font_dir.is_dir():
            continue
        for font_path in sorted(font_dir.glob("NotoSansSC-Regular.*")):
            try:
                font_manager.fontManager.addfont(str(font_path))
                family = font_manager.FontProperties(fname=str(font_path)).get_name()
                plt.rcParams["font.family"] = "sans-serif"
                plt.rcParams["font.sans-serif"] = [family, "DejaVu Sans"]
                plt.rcParams["axes.unicode_minus"] = False
                return
            except Exception:
                continue
    plt.rcParams["axes.unicode_minus"] = False


def _xy(rows: List[dict], key: str) -> Tuple[List[int], List[float]]:
    xs, ys = [], []
    for r in rows:
        v = r.get(key)
        if v is None:
            continue
        try:
            ys.append(float(v))
        except (TypeError, ValueError):
            continue
        xs.append(int(r["rerank_top_k"]))
    return xs, ys


def _annotate_points(
    ax,
    xs: List[int],
    ys: List[float],
    *,
    fmt: str,
    dy: int,
    color: Optional[str] = None,
) -> None:
    """在每个数据点旁水平标数值。"""
    if not xs:
        return
    va = "bottom" if dy >= 0 else "top"
    for x, y in zip(xs, ys):
        ax.annotate(
            fmt.format(y),
            xy=(x, y),
            xytext=(0, dy),
            textcoords="offset points",
            ha="center",
            va=va,
            fontsize=7,
            color=color,
            rotation=0,
            clip_on=False,
        )


def _pad_ylim(ax, series: List[List[float]], *, lower: Optional[float] = None) -> None:
    ys = [y for g in series for y in g]
    if not ys:
        return
    lo = min(ys) if lower is None else lower
    hi = max(ys)
    span = max(hi - lo, 1e-6)
    top = hi + 0.16 * span
    if lower is not None and top < 1.08 and hi <= 1.0:
        top = 1.08
    ax.set_ylim(lo if lower is not None else lo - 0.08 * span, top)


def _plot_curves(rows: List[dict], out_dir: Path) -> Dict[str, str]:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    _configure_chinese_font()
    out_dir.mkdir(parents=True, exist_ok=True)
    saved: Dict[str, str] = {}
    n_pts = max(len(rows), 8)
    fig_w = max(8.5, min(16.0, 0.55 * n_pts + 3.0))

    def _style(ax, xs_all: List[int]) -> None:
        ax.grid(True, linestyle="--", alpha=0.4)
        ax.set_xlabel("rerank_top_k")
        ticks = sorted(set(xs_all))
        if ticks:
            ax.set_xticks(ticks)

    fig, ax = plt.subplots(figsize=(fig_w, 5.4))
    x_acc, y_acc = _xy(rows, "accuracy")
    x_rec, y_rec = _xy(rows, "mean_recall")
    xs_all: List[int] = []
    if x_acc:
        line, = ax.plot(x_acc, y_acc, marker="o", label="准确率")
        _annotate_points(ax, x_acc, y_acc, fmt="{:.3f}", dy=8, color=line.get_color())
        xs_all.extend(x_acc)
    if x_rec:
        line, = ax.plot(x_rec, y_rec, marker="s", label="文档召回率")
        _annotate_points(ax, x_rec, y_rec, fmt="{:.3f}", dy=-10, color=line.get_color())
        xs_all.extend(x_rec)
    _style(ax, xs_all)
    ax.set_ylabel("比例")
    _pad_ylim(ax, [y_acc, y_rec], lower=0.0)
    ax.set_title("准确率 / 召回率 vs rerank_top_k")
    ax.legend()
    fig.tight_layout()
    p1 = out_dir / "rerank_topk_acc_recall.png"
    fig.savefig(p1, dpi=150)
    plt.close(fig)
    saved["acc_recall"] = str(p1)

    fig, ax = plt.subplots(figsize=(fig_w, 5.4))
    series = (
        ("mean_query_s", "问答耗时", "o", 8),
        ("mean_retrieve_s", "检索耗时", "s", 8),
        ("mean_rerank_s", "rerank 耗时", "^", 8),
    )
    any_line = False
    ys_all: List[List[float]] = []
    xs_all = []
    for key, label, marker, dy in series:
        xs, ys = _xy(rows, key)
        if not xs:
            continue
        line, = ax.plot(xs, ys, marker=marker, label=label)
        _annotate_points(ax, xs, ys, fmt="{:.2f}", dy=dy, color=line.get_color())
        any_line = True
        ys_all.append(ys)
        xs_all.extend(xs)
    _style(ax, xs_all)
    ax.set_ylabel("秒")
    _pad_ylim(ax, ys_all)
    ax.set_title("延迟 vs rerank_top_k")
    if any_line:
        ax.legend()
    fig.tight_layout()
    p2 = out_dir / "rerank_topk_latency.png"
    fig.savefig(p2, dpi=150)
    plt.close(fig)
    saved["latency"] = str(p2)
    return saved


def _table_fields() -> List[str]:
    return [
        "rerank_top_k",
        "n_results",
        "n_judged",
        "n_correct",
        "n_wrong",
        "accuracy",
        "mean_recall",
        "mean_query_s",
        "mean_retrieve_s",
        "mean_rerank_s",
        "mean_wall_s",
        "dir",
        "evals_path",
    ]


def _row_for_table(r: dict) -> Dict[str, Any]:
    return {k: r.get(k) for k in _table_fields()}


def _write_csv(rows: List[dict], path: Path) -> Path:
    fields = _table_fields()
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(_row_for_table(r))
    return path


def _write_xlsx(rows: List[dict], path: Path) -> Optional[Path]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "rerank_top_k"
    fields = _table_fields()
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        rec = _row_for_table(r)
        ws.append([rec.get(k) for k in fields])
    hop_rows = []
    for r in rows:
        by_hop = r.get("by_hop") or {}
        if not isinstance(by_hop, dict):
            continue
        for hop, stats in sorted(by_hop.items(), key=lambda x: str(x[0])):
            if not isinstance(stats, dict):
                continue
            acc = stats.get("llm_acc") if isinstance(stats.get("llm_acc"), dict) else {}
            rec = stats.get("doc_recall") if isinstance(stats.get("doc_recall"), dict) else {}
            hop_rows.append({
                "rerank_top_k": r.get("rerank_top_k"),
                "hop": hop,
                "n": stats.get("n") or acc.get("n_judged"),
                "accuracy": acc.get("accuracy"),
                "mean_recall": rec.get("mean_recall"),
            })
    if hop_rows:
        ws2 = wb.create_sheet("by_hop")
        hdr = ["rerank_top_k", "hop", "n", "accuracy", "mean_recall"]
        ws2.append(hdr)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for rec in hop_rows:
            ws2.append([rec.get(k) for k in hdr])
    wb.save(path)
    return path


def _write_summary(rows: List[dict], results_root: Path) -> Path:
    plots = {}
    try:
        plots = _plot_curves(rows, results_root)
    except Exception as e:
        print(f"[sweep] plot failed: {e}", file=sys.stderr)

    slim = []
    for r in rows:
        item = dict(r)
        item.pop("report", None)
        slim.append(item)

    csv_path = results_root / "rerank_topk_sweep_summary.csv"
    _write_csv(slim, csv_path)
    xlsx_path = results_root / "rerank_topk_sweep_summary.xlsx"
    xlsx = _write_xlsx(slim, xlsx_path)

    doc = {
        "meta": {
            "results_root": str(results_root),
            "n_runs": len(rows),
            "rerank_top_ks": [r["rerank_top_k"] for r in rows],
        },
        "runs": slim,
        "plots": plots,
        "tables": {
            "csv": str(csv_path),
            "xlsx": str(xlsx) if xlsx else None,
        },
    }
    out = results_root / "rerank_topk_sweep_summary.json"
    out.write_text(
        json.dumps(doc, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[sweep] summary → {out}", file=sys.stderr)
    print(f"[sweep] table csv → {csv_path}", file=sys.stderr)
    if xlsx:
        print(f"[sweep] table xlsx → {xlsx}", file=sys.stderr)
    for name, path in plots.items():
        print(f"[sweep] plot {name} → {path}", file=sys.stderr)
    return out


def _print_table(rows: List[dict]) -> None:
    if not rows:
        print("[sweep] 没有可汇总的 evals", file=sys.stderr)
        return
    hdr = (
        f"{'k':>4}  {'acc':>7}  {'recall':>7}  "
        f"{'query_s':>8}  {'retr_s':>8}  {'rerank_s':>8}  n"
    )
    print(hdr, file=sys.stderr)
    for r in rows:
        def _f(key, w):
            v = r.get(key)
            if v is None:
                return f"{'—':>{w}}"
            return f"{float(v):>{w}.3f}"

        print(
            f"{r['rerank_top_k']:>4}  {_f('accuracy', 7)}  {_f('mean_recall', 7)}  "
            f"{_f('mean_query_s', 8)}  {_f('mean_retrieve_s', 8)}  "
            f"{_f('mean_rerank_s', 8)}  {r.get('n_results') or '—'}",
            file=sys.stderr,
        )


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="rerank_top_k 敏感性分析：generate → evaluate → summarize",
    )
    p.add_argument(
        "step",
        nargs="?",
        default=STEP,
        choices=("generate", "evaluate", "summarize", "all"),
        help="跑哪一步（默认脚本顶部 STEP）",
    )
    return p.parse_args()


def main() -> int:
    args = _parse_args()
    step = str(args.step or STEP).strip().lower()
    ks = _uniq_ks(RERANK_TOP_KS)
    config_path = str(_as_path(BENCHMARK_CONFIG))
    results_root = _as_path(RESULTS_ROOT)
    results_root.mkdir(parents=True, exist_ok=True)

    base_cfg = BenchmarkConfig.from_yaml(config_path)
    evals_name = base_cfg.eval_results_filename or "evals_complex.json"
    questions = _questions_path(results_root, base_cfg)

    print(f"[sweep] step       = {step}", file=sys.stderr)
    print(f"[sweep] config     = {config_path}", file=sys.stderr)
    print(f"[sweep] results    = {results_root}", file=sys.stderr)

    if step in ("generate", "all"):
        if SKIP_EXISTING and questions.is_file():
            print(f"[sweep] skip generate（已有 {questions}）", file=sys.stderr)
        else:
            try:
                _run_generate(
                    config_path=config_path,
                    results_root=results_root,
                    hop_counts=HOP_COUNTS,
                    seed=SEED,
                )
            except KeyboardInterrupt:
                print("[sweep] interrupted during generate", file=sys.stderr)
                return 130

    if step in ("evaluate", "all"):
        if not ks:
            print("[sweep] RERANK_TOP_KS 为空", file=sys.stderr)
            return 2
        try:
            _run_evaluate(
                config_path=config_path,
                results_root=results_root,
                questions_path=questions,
                ks=ks,
                evals_name=evals_name,
                skip_existing=SKIP_EXISTING,
                pin_indexes=PIN_INDEXES,
            )
        except KeyboardInterrupt:
            print("[sweep] interrupted during evaluate", file=sys.stderr)
            return 130
        except FileNotFoundError as e:
            print(f"[sweep] {e}", file=sys.stderr)
            return 2

    if step in ("summarize", "all"):
        print("[sweep] summarizing", results_root, file=sys.stderr)
        rows = _discover_runs(results_root, evals_name)
        _print_table(rows)
        if rows:
            _write_summary(rows, results_root)

    print("[sweep] all done", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
