#!/usr/bin/env python3
"""把 benchmark3 的 report / evals 写成总结 Excel。"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

from benchmark2.evaluator import SYSTEM_HYPERGRAPH, SYSTEM_LLM_ONLY
from benchmark2.report import build_report_document, project_system_row

from .dataset import DIM_KEYS, dimension_key

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:  # pragma: no cover
    raise ImportError("写出 Excel 需要 openpyxl，请 pip install openpyxl") from e


SYSTEM_LABELS = {
    SYSTEM_HYPERGRAPH: "超图",
    SYSTEM_LLM_ONLY: "纯LLM",
}
EXCEL_CELL_MAX = 32767
_INVALID_SHEET_RE = re.compile(r"[\[\]\*?:/\\]")

_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FILL_HG = PatternFill("solid", fgColor="D6EAF8")
_FILL_POS = PatternFill("solid", fgColor="D5F5E3")
_FILL_NEG = PatternFill("solid", fgColor="FADBD8")
_FILL_MID = PatternFill("solid", fgColor="FCF3CF")
_FILL_ALT = PatternFill("solid", fgColor="F7F9FB")
_FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
_FONT_HEADER = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_FONT_TITLE = Font(name="微软雅黑", bold=True, color="FFFFFF", size=14)
_FONT_CELL = Font(name="微软雅黑", size=10)
_FONT_BOLD = Font(name="微软雅黑", bold=True, size=10)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_R = Alignment(horizontal="right", vertical="center")


def _sheet_name(name: str, used: Optional[set] = None) -> str:
    s = _INVALID_SHEET_RE.sub(" ", str(name or "Sheet")).strip() or "Sheet"
    s = s[:31]
    if used is None:
        return s
    base, n = s, 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = base[: 31 - len(suffix)] + suffix
    used.add(s)
    return s


def _cell_text(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v)
    if len(s) > EXCEL_CELL_MAX:
        s = s[: EXCEL_CELL_MAX - 1] + "…"
    return s


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _fmt_pct(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _sys_label(key: str) -> str:
    return SYSTEM_LABELS.get(key, key)


def _nested(d: Optional[dict], *keys: str, default: Any = None) -> Any:
    cur: Any = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur


def _apply_number_format(cell, fmt: Optional[str]) -> None:
    if fmt == "pct":
        cell.number_format = "0.00%"
    elif fmt == "int":
        cell.number_format = "#,##0"
    elif fmt == "float2":
        cell.number_format = "0.00"
    elif fmt == "float3":
        cell.number_format = "0.000"


def _auto_width(ws: Worksheet, *, min_w: float = 8, max_w: float = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0.0
        for cell in col:
            if cell.value is None:
                continue
            text = str(cell.value).replace("\n", " ")
            w = 0.0
            for ch in text[:80]:
                w += 1.7 if ord(ch) > 127 else 1.05
            longest = max(longest, w)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 1.5))


def _write_title(ws: Worksheet, title: str, n_cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, n_cols))
    cell = ws.cell(1, 1, title)
    cell.fill = _FILL_TITLE
    cell.font = _FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def _write_table(
    ws: Worksheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    start_row: int = 1,
    col_formats: Optional[Sequence[Optional[str]]] = None,
    freeze: bool = True,
    autofilter: bool = True,
    zebra: bool = True,
    max_width: float = 48,
    wrap: bool = False,
) -> int:
    n_cols = len(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _THIN
    ws.row_dimensions[start_row].height = 22

    align_text = _ALIGN_L if wrap else Alignment(horizontal="left", vertical="center")
    for i, row in enumerate(rows):
        r = start_row + 1 + i
        alt = zebra and (i % 2 == 1)
        for c in range(1, n_cols + 1):
            raw = row[c - 1] if c - 1 < len(row) else None
            val = _cell_text(raw)
            cell = ws.cell(r, c, val)
            cell.font = _FONT_CELL
            cell.border = _THIN
            fmt = col_formats[c - 1] if col_formats and c - 1 < len(col_formats) else None
            _apply_number_format(cell, fmt)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = _ALIGN_R
            else:
                cell.alignment = align_text
            if alt:
                cell.fill = _FILL_ALT
        if wrap:
            ws.row_dimensions[r].height = 72

    last = start_row + len(rows)
    if autofilter and rows:
        ws.auto_filter.ref = (
            f"{get_column_letter(1)}{start_row}:{get_column_letter(n_cols)}{last}"
        )
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"
    _auto_width(ws, max_w=max_width)
    ws.sheet_view.showGridLines = False
    return last + 1


def _flatten_meta_value(v: Any) -> Any:
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    return json.dumps(v, ensure_ascii=False)


def _systems(summary: Dict[str, Any]) -> List[str]:
    raw = summary.get("systems")
    out: List[str] = []
    if isinstance(raw, list):
        for k in raw:
            if k in SYSTEM_LABELS:
                out.append(k)
    if not out:
        for k in (SYSTEM_HYPERGRAPH, SYSTEM_LLM_ONLY):
            block = summary.get(k)
            if isinstance(block, dict) and block.get("judgment") is not None:
                out.append(k)
    return out or [SYSTEM_HYPERGRAPH]


def _system_block(result: Dict[str, Any], system: str) -> Dict[str, Any]:
    block = result.get(system)
    if isinstance(block, dict) and block:
        return block
    row = project_system_row(result, system)
    return {
        "answer": result.get("rag_answer") if system == SYSTEM_HYPERGRAPH else result.get("llm_only_answer"),
        "raw_answer": result.get("rag_raw_answer") if system == SYSTEM_HYPERGRAPH else None,
        "llm_acc": row.get("llm_acc"),
        "score": row.get("score"),
        "query_status": row.get("query_status"),
        "query_error": row.get("query_error"),
        "judge_status": row.get("judge_status"),
        "judge_error": row.get("judge_error"),
        "metrics": row.get("metrics") or {},
        "judge_reason": result.get("judge_reason") if system == SYSTEM_HYPERGRAPH else (block or {}).get("judge_reason"),
        "dimension_scores": row.get("dimension_scores") or [],
        "retrieval_sources": result.get("retrieval_sources") if system == SYSTEM_HYPERGRAPH else [],
    }


def _dim_map(block: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for i, d in enumerate(block.get("dimension_scores") or [], 1):
        if not isinstance(d, dict):
            continue
        key = dimension_key(i, str(d.get("dimension") or ""))
        out[key] = d
    return out


def _judgment_fill(label: Any) -> Optional[PatternFill]:
    s = str(label or "")
    if s == "正确":
        return _FILL_POS
    if s == "部分正确":
        return _FILL_MID
    if s == "错误":
        return _FILL_NEG
    return None


def _score_fill(score: Any) -> Optional[PatternFill]:
    v = _int(score)
    if v is None:
        return None
    if v >= 2:
        return _FILL_POS
    if v == 1:
        return _FILL_MID
    return _FILL_NEG


def _add_toc(wb: Workbook, items: Sequence[Tuple[str, str]], *, title: str) -> None:
    ws = wb.create_sheet(_sheet_name("目录"), 0)
    _write_title(ws, title, 2)
    ws.cell(2, 1, "点击工作表名跳转。")
    ws.cell(2, 1).font = Font(name="微软雅黑", italic=True, size=10, color="666666")
    _write_table(
        ws,
        ("工作表", "说明"),
        [(name, desc) for name, desc in items],
        start_row=4,
        freeze=False,
        autofilter=False,
    )
    for i, (sheet, _) in enumerate(items):
        cell = ws.cell(5 + i, 1, sheet)
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet}'!A1", display=sheet)
        cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 72


def _add_meta(wb: Workbook, report: Dict[str, Any], eval_data: Optional[Dict[str, Any]]) -> str:
    name = _sheet_name("运行信息")
    ws = wb.create_sheet(name)
    meta = dict(report.get("meta") or {})
    ev_meta = (eval_data or {}).get("meta") if isinstance(eval_data, dict) else {}
    if isinstance(ev_meta, dict):
        for k, v in ev_meta.items():
            meta.setdefault(k, v)
    skip = {"config", "dataset_meta", "progress", "stats_meta"}
    rows: List[List[Any]] = []
    preferred = [
        "created_at",
        "updated_at",
        "eval_done",
        "done",
        "query_mode",
        "judge_model",
        "enable_llm_only",
        "n_results",
        "n_questions",
        "n_total_planned",
        "n_resumed",
        "eval_elapsed_s",
        "elapsed_s",
        "csv_path",
        "source_path",
        "config_file",
        "dhmf_config_path",
    ]
    seen = set()
    for k in preferred:
        if k in meta:
            rows.append([k, _flatten_meta_value(meta[k])])
            seen.add(k)
    for k in sorted(meta.keys()):
        if k in seen or k in skip:
            continue
        rows.append([k, _flatten_meta_value(meta[k])])
    cfg = meta.get("config") if isinstance(meta.get("config"), dict) else None
    if cfg:
        rows.append(["— 配置快照 —", ""])
        for k, v in cfg.items():
            rows.append([f"config.{k}", _flatten_meta_value(v)])
    _write_title(ws, "运行信息", 2)
    _write_table(ws, ("字段", "值"), rows, start_row=3, freeze=True, autofilter=False, max_width=80)
    return name


def _dimension_stats(results: Sequence[Dict[str, Any]], system: str) -> Dict[str, Dict[str, Any]]:
    bag: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"n": 0, "n_pass": 0, "n_fail": 0, "scores": [], "comments": 0}
    )
    for r in results:
        block = _system_block(r, system)
        used = set()
        for i, d in enumerate(block.get("dimension_scores") or [], 1):
            if not isinstance(d, dict):
                continue
            key = dimension_key(i, str(d.get("dimension") or ""))
            if key in used:
                key = f"{key}#{i}"
            used.add(key)
            rec = bag[key]
            rec["n"] += 1
            if d.get("pass") is True:
                rec["n_pass"] += 1
            elif d.get("pass") is False:
                rec["n_fail"] += 1
            sc = _num(d.get("score"))
            if sc is not None:
                rec["scores"].append(sc)
            if d.get("comment"):
                rec["comments"] += 1
    ordered: Dict[str, Dict[str, Any]] = {}
    for key in list(DIM_KEYS) + [k for k in bag if k not in DIM_KEYS]:
        rec = bag.get(key)
        if not rec or rec["n"] == 0:
            continue
        scores = rec["scores"]
        ordered[key] = {
            "n": rec["n"],
            "n_pass": rec["n_pass"],
            "n_fail": rec["n_fail"],
            "pass_rate": (rec["n_pass"] / rec["n"]) if rec["n"] else None,
            "mean": (sum(scores) / len(scores)) if scores else None,
        }
    return ordered


def _add_overview(
    wb: Workbook,
    summary: Dict[str, Any],
    results: Sequence[Dict[str, Any]],
) -> str:
    name = _sheet_name("总览")
    ws = wb.create_sheet(name)
    systems = _systems(summary)
    sys_key = SYSTEM_HYPERGRAPH if SYSTEM_HYPERGRAPH in systems else systems[0]
    block = summary.get(sys_key) or {}
    j = block.get("judgment") or {}
    s = block.get("score") or {}
    p = block.get("pipeline") or {}
    lat = block.get("latency") or {}

    _write_title(ws, "超图评测总览", 4)
    overview = [
        ["题目数", _int(block.get("n_total")) or len(results), None, None],
        ["流水线成功", _int(p.get("n_ok")), "int", None],
        ["问答失败", _int(p.get("n_query_fail")), "int", False],
        ["评判失败", _int(p.get("n_judge_fail")), "int", False],
        ["已评判", _int(j.get("n_judged")), "int", True],
        ["正确", _int(j.get("n_correct")), "int", True],
        ["部分正确", _int(j.get("n_partial")), "int", None],
        ["错误", _int(j.get("n_wrong")), "int", False],
        ["正确率", _fmt_pct(j.get("accuracy")), "pct", True],
        ["宽松正确率（正确+部分）", _fmt_pct(j.get("lenient_accuracy")), "pct", True],
        ["错误率", _fmt_pct(j.get("error_rate")), "pct", False],
        ["均分 (0-3)", _num(s.get("mean")), "float2", True],
        ["归一化均分", _fmt_pct(s.get("mean_normalized")), "pct", True],
        ["均问答耗时 s", _num(lat.get("mean_query_s")), "float3", False],
        ["均检索耗时 s", _num(lat.get("mean_retrieve_s")), "float3", False],
        ["均评判耗时 s", _num(lat.get("mean_judge_s")), "float3", False],
    ]
    rows = [[a, b] for a, b, *_ in overview]
    _write_table(
        ws,
        ("指标", "超图"),
        rows,
        start_row=3,
        freeze=False,
        autofilter=False,
    )
    for i, (_name, _val, fmt, better) in enumerate(overview):
        cell = ws.cell(4 + i, 2)
        _apply_number_format(cell, fmt)
        cell.fill = _FILL_HG
        if fmt == "pct" and better is True and cell.value is not None:
            try:
                cell.fill = _FILL_POS if float(cell.value) >= 0.6 else (
                    _FILL_MID if float(cell.value) >= 0.3 else _FILL_NEG
                )
            except (TypeError, ValueError):
                pass

    hist_start = 4 + len(overview) + 2
    ws.cell(hist_start, 1, "分数分布").font = _FONT_BOLD
    hist = (s.get("histogram") or {}) if isinstance(s, dict) else {}
    hist_rows = [[int(sc), _int(hist.get(str(sc))) or 0] for sc in range(4)]
    _write_table(
        ws,
        ("分数", "题数"),
        hist_rows,
        start_row=hist_start + 1,
        col_formats=["int", "int"],
        freeze=False,
        autofilter=False,
        zebra=False,
    )
    chart = BarChart()
    chart.type = "col"
    chart.title = "分数分布"
    chart.y_axis.title = "题数"
    chart.x_axis.title = "分数"
    data_ref = Reference(
        ws, min_col=2, min_row=hist_start + 1, max_row=hist_start + 1 + len(hist_rows)
    )
    cats = Reference(
        ws, min_col=1, min_row=hist_start + 2, max_row=hist_start + 1 + len(hist_rows)
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.width = 12
    chart.height = 7
    ws.add_chart(chart, "D" + str(hist_start + 1))

    dim_start = hist_start + 8
    ws.cell(dim_start, 1, "得分维度通过率").font = _FONT_BOLD
    dim_stats = _dimension_stats(results, sys_key)
    dim_rows = []
    for key, rec in dim_stats.items():
        dim_rows.append(
            [key, rec["n"], rec["n_pass"], rec["n_fail"], rec["pass_rate"], rec["mean"]]
        )
    _write_table(
        ws,
        ("维度", "题数", "通过", "未通过", "通过率", "均分"),
        dim_rows,
        start_row=dim_start + 1,
        col_formats=[None, "int", "int", "int", "pct", "float2"],
        freeze=False,
        autofilter=False,
    )
    for i, row in enumerate(dim_rows):
        r = dim_start + 2 + i
        rate_cell = ws.cell(r, 5)
        fill = None
        if row[4] is not None:
            fill = _FILL_POS if row[4] >= 0.6 else (_FILL_MID if row[4] >= 0.3 else _FILL_NEG)
        if fill is not None:
            rate_cell.fill = fill
    return name


def _add_details(wb: Workbook, results: Sequence[Dict[str, Any]], systems: Sequence[str]) -> str:
    name = _sheet_name("逐题明细")
    ws = wb.create_sheet(name)
    has_lo = SYSTEM_LLM_ONLY in systems
    headers = [
        "ID",
        "来源专利",
        "主题",
        "问题",
        "预期回答",
        "超图评判",
        "超图分数",
        "核心机理",
        "配方可用性",
        "工艺与协同",
        "负向分",
        "超图原因",
        "超图回答",
        "检索来源数",
        "问答耗时s",
        "query状态",
        "错误",
        "来源段落",
        "md文件",
    ]
    fmts: List[Optional[str]] = [
        None, None, None, None, None,
        None, "int", "int", "int", "int", "int",
        None, None, "int", "float3", "int", None, None, None,
    ]
    if has_lo:
        headers += ["纯LLM评判", "纯LLM分数", "纯LLM回答", "纯LLM原因"]
        fmts += [None, "int", None, None]

    rows: List[List[Any]] = []
    for r in results:
        hg = _system_block(r, SYSTEM_HYPERGRAPH)
        dm = _dim_map(hg)
        hm = hg.get("metrics") or {}
        cats = r.get("categories") or {}
        row = [
            r.get("id"),
            r.get("source_patent") or cats.get("来源专利"),
            r.get("topic") or cats.get("主题"),
            r.get("question"),
            r.get("expected_answer"),
            hg.get("llm_acc"),
            _int(hg.get("score")),
            _int((dm.get("核心机理") or {}).get("score")),
            _int((dm.get("配方可用性") or {}).get("score")),
            _int((dm.get("工艺与协同") or {}).get("score")),
            _int((dm.get("负向分") or {}).get("score")),
            hg.get("judge_reason"),
            hg.get("answer") or hg.get("raw_answer"),
            len(hg.get("retrieval_sources") or []),
            _num(hm.get("query_latency_s")),
            _int(hg.get("query_status")),
            hg.get("query_error") or hg.get("judge_error"),
            r.get("source_ref") or r.get("knowledge_source"),
            r.get("md_file") or r.get("note"),
        ]
        if has_lo:
            lo = _system_block(r, SYSTEM_LLM_ONLY)
            row += [
                lo.get("llm_acc"),
                _int(lo.get("score")),
                lo.get("answer") or lo.get("raw_answer"),
                lo.get("judge_reason"),
            ]
        rows.append(row)

    _write_title(ws, "逐题明细（点选单元格可在编辑栏看全文）", len(headers))
    _write_table(
        ws,
        headers,
        rows,
        start_row=3,
        col_formats=fmts,
        freeze=True,
        autofilter=True,
        max_width=28,
        wrap=True,
    )
    ws.freeze_panes = "D4"
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["L"].width = 28
    ws.column_dimensions["M"].width = 36
    judge_col = headers.index("超图评判") + 1
    score_cols = [
        headers.index(h) + 1
        for h in ("超图分数", "核心机理", "配方可用性", "工艺与协同", "负向分")
        if h in headers
    ]
    for i in range(len(rows)):
        r = 4 + i
        fill = _judgment_fill(ws.cell(r, judge_col).value)
        if fill is not None:
            ws.cell(r, judge_col).fill = fill
        for c in score_cols:
            fill = _score_fill(ws.cell(r, c).value)
            if fill is not None:
                ws.cell(r, c).fill = fill
    return name


def _add_dimension_detail(
    wb: Workbook, results: Sequence[Dict[str, Any]], systems: Sequence[str]
) -> str:
    name = _sheet_name("维度评分")
    ws = wb.create_sheet(name)
    rows: List[List[Any]] = []
    for r in results:
        for system in systems:
            block = _system_block(r, system)
            for i, d in enumerate(block.get("dimension_scores") or [], 1):
                if not isinstance(d, dict):
                    continue
                raw_name = str(d.get("dimension") or "")
                rows.append(
                    [
                        r.get("id"),
                        r.get("source_patent") or (r.get("categories") or {}).get("来源专利"),
                        _sys_label(system),
                        i,
                        dimension_key(i, raw_name),
                        raw_name,
                        _int(d.get("score")),
                        d.get("pass"),
                        d.get("comment"),
                    ]
                )
    headers = ["ID", "来源专利", "系统", "序号", "维度类型", "评分维度", "分数", "通过", "评语"]
    _write_title(ws, "各评分维度明细", len(headers))
    _write_table(
        ws,
        headers,
        rows,
        start_row=3,
        col_formats=[None, None, None, "int", None, None, "int", None, None],
        freeze=True,
        autofilter=bool(rows),
        max_width=40,
        wrap=True,
    )
    score_col = 7
    for i in range(len(rows)):
        cell = ws.cell(4 + i, score_col)
        fill = _score_fill(cell.value)
        if fill is not None:
            cell.fill = fill
        pass_cell = ws.cell(4 + i, 8)
        if pass_cell.value is True:
            pass_cell.fill = _FILL_POS
        elif pass_cell.value is False:
            pass_cell.fill = _FILL_NEG
    return name


def _add_dimension_summary(
    wb: Workbook, results: Sequence[Dict[str, Any]], systems: Sequence[str]
) -> str:
    name = _sheet_name("维度汇总")
    ws = wb.create_sheet(name)
    rows: List[List[Any]] = []
    for system in systems:
        stats = _dimension_stats(results, system)
        for key, rec in stats.items():
            rows.append(
                [
                    _sys_label(system),
                    key,
                    rec["n"],
                    rec["n_pass"],
                    rec["n_fail"],
                    rec["pass_rate"],
                    rec["mean"],
                ]
            )
    headers = ["系统", "维度类型", "题数", "通过", "未通过", "通过率", "均分"]
    _write_title(ws, "得分维度汇总（核心机理 / 配方 / 工艺 / 负向）", len(headers))
    _write_table(
        ws,
        headers,
        rows,
        start_row=3,
        col_formats=[None, None, "int", "int", "int", "pct", "float2"],
        freeze=True,
        autofilter=bool(rows),
    )
    for i, row in enumerate(rows):
        cell = ws.cell(4 + i, 6)
        if row[5] is None:
            continue
        cell.fill = _FILL_POS if row[5] >= 0.6 else (_FILL_MID if row[5] >= 0.3 else _FILL_NEG)
    return name


def _add_failures(
    wb: Workbook, results: Sequence[Dict[str, Any]], systems: Sequence[str]
) -> str:
    name = _sheet_name("失败列表")
    ws = wb.create_sheet(name)
    rows: List[List[Any]] = []
    for r in results:
        for system in systems:
            b = _system_block(r, system)
            qerr = b.get("query_error")
            jerr = b.get("judge_error")
            qs, js = b.get("query_status"), b.get("judge_status")
            if qerr or jerr or qs == 0 or js == 0:
                rows.append(
                    [
                        r.get("id"),
                        _sys_label(system),
                        r.get("source_patent"),
                        r.get("question"),
                        qs,
                        qerr,
                        js,
                        jerr,
                    ]
                )
    headers = ["ID", "系统", "来源专利", "问题", "query状态", "query错误", "judge状态", "judge错误"]
    _write_title(
        ws,
        f"问答 / 评判失败（{len(rows)} 条）" if rows else "问答 / 评判失败（无）",
        len(headers),
    )
    _write_table(
        ws,
        headers,
        rows,
        start_row=3,
        freeze=True,
        autofilter=bool(rows),
        max_width=40,
    )
    return name


def write_report_workbook(
    report: Dict[str, Any],
    out_path: Union[str, Path],
    *,
    eval_data: Optional[Dict[str, Any]] = None,
    title: str = "建筑涂料专利集超图评测",
) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    summary = report.get("summary") or {}
    if eval_data is None and isinstance(report.get("results"), list):
        eval_data = report
    results = list((eval_data or {}).get("results") or [])
    systems = _systems(summary)

    wb = Workbook()
    default = wb.active
    default.title = "tmp"
    toc: List[Tuple[str, str]] = []

    toc.append((_add_overview(wb, summary, results), "正确率、均分、分数分布、四维通过率"))
    if results:
        toc.append((_add_details(wb, results, systems), "每题超图回答、评判、四维分数"))
        toc.append((_add_dimension_detail(wb, results, systems), "评分维度逐条展开"))
        toc.append((_add_dimension_summary(wb, results, systems), "四类得分维度通过率"))
        toc.append((_add_failures(wb, results, systems), "问答或评判失败清单"))
    toc.append((_add_meta(wb, report, eval_data), "评测配置、模型、题量、耗时"))

    _add_toc(wb, toc, title=title)
    if "tmp" in wb.sheetnames:
        del wb["tmp"]
    wb.save(str(out))
    return out


def main(argv: Optional[List[str]] = None) -> int:
    import argparse
    import sys

    root = Path(__file__).resolve().parent.parent
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))

    from benchmark3.config import DEFAULT_CONFIG_PATH, Benchmark3Config

    p = argparse.ArgumentParser(description="benchmark3 report/evals JSON → 总结 Excel")
    p.add_argument("--config", default=DEFAULT_CONFIG_PATH)
    p.add_argument("--report", default=None)
    p.add_argument("--evals", default=None)
    p.add_argument("--out", default=None)
    args = p.parse_args(argv)

    cfg = Benchmark3Config.from_yaml(args.config).resolve_paths()
    report_p = Path(args.report) if args.report else cfg.report_file()
    evals_p = Path(args.evals) if args.evals else cfg.eval_results_file()
    out_p = Path(args.out) if args.out else cfg.report_excel_file()

    report = None
    evals = None
    if evals_p.is_file():
        with open(evals_p, "r", encoding="utf-8") as f:
            evals = json.load(f)
    if report_p.is_file():
        with open(report_p, "r", encoding="utf-8") as f:
            report = json.load(f)
    if report is None:
        if evals is None:
            raise FileNotFoundError(f"未找到 report/evals: {report_p} / {evals_p}")
        report = build_report_document(evals, source_path=str(evals_p))
    write_report_workbook(report, out_p, eval_data=evals)
    print(f"[saved] {out_p}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
