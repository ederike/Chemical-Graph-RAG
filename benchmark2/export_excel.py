#!/usr/bin/env python3
"""
把 benchmark2 的 report / evals JSON 汇总成多表 Excel。

默认读 config.yaml 的 report / eval 路径，写出同名 .xlsx::

    python -m benchmark2.export_excel
    python -m benchmark2.export_excel \\
        --report benchmark2_results/report_queation1.json \\
        --evals  benchmark2_results/evals_queation1.json \\
        --out    benchmark2_results/report_queation1.xlsx

也可用 run.mode=excel，或在 report / all 结束后自动写出。
"""

from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

from .evaluator import SYSTEM_HYPERGRAPH, SYSTEM_LLM_ONLY
from .report import RETRIEVE_STAGE_KEYS, build_report_document, project_system_row

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:  # pragma: no cover
    raise ImportError("写出 Excel 需要 openpyxl，请 pip install openpyxl") from e


SYSTEM_LABELS = {
    SYSTEM_HYPERGRAPH: "超图",
    SYSTEM_LLM_ONLY: "纯LLM",
}
JUDGE_ORDER = ("正确", "部分正确", "错误", "未知")
EXCEL_CELL_MAX = 32767
_INVALID_SHEET_RE = re.compile(r"[\[\]\*?:/\\]")

# 样式
_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FILL_HG = PatternFill("solid", fgColor="D6EAF8")
_FILL_LO = PatternFill("solid", fgColor="FDEBD0")
_FILL_POS = PatternFill("solid", fgColor="D5F5E3")
_FILL_NEG = PatternFill("solid", fgColor="FADBD8")
_FILL_TIE = PatternFill("solid", fgColor="F4F6F7")
_FILL_ALT = PatternFill("solid", fgColor="F7F9FB")
_FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
_FONT_HEADER = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_FONT_TITLE = Font(name="微软雅黑", bold=True, color="FFFFFF", size=14)
_FONT_CELL = Font(name="微软雅黑", size=10)
_FONT_BOLD = Font(name="微软雅黑", bold=True, size=10)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_R = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# 小工具
# ---------------------------------------------------------------------------


def _sheet_name(name: str, used: Optional[set] = None) -> str:
    s = _INVALID_SHEET_RE.sub(" ", str(name or "Sheet")).strip() or "Sheet"
    s = s[:31]
    if used is None:
        return s
    base, n = s, 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = (base[: 31 - len(suffix)] + suffix)
    used.add(s)
    return s


def _cell_text(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v)
    if len(s) > EXCEL_CELL_MAX:
        s = s[: EXCEL_CELL_MAX - 1] + "…"
    return s


def _fmt_pct(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _delta_fill(delta: Optional[float], *, higher_is_better: bool) -> Optional[PatternFill]:
    if delta is None:
        return None
    try:
        d = float(delta)
    except (TypeError, ValueError):
        return None
    if d == 0:
        return _FILL_TIE
    better = d > 0 if higher_is_better else d < 0
    return _FILL_POS if better else _FILL_NEG


def _systems(summary: Dict[str, Any]) -> List[str]:
    raw = summary.get("systems")
    out: List[str] = []
    if isinstance(raw, list):
        for k in raw:
            if k in SYSTEM_LABELS:
                out.append(k)
    if not out:
        for k in (SYSTEM_HYPERGRAPH, SYSTEM_LLM_ONLY):
            block = summary.get(k)
            if isinstance(block, dict) and block.get("judgment") is not None:
                out.append(k)
    return out or [SYSTEM_HYPERGRAPH]


def _sys_label(key: str) -> str:
    return SYSTEM_LABELS.get(key, key)


def _nested(d: Optional[dict], *keys: str, default: Any = None) -> Any:
    cur: Any = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur


def _iter_category_groups(
    sys_summary: Dict[str, Any],
) -> Iterable[Tuple[str, str, str, Dict[str, Any]]]:
    for kind, key in (("主分类", "by_category"), ("附加分类", "by_category_extra")):
        block = sys_summary.get(key) or {}
        if not isinstance(block, dict):
            continue
        for field, fb in block.items():
            groups = (fb or {}).get("groups") or {}
            for label, g in groups.items():
                yield kind, str(field), str(label), g or {}


def _auto_width(ws: Worksheet, *, min_w: float = 8, max_w: float = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            if cell.value is None:
                continue
            text = str(cell.value).replace("\n", " ")
            # 中文按 1.7 宽估算
            w = 0.0
            for ch in text[:80]:
                w += 1.7 if ord(ch) > 127 else 1.05
            longest = max(longest, w)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 1.5))


def _apply_number_format(cell, fmt: Optional[str]) -> None:
    if fmt == "pct":
        cell.number_format = "0.00%"
    elif fmt == "int":
        cell.number_format = "#,##0"
    elif fmt == "float2":
        cell.number_format = "0.00"
    elif fmt == "float3":
        cell.number_format = "0.000"
    elif fmt == "float4":
        cell.number_format = "0.0000"


def _write_title(ws: Worksheet, title: str, n_cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, n_cols))
    cell = ws.cell(1, 1, title)
    cell.fill = _FILL_TITLE
    cell.font = _FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def _write_table(
    ws: Worksheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    start_row: int = 1,
    col_formats: Optional[Sequence[Optional[str]]] = None,
    header_fills: Optional[Sequence[Optional[PatternFill]]] = None,
    freeze: bool = True,
    autofilter: bool = True,
    zebra: bool = True,
    max_width: float = 48,
) -> int:
    """写一张表，返回下一空行号。"""
    n_cols = len(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = _FONT_HEADER
        cell.fill = (
            header_fills[c - 1]
            if header_fills and header_fills[c - 1] is not None
            else _FILL_HEADER
        )
        cell.alignment = _ALIGN_C
        cell.border = _THIN
    ws.row_dimensions[start_row].height = 22

    for i, row in enumerate(rows):
        r = start_row + 1 + i
        alt = zebra and (i % 2 == 1)
        for c in range(1, n_cols + 1):
            raw = row[c - 1] if c - 1 < len(row) else None
            val = _cell_text(raw)
            cell = ws.cell(r, c, val)
            cell.font = _FONT_CELL
            cell.border = _THIN
            fmt = col_formats[c - 1] if col_formats and c - 1 < len(col_formats) else None
            _apply_number_format(cell, fmt)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = _ALIGN_R
            else:
                cell.alignment = _ALIGN_L
            if alt:
                cell.fill = _FILL_ALT

    last = start_row + len(rows)
    if autofilter and rows:
        ws.auto_filter.ref = (
            f"{get_column_letter(1)}{start_row}:{get_column_letter(n_cols)}{last}"
        )
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"
    _auto_width(ws, max_w=max_width)
    ws.sheet_view.showGridLines = False
    return last + 1


# ---------------------------------------------------------------------------
# 各 sheet
# ---------------------------------------------------------------------------


def _add_toc(wb: Workbook, items: Sequence[Tuple[str, str]]) -> None:
    ws = wb.create_sheet(_sheet_name("目录"), 0)
    _write_title(ws, "Benchmark2 评测报告", 2)
    ws.cell(2, 1, "点击工作表名跳转。JSON 嵌套已摊平成表格。")
    ws.cell(2, 1).font = Font(name="微软雅黑", italic=True, size=10, color="666666")
    headers = ("工作表", "说明")
    rows = [(name, desc) for name, desc in items]
    _write_table(ws, headers, rows, start_row=4, freeze=False, autofilter=False, zebra=True)
    for i, (sheet, _) in enumerate(items):
        cell = ws.cell(5 + i, 1, sheet)
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet}'!A1", display=sheet)
        cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        cell.alignment = _ALIGN_L
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 72


def _add_meta(wb: Workbook, report: Dict[str, Any], eval_data: Optional[Dict[str, Any]]) -> str:
    name = _sheet_name("运行信息")
    ws = wb.create_sheet(name)
    meta = dict(report.get("meta") or {})
    ev_meta = (eval_data or {}).get("meta") if isinstance(eval_data, dict) else {}
    if isinstance(ev_meta, dict):
        for k, v in ev_meta.items():
            meta.setdefault(k, v)
    skip = {"config", "dataset_meta", "progress", "stats_meta"}
    rows: List[List[Any]] = []
    preferred = [
        "created_at",
        "updated_at",
        "eval_done",
        "done",
        "query_mode",
        "judge_model",
        "answer_model",
        "enable_llm_only",
        "n_results",
        "n_questions",
        "n_total_planned",
        "n_resumed",
        "eval_elapsed_s",
        "elapsed_s",
        "excel_path",
        "source_path",
        "config_file",
        "dhmf_config_path",
    ]
    seen = set()
    for k in preferred:
        if k in meta:
            rows.append([k, _flatten_meta_value(meta[k])])
            seen.add(k)
    for k in sorted(meta.keys()):
        if k in seen or k in skip:
            continue
        rows.append([k, _flatten_meta_value(meta[k])])
    cfg = meta.get("config") if isinstance(meta.get("config"), dict) else None
    if cfg:
        rows.append(["— 配置快照 —", ""])
        for k, v in cfg.items():
            rows.append([f"config.{k}", _flatten_meta_value(v)])
    _write_title(ws, "运行信息", 2)
    _write_table(
        ws,
        ("字段", "值"),
        rows,
        start_row=3,
        freeze=True,
        autofilter=False,
        max_width=80,
    )
    return name


def _flatten_meta_value(v: Any) -> Any:
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    return json.dumps(v, ensure_ascii=False)


def _metric_row_overall(
    sys_blocks: Dict[str, Dict[str, Any]], systems: Sequence[str]
) -> Tuple[List[List[Any]], List[Tuple[str, str, Tuple[str, ...], Optional[bool]]], bool]:
    """摊平总体指标为 指标 | 系统… | 差值 行。"""

    def g(sys: str, *path, default=None):
        return _nested(sys_blocks.get(sys), *path, default=default)

    specs: List[Tuple[str, str, Tuple[str, ...], Optional[bool]]] = [
        ("题目数 n", "int", ("n_total",), None),
        ("流水线成功", "int", ("pipeline", "n_ok"), True),
        ("问答失败", "int", ("pipeline", "n_query_fail"), False),
        ("评判失败", "int", ("pipeline", "n_judge_fail"), False),
        ("已评判", "int", ("judgment", "n_judged"), True),
        ("正确", "int", ("judgment", "n_correct"), True),
        ("部分正确", "int", ("judgment", "n_partial"), None),
        ("错误", "int", ("judgment", "n_wrong"), False),
        ("未知", "int", ("judgment", "counts", "未知"), None),
        ("正确率", "pct", ("judgment", "accuracy"), True),
        ("宽松正确率（正确+部分）", "pct", ("judgment", "lenient_accuracy"), True),
        ("错误率", "pct", ("judgment", "error_rate"), False),
        ("部分正确率", "pct", ("judgment", "partial_rate"), None),
        ("均分 (0-3)", "float2", ("score", "mean"), True),
        ("归一化均分", "pct", ("score", "mean_normalized"), True),
        ("已打分", "int", ("score", "n_scored"), True),
        ("分数=0", "int", ("score", "histogram", "0"), False),
        ("分数=1", "int", ("score", "histogram", "1"), None),
        ("分数=2", "int", ("score", "histogram", "2"), True),
        ("分数=3", "int", ("score", "histogram", "3"), True),
        ("均问答耗时 s", "float3", ("latency", "mean_query_s"), False),
        ("均检索耗时 s", "float3", ("latency", "mean_retrieve_s"), False),
        ("均墙钟耗时 s", "float3", ("latency", "mean_wall_s"), False),
        ("均评判耗时 s", "float3", ("latency", "mean_judge_s"), False),
        ("问答总耗时 s", "float3", ("latency", "sum_query_s"), False),
        ("检索总耗时 s", "float3", ("latency", "sum_retrieve_s"), False),
        ("墙钟总耗时 s", "float3", ("latency", "sum_wall_s"), False),
        ("prompt tokens 合计", "int", ("tokens", "sum_prompt"), None),
        ("completion tokens 合计", "int", ("tokens", "sum_completion"), None),
        ("tokens 合计", "int", ("tokens", "sum_total"), None),
        ("均 prompt tokens", "float2", ("tokens", "mean_prompt"), None),
        ("均 completion tokens", "float2", ("tokens", "mean_completion"), None),
        ("评判 prompt tokens 合计", "int", ("tokens", "sum_judge_prompt"), None),
        ("评判 completion tokens 合计", "int", ("tokens", "sum_judge_completion"), None),
        ("评判 tokens 合计", "int", ("tokens", "sum_judge_total"), None),
    ]

    has_pair = len(systems) >= 2
    rows: List[List[Any]] = []
    for name, fmt, path, better in specs:
        vals = [g(s, *path) for s in systems]
        if fmt == "pct":
            vals = [_fmt_pct(v) for v in vals]
        elif fmt == "int":
            vals = [_int(v) if v is not None else None for v in vals]
        else:
            vals = [_num(v) if v is not None else None for v in vals]
        delta = None
        if has_pair and vals[0] is not None and vals[1] is not None:
            try:
                delta = round(float(vals[0]) - float(vals[1]), 6)
            except (TypeError, ValueError):
                delta = None
        row: List[Any] = [name, *vals]
        if has_pair:
            row.append(delta)
        rows.append(row)
    return rows, specs, has_pair


def _add_overall(wb: Workbook, summary: Dict[str, Any]) -> str:
    sheet_name = _sheet_name("总体指标")
    ws = wb.create_sheet(sheet_name)
    systems = _systems(summary)
    sys_blocks = {s: summary.get(s) or {} for s in systems}
    rows, specs, has_pair = _metric_row_overall(sys_blocks, systems)

    headers = ["指标"] + [_sys_label(s) for s in systems]
    col_formats: List[Optional[str]] = [None] + [None] * len(systems)
    if has_pair:
        headers.append("差值（超图 − 纯LLM）")
        col_formats.append("float4")

    n_cols = len(headers)
    _write_title(ws, "总体指标：超图 vs 纯LLM" if has_pair else "总体指标", n_cols)
    # 表头
    start = 3
    _write_table(
        ws,
        headers,
        rows,
        start_row=start,
        col_formats=col_formats,
        freeze=True,
        autofilter=False,
        zebra=True,
    )
    # 按行套数字格式 + 差值着色 + 系统列底色
    for i, (_metric, fmt, path, better) in enumerate(specs):
        r = start + 1 + i
        for c, s in enumerate(systems, 2):
            cell = ws.cell(r, c)
            _apply_number_format(cell, fmt)
            cell.fill = _FILL_HG if s == SYSTEM_HYPERGRAPH else (
                _FILL_LO if s == SYSTEM_LLM_ONLY else _FILL_ALT
            )
        if has_pair:
            dcell = ws.cell(r, 2 + len(systems))
            _apply_number_format(dcell, fmt)
            if better is not None:
                fill = _delta_fill(dcell.value, higher_is_better=bool(better))
                if fill is not None:
                    dcell.fill = fill

    # 分数直方图小表 + 柱状图
    hist_start = start + len(rows) + 3
    ws.cell(hist_start, 1, "分数直方图").font = _FONT_BOLD
    hist_headers = ["分数"] + [_sys_label(s) for s in systems]
    hist_rows = []
    for sc in ("0", "1", "2", "3"):
        hist_rows.append(
            [int(sc)]
            + [
                _int(_nested(sys_blocks.get(s), "score", "histogram", sc, default=0)) or 0
                for s in systems
            ]
        )
    _write_table(
        ws,
        hist_headers,
        hist_rows,
        start_row=hist_start + 1,
        col_formats=["int"] + ["int"] * len(systems),
        freeze=False,
        autofilter=False,
        zebra=False,
    )
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "分数分布"
    chart.y_axis.title = "题数"
    chart.x_axis.title = "分数"
    data_ref = Reference(
        ws,
        min_col=2,
        min_row=hist_start + 1,
        max_col=1 + len(systems),
        max_row=hist_start + 1 + len(hist_rows),
    )
    cats = Reference(
        ws, min_col=1, min_row=hist_start + 2, max_row=hist_start + 1 + len(hist_rows)
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.y_axis.scaling.min = 0
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "F" + str(hist_start))
    return sheet_name


def _parse_crosstab(xt: Dict[str, Any]) -> Dict[Tuple[str, str], int]:
    matrix: Dict[Tuple[str, str], int] = {}
    for k, v in (xt or {}).items():
        key = str(k)
        if not key.startswith("超图") or "_纯LLM" not in key:
            continue
        rest = key[len("超图") :]
        hj, lj = rest.split("_纯LLM", 1)
        try:
            matrix[(hj, lj)] = int(v)
        except (TypeError, ValueError):
            continue
    return matrix


def _add_comparison(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    cmp_ = summary.get("comparison")
    if not isinstance(cmp_, dict) or not cmp_:
        return None
    name = _sheet_name("系统对比")
    ws = wb.create_sheet(name)
    _write_title(ws, "超图 vs 纯LLM 对比", 5)

    win = cmp_.get("win") or {}
    n_paired = _int(cmp_.get("n_paired")) or 0
    n_hg = _int(win.get("hypergraph")) or 0
    n_lo = _int(win.get("llm_only")) or 0
    n_tie = _int(win.get("tie")) or 0

    overview = [
        ["配对题数", n_paired],
        ["均分差（超图 − 纯LLM）", _num(cmp_.get("score_delta_mean"))],
        ["说明", cmp_.get("score_delta_note") or "正值表示超图均分更高"],
        ["超图胜", n_hg],
        ["纯LLM胜", n_lo],
        ["平局", n_tie],
        ["超图胜率", (n_hg / n_paired) if n_paired else None],
        ["纯LLM胜率", (n_lo / n_paired) if n_paired else None],
        ["平局率", (n_tie / n_paired) if n_paired else None],
    ]
    _write_table(
        ws,
        ("项目", "值"),
        overview,
        start_row=3,
        freeze=False,
        autofilter=False,
    )
    for r, row in enumerate(overview, 4):
        if "率" in str(row[0]):
            _apply_number_format(ws.cell(r, 2), "pct")
        elif "均分差" in str(row[0]):
            _apply_number_format(ws.cell(r, 2), "float2")
            fill = _delta_fill(ws.cell(r, 2).value, higher_is_better=True)
            if fill is not None:
                ws.cell(r, 2).fill = fill
        elif row[0] == "超图胜":
            ws.cell(r, 2).fill = _FILL_POS
        elif row[0] == "纯LLM胜":
            ws.cell(r, 2).fill = _FILL_NEG
        elif row[0] == "平局":
            ws.cell(r, 2).fill = _FILL_TIE

    # 胜负柱状图数据
    bar_row = 15
    ws.cell(bar_row, 1, "胜负").font = _FONT_BOLD
    _write_table(
        ws,
        ("结果", "题数"),
        [["超图胜", n_hg], ["纯LLM胜", n_lo], ["平局", n_tie]],
        start_row=bar_row + 1,
        col_formats=[None, "int"],
        freeze=False,
        autofilter=False,
        zebra=False,
    )
    chart = BarChart()
    chart.type = "col"
    chart.title = "分数胜负"
    chart.y_axis.title = "题数"
    data_ref = Reference(ws, min_col=2, min_row=bar_row + 1, max_row=bar_row + 4)
    cats = Reference(ws, min_col=1, min_row=bar_row + 2, max_row=bar_row + 4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.width = 12
    chart.height = 7
    ws.add_chart(chart, "D16")

    # 评判交叉表
    xt_start = 22
    ws.cell(xt_start, 1, "评判交叉表（行=超图，列=纯LLM）").font = _FONT_BOLD
    matrix = _parse_crosstab(cmp_.get("judgment_crosstab") or {})
    labels = list(JUDGE_ORDER)
    xt_headers = ["超图 \\ 纯LLM"] + [f"纯LLM·{x}" for x in labels] + ["行合计"]
    xt_rows = []
    col_tot = {lj: 0 for lj in labels}
    for hj in labels:
        row = [f"超图·{hj}"]
        s = 0
        for lj in labels:
            v = matrix.get((hj, lj), 0)
            row.append(v)
            s += v
            col_tot[lj] += v
        row.append(s)
        xt_rows.append(row)
    xt_rows.append(["列合计"] + [col_tot[lj] for lj in labels] + [sum(col_tot.values())])
    _write_table(
        ws,
        xt_headers,
        xt_rows,
        start_row=xt_start + 1,
        col_formats=[None] + ["int"] * (len(labels) + 1),
        freeze=False,
        autofilter=False,
        zebra=False,
    )
    return name


def _group_metric_row(kind: str, field: str, label: str, g: Dict[str, Any], system: str) -> List[Any]:
    j = g.get("judgment") or {}
    s = g.get("score") or {}
    lat = g.get("latency") or {}
    p = g.get("pipeline") or {}
    hist = s.get("histogram") or {}
    counts = j.get("counts") or {}
    return [
        _sys_label(system),
        kind,
        field,
        label,
        _int(g.get("n")),
        _fmt_pct(g.get("eval_ratio")),
        _int(g.get("dataset_count")),
        _fmt_pct(g.get("dataset_ratio")),
        _fmt_pct(g.get("coverage")),
        _int(j.get("n_correct") if j.get("n_correct") is not None else counts.get("正确")),
        _int(j.get("n_partial") if j.get("n_partial") is not None else counts.get("部分正确")),
        _int(j.get("n_wrong") if j.get("n_wrong") is not None else counts.get("错误")),
        _int(counts.get("未知")),
        _int(j.get("n_judged")),
        _fmt_pct(j.get("accuracy")),
        _fmt_pct(j.get("lenient_accuracy")),
        _fmt_pct(j.get("error_rate")),
        _fmt_pct(j.get("partial_rate")),
        _num(s.get("mean")),
        _fmt_pct(s.get("mean_normalized")),
        _int(hist.get("0")),
        _int(hist.get("1")),
        _int(hist.get("2")),
        _int(hist.get("3")),
        _int(s.get("n_scored")),
        _num(lat.get("mean_query_s")),
        _num(lat.get("mean_retrieve_s")),
        _num(lat.get("mean_wall_s")),
        _int(p.get("n_query_fail")),
        _int(p.get("n_judge_fail")),
        _int(p.get("n_ok")),
    ]


_CAT_HEADERS = [
    "系统",
    "分类类型",
    "分类字段",
    "分组",
    "n",
    "评测占比",
    "全集数量",
    "全集占比",
    "覆盖率",
    "正确",
    "部分正确",
    "错误",
    "未知",
    "已评判",
    "正确率",
    "宽松正确率",
    "错误率",
    "部分正确率",
    "均分",
    "归一化均分",
    "分0",
    "分1",
    "分2",
    "分3",
    "已打分",
    "均问答s",
    "均检索s",
    "均墙钟s",
    "问答失败",
    "评判失败",
    "流水线成功",
]
_CAT_FMTS = [
    None,
    None,
    None,
    None,
    "int",
    "pct",
    "int",
    "pct",
    "pct",
    "int",
    "int",
    "int",
    "int",
    "int",
    "pct",
    "pct",
    "pct",
    "pct",
    "float2",
    "pct",
    "int",
    "int",
    "int",
    "int",
    "int",
    "float3",
    "float3",
    "float3",
    "int",
    "int",
    "int",
]


def _add_category_long(wb: Workbook, summary: Dict[str, Any]) -> str:
    name = _sheet_name("分类汇总")
    ws = wb.create_sheet(name)
    rows: List[List[Any]] = []
    for system in _systems(summary):
        block = summary.get(system) or {}
        for kind, field, label, g in _iter_category_groups(block):
            rows.append(_group_metric_row(kind, field, label, g, system))
    _write_title(ws, "按分类切分（长表，可用筛选）", len(_CAT_HEADERS))
    _write_table(
        ws,
        _CAT_HEADERS,
        rows,
        start_row=3,
        col_formats=_CAT_FMTS,
        freeze=True,
        autofilter=True,
        max_width=18,
    )
    return name


def _add_category_compare(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    systems = _systems(summary)
    if SYSTEM_HYPERGRAPH not in systems or SYSTEM_LLM_ONLY not in systems:
        return None
    name = _sheet_name("分类对比")
    ws = wb.create_sheet(name)

    hg_map: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    lo_map: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for kind, field, label, g in _iter_category_groups(summary.get(SYSTEM_HYPERGRAPH) or {}):
        hg_map[(kind, field, label)] = g
    for kind, field, label, g in _iter_category_groups(summary.get(SYSTEM_LLM_ONLY) or {}):
        lo_map[(kind, field, label)] = g
    keys = list(OrderedDict.fromkeys(list(hg_map.keys()) + list(lo_map.keys())))

    headers = [
        "分类类型",
        "分类字段",
        "分组",
        "n",
        "正确率_超图",
        "正确率_纯LLM",
        "正确率差",
        "宽松_超图",
        "宽松_纯LLM",
        "宽松差",
        "均分_超图",
        "均分_纯LLM",
        "均分差",
        "错误率_超图",
        "错误率_纯LLM",
        "错误率差",
        "正确_超图",
        "正确_纯LLM",
        "部分_超图",
        "部分_纯LLM",
        "错误_超图",
        "错误_纯LLM",
        "均问答s_超图",
        "均问答s_纯LLM",
    ]
    fmts = [
        None,
        None,
        None,
        "int",
        "pct",
        "pct",
        "pct",
        "pct",
        "pct",
        "pct",
        "float2",
        "float2",
        "float2",
        "pct",
        "pct",
        "pct",
        "int",
        "int",
        "int",
        "int",
        "int",
        "int",
        "float3",
        "float3",
    ]
    rows = []
    for kind, field, label in keys:
        hg, lo = hg_map.get((kind, field, label), {}), lo_map.get((kind, field, label), {})
        hj, lj = hg.get("judgment") or {}, lo.get("judgment") or {}
        hs, ls = hg.get("score") or {}, lo.get("score") or {}
        hlat, llat = hg.get("latency") or {}, lo.get("latency") or {}
        acc_h, acc_l = _fmt_pct(hj.get("accuracy")), _fmt_pct(lj.get("accuracy"))
        len_h, len_l = _fmt_pct(hj.get("lenient_accuracy")), _fmt_pct(lj.get("lenient_accuracy"))
        mean_h, mean_l = _num(hs.get("mean")), _num(ls.get("mean"))
        err_h, err_l = _fmt_pct(hj.get("error_rate")), _fmt_pct(lj.get("error_rate"))

        def sub(a, b):
            if a is None or b is None:
                return None
            return round(float(a) - float(b), 6)

        rows.append(
            [
                kind,
                field,
                label,
                _int(hg.get("n")) or _int(lo.get("n")),
                acc_h,
                acc_l,
                sub(acc_h, acc_l),
                len_h,
                len_l,
                sub(len_h, len_l),
                mean_h,
                mean_l,
                sub(mean_h, mean_l),
                err_h,
                err_l,
                sub(err_h, err_l),
                _int(hj.get("n_correct")),
                _int(lj.get("n_correct")),
                _int(hj.get("n_partial")),
                _int(lj.get("n_partial")),
                _int(hj.get("n_wrong")),
                _int(lj.get("n_wrong")),
                _num(hlat.get("mean_query_s")),
                _num(llat.get("mean_query_s")),
            ]
        )

    _write_title(ws, "同一分类下超图 vs 纯LLM（差 = 超图 − 纯LLM）", len(headers))
    start = 3
    _write_table(
        ws,
        headers,
        rows,
        start_row=start,
        col_formats=fmts,
        freeze=True,
        autofilter=True,
        max_width=16,
        zebra=True,
    )
    # 差值列着色
    higher_better_idx = {7, 10, 13}  # 正确率差, 宽松差, 均分差  (1-based)
    lower_better_idx = {16}  # 错误率差
    for i in range(len(rows)):
        r = start + 1 + i
        for c in higher_better_idx:
            fill = _delta_fill(ws.cell(r, c).value, higher_is_better=True)
            if fill is not None:
                ws.cell(r, c).fill = fill
        for c in lower_better_idx:
            fill = _delta_fill(ws.cell(r, c).value, higher_is_better=False)
            if fill is not None:
                ws.cell(r, c).fill = fill
        ws.cell(r, 5).fill = _FILL_HG
        ws.cell(r, 6).fill = _FILL_LO
        ws.cell(r, 8).fill = _FILL_HG
        ws.cell(r, 9).fill = _FILL_LO
        ws.cell(r, 11).fill = _FILL_HG
        ws.cell(r, 12).fill = _FILL_LO
    return name


def _add_latency_tokens(wb: Workbook, summary: Dict[str, Any]) -> str:
    name = _sheet_name("延迟与Token")
    ws = wb.create_sheet(name)
    _write_title(ws, "检索阶段延迟 + Token", 6)
    systems = _systems(summary)

    lat_headers = ["系统", "阶段", "均值 s", "合计 s"]
    lat_rows: List[List[Any]] = []
    stages = [
        ("问答 query", "mean_query_s", "sum_query_s"),
        ("检索 retrieve", "mean_retrieve_s", "sum_retrieve_s"),
        ("墙钟 wall", "mean_wall_s", "sum_wall_s"),
        ("评判 judge", "mean_judge_s", "sum_judge_s"),
    ]
    for src, mk, sk in RETRIEVE_STAGE_KEYS:
        label = src.replace("_latency_s", "")
        stages.append((f"检索·{label}", mk, sk))
    for system in systems:
        lat = (summary.get(system) or {}).get("latency") or {}
        for label, mk, sk in stages:
            lat_rows.append(
                [_sys_label(system), label, _num(lat.get(mk)), _num(lat.get(sk))]
            )
    _write_table(
        ws,
        lat_headers,
        lat_rows,
        start_row=3,
        col_formats=[None, None, "float3", "float3"],
        freeze=False,
        autofilter=True,
        zebra=True,
    )

    tok_start = 5 + len(lat_rows)
    ws.cell(tok_start, 1, "Token").font = _FONT_BOLD
    tok_headers = ["系统", "prompt合计", "completion合计", "合计", "均prompt", "均completion", "评判prompt合计", "评判completion合计", "评判合计"]
    tok_rows = []
    for system in systems:
        t = (summary.get(system) or {}).get("tokens") or {}
        tok_rows.append(
            [
                _sys_label(system),
                _int(t.get("sum_prompt")),
                _int(t.get("sum_completion")),
                _int(t.get("sum_total")),
                _num(t.get("mean_prompt")),
                _num(t.get("mean_completion")),
                _int(t.get("sum_judge_prompt")),
                _int(t.get("sum_judge_completion")),
                _int(t.get("sum_judge_total")),
            ]
        )
    _write_table(
        ws,
        tok_headers,
        tok_rows,
        start_row=tok_start + 1,
        col_formats=[None, "int", "int", "int", "float2", "float2", "int", "int", "int"],
        freeze=False,
        autofilter=False,
        zebra=False,
        max_width=18,
    )
    return name


def _result_system_block(result: Dict[str, Any], system: str) -> Dict[str, Any]:
    block = result.get(system)
    if isinstance(block, dict) and block:
        return block
    row = project_system_row(result, system)
    return {
        "answer": result.get("rag_answer") if system == SYSTEM_HYPERGRAPH else result.get("llm_only_answer"),
        "raw_answer": result.get("rag_raw_answer") if system == SYSTEM_HYPERGRAPH else None,
        **{k: row.get(k) for k in (
            "llm_acc",
            "score",
            "query_status",
            "query_error",
            "judge_status",
            "judge_error",
            "metrics",
        )},
        "judge_reason": result.get("judge_reason") if system == SYSTEM_HYPERGRAPH else (block or {}).get("judge_reason"),
        "dimension_scores": row.get("dimension_scores") or [],
        "retrieval_sources": result.get("retrieval_sources") if system == SYSTEM_HYPERGRAPH else [],
    }


def _winner(hs: Optional[float], ls: Optional[float]) -> str:
    if hs is None or ls is None:
        return ""
    if hs > ls:
        return "超图"
    if ls > hs:
        return "纯LLM"
    return "平"


def _category_fields_from_results(results: Sequence[Dict[str, Any]]) -> List[str]:
    seen: List[str] = []
    bag = set()
    for r in results:
        cats = r.get("categories") or {}
        for k in cats.keys():
            if k not in bag:
                seen.append(str(k))
                bag.add(k)
    return seen


def _detail_headers(cat_fields: Sequence[str], has_lo: bool) -> Tuple[List[str], List[Optional[str]]]:
    headers = ["ID", "问题", "标准答案"]
    fmts: List[Optional[str]] = [None, None, None]
    for f in cat_fields:
        headers.append(f)
        fmts.append(None)
    headers += [
        "超图评判",
        "超图分数",
        "超图query状态",
        "超图问答s",
        "超图检索s",
        "超图tokens",
        "超图检索来源数",
        "超图评判原因",
        "超图答案",
        "超图错误",
    ]
    fmts += [None, "int", "int", "float3", "float3", "int", "int", None, None, None]
    if has_lo:
        headers += [
            "纯LLM评判",
            "纯LLM分数",
            "纯LLM query状态",
            "纯LLM问答s",
            "纯LLM tokens",
            "纯LLM评判原因",
            "纯LLM答案",
            "纯LLM错误",
            "分差",
            "胜者",
            "评判是否一致",
        ]
        fmts += [None, "int", "int", "float3", "int", None, None, None, "int", None, None]
    return headers, fmts


def _detail_row(
    r: Dict[str, Any],
    cat_fields: Sequence[str],
    has_lo: bool,
) -> List[Any]:
    cats = r.get("categories") or {}
    hg = _result_system_block(r, SYSTEM_HYPERGRAPH)
    hm = hg.get("metrics") or {}
    hs = _num(hg.get("score"))
    row: List[Any] = [
        r.get("id"),
        r.get("question"),
        r.get("expected_answer"),
    ]
    for f in cat_fields:
        row.append(cats.get(f))
    hg_err = hg.get("query_error") or hg.get("judge_error")
    row += [
        hg.get("llm_acc"),
        _int(hg.get("score")),
        _int(hg.get("query_status")),
        _num(hm.get("query_latency_s")),
        _num(hm.get("retrieve_latency_s")),
        _int(hm.get("total_tokens")),
        len(hg.get("retrieval_sources") or []),
        hg.get("judge_reason"),
        hg.get("answer") or hg.get("raw_answer"),
        hg_err,
    ]
    if has_lo:
        lo = _result_system_block(r, SYSTEM_LLM_ONLY)
        lm = lo.get("metrics") or {}
        ls = _num(lo.get("score"))
        lo_err = lo.get("query_error") or lo.get("judge_error")
        hj, lj = hg.get("llm_acc"), lo.get("llm_acc")
        same = ""
        if hj or lj:
            same = "是" if hj == lj else "否"
        delta = None
        if hs is not None and ls is not None:
            delta = int(hs - ls) if float(hs).is_integer() and float(ls).is_integer() else hs - ls
        row += [
            lo.get("llm_acc"),
            _int(lo.get("score")),
            _int(lo.get("query_status")),
            _num(lm.get("query_latency_s")),
            _int(lm.get("total_tokens")),
            lo.get("judge_reason"),
            lo.get("answer") or lo.get("raw_answer"),
            lo_err,
            delta,
            _winner(hs, ls),
            same,
        ]
    return row


def _add_details(
    wb: Workbook,
    results: Sequence[Dict[str, Any]],
    systems: Sequence[str],
) -> Tuple[str, Optional[str], Optional[str], Optional[str]]:
    cat_fields = _category_fields_from_results(results)
    has_lo = SYSTEM_LLM_ONLY in systems
    headers, fmts = _detail_headers(cat_fields, has_lo)
    all_rows = [_detail_row(r, cat_fields, has_lo) for r in results]

    name = _sheet_name("逐题明细")
    ws = wb.create_sheet(name)
    _write_title(ws, "逐题明细（点选单元格可在编辑栏看全文）", len(headers))
    _write_table(
        ws,
        headers,
        all_rows,
        start_row=3,
        col_formats=fmts,
        freeze=True,
        autofilter=True,
        max_width=28,
        zebra=True,
    )
    ws.freeze_panes = "D4"
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 28

    # 差异题
    diff_name = None
    if has_lo:
        winner_idx = headers.index("胜者") if "胜者" in headers else None
        same_idx = headers.index("评判是否一致") if "评判是否一致" in headers else None
        diff_rows = []
        for row in all_rows:
            win = row[winner_idx] if winner_idx is not None else ""
            same = row[same_idx] if same_idx is not None else ""
            if win in ("超图", "纯LLM") or same == "否":
                diff_rows.append(row)
        diff_name = _sheet_name("差异题")
        dws = wb.create_sheet(diff_name)
        _write_title(
            dws,
            f"分数或评判不一致的题（{len(diff_rows)} / {len(all_rows)}）",
            len(headers),
        )
        _write_table(
            dws,
            headers,
            diff_rows,
            start_row=3,
            col_formats=fmts,
            freeze=True,
            autofilter=True,
            max_width=28,
            zebra=True,
        )
        dws.freeze_panes = "D4"
        dws.column_dimensions["B"].width = 36
        # 胜者着色
        if winner_idx is not None:
            c = winner_idx + 1
            for i in range(len(diff_rows)):
                cell = dws.cell(4 + i, c)
                if cell.value == "超图":
                    cell.fill = _FILL_POS
                elif cell.value == "纯LLM":
                    cell.fill = _FILL_NEG

    fail_name = None
    fail_rows: List[List[Any]] = []
    fail_headers = [
        "ID",
        "系统",
        "问题",
        "query状态",
        "query错误",
        "judge状态",
        "judge错误",
    ]
    for r in results:
        for system in systems:
            b = _result_system_block(r, system)
            qerr = b.get("query_error")
            jerr = b.get("judge_error")
            qs, js = b.get("query_status"), b.get("judge_status")
            if qerr or jerr or qs == 0 or js == 0:
                fail_rows.append(
                    [
                        r.get("id"),
                        _sys_label(system),
                        r.get("question"),
                        qs,
                        qerr,
                        js,
                        jerr,
                    ]
                )
    fail_name = _sheet_name("失败列表")
    fws = wb.create_sheet(fail_name)
    _write_title(
        fws,
        f"问答 / 评判失败（{len(fail_rows)} 条）" if fail_rows else "问答 / 评判失败（无）",
        len(fail_headers),
    )
    _write_table(
        fws,
        fail_headers,
        fail_rows,
        start_row=3,
        freeze=True,
        autofilter=bool(fail_rows),
        max_width=40,
    )

    dim_name = None
    dim_rows: List[List[Any]] = []
    for r in results:
        for system in systems:
            b = _result_system_block(r, system)
            for ds in b.get("dimension_scores") or []:
                if not isinstance(ds, dict):
                    continue
                dim_rows.append(
                    [
                        r.get("id"),
                        _sys_label(system),
                        ds.get("dimension"),
                        _int(ds.get("score")),
                        ds.get("pass"),
                        ds.get("comment"),
                    ]
                )
    dim_name = _sheet_name("维度评分")
    dim_ws = wb.create_sheet(dim_name)
    dim_headers = ["ID", "系统", "评分维度", "分数", "通过", "评语"]
    _write_title(dim_ws, "各评分维度明细", len(dim_headers))
    _write_table(
        dim_ws,
        dim_headers,
        dim_rows,
        start_row=3,
        col_formats=[None, None, None, "int", None, None],
        freeze=True,
        autofilter=bool(dim_rows),
        max_width=40,
    )
    return name, diff_name, fail_name, dim_name


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------


def write_report_workbook(
    report: Dict[str, Any],
    out_path: Union[str, Path],
    *,
    eval_data: Optional[Dict[str, Any]] = None,
) -> Path:
    """把 report（及可选 evals）写成多 sheet 的 xlsx。"""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    summary = report.get("summary") or {}
    if eval_data is None and isinstance(report.get("results"), list):
        eval_data = report
    results = list((eval_data or {}).get("results") or [])
    systems = _systems(summary)

    wb = Workbook()
    # 占位，写完目录后删
    default = wb.active
    default.title = "tmp"

    toc: List[Tuple[str, str]] = []

    toc.append((_add_meta(wb, report, eval_data), "评测配置、模型、题量、耗时"))
    toc.append((_add_overall(wb, summary), "两路总体正确率 / 均分 / 耗时 / Token，含分数直方图"))
    cmp_name = _add_comparison(wb, summary)
    if cmp_name:
        toc.append((cmp_name, "配对胜负、均分差、评判交叉表"))
    toc.append((_add_category_long(wb, summary), "主分类 + 附加分类的长表，可筛选"))
    cmp_cat = _add_category_compare(wb, summary)
    if cmp_cat:
        toc.append((cmp_cat, "同一分组下超图 vs 纯LLM 的正确率/均分差，绿=超图更好"))
    toc.append((_add_latency_tokens(wb, summary), "检索各阶段耗时和 token 用量"))

    if results:
        d_name, diff_name, fail_name, dim_name = _add_details(wb, results, systems)
        toc.append((d_name, "每题两路评判、分数、答案、原因"))
        if diff_name:
            toc.append((diff_name, "分数或评判不一致的题，便于抽查看法差异"))
        if fail_name:
            toc.append((fail_name, "问答或评判失败清单"))
        if dim_name:
            toc.append((dim_name, "评分维度逐条展开"))
    else:
        ws = wb.create_sheet(_sheet_name("逐题明细"))
        ws["A1"] = "未提供 evals JSON，没有逐题明细。重新导出时加上 --evals。"
        ws["A1"].font = _FONT_CELL
        toc.append((ws.title, "无逐题数据（缺少 evals）"))

    _add_toc(wb, toc)
    if "tmp" in wb.sheetnames:
        del wb["tmp"]

    wb.save(str(out))
    return out


def load_report_and_evals(
    report_path: Optional[Union[str, Path]] = None,
    evals_path: Optional[Union[str, Path]] = None,
) -> Tuple[Dict[str, Any], Optional[Dict[str, Any]]]:
    """读 JSON。若只给 evals，会现场生成 report。"""
    report: Optional[Dict[str, Any]] = None
    evals: Optional[Dict[str, Any]] = None

    def _load(p: Union[str, Path]) -> Dict[str, Any]:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)

    if report_path:
        data = _load(report_path)
        if isinstance(data.get("results"), list):
            evals = data
            report = build_report_document(data, source_path=str(report_path))
        elif isinstance(data.get("summary"), dict):
            report = data
        else:
            raise ValueError(f"无法识别的 JSON（缺少 summary / results）: {report_path}")

    if evals_path:
        evals = _load(evals_path)
        if not isinstance(evals.get("results"), list):
            raise ValueError(f"非法评测结果（缺少 results）: {evals_path}")
        if report is None:
            report = build_report_document(evals, source_path=str(evals_path))

    if report is None:
        raise FileNotFoundError("需要 --report 或 --evals（或 config.yaml 里的对应路径）")
    return report, evals


def export_paths(
    *,
    report_path: Optional[Union[str, Path]] = None,
    evals_path: Optional[Union[str, Path]] = None,
    out_path: Optional[Union[str, Path]] = None,
) -> Path:
    report, evals = load_report_and_evals(report_path, evals_path)
    if out_path is None:
        src = Path(report_path or evals_path or "report.json")
        out_path = src.with_suffix(".xlsx")
    return write_report_workbook(report, out_path, eval_data=evals)


def main(argv: Optional[List[str]] = None) -> int:
    import argparse
    import sys

    root = Path(__file__).resolve().parent.parent
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))

    from benchmark.utils import resolve_path
    from benchmark2.config import DEFAULT_CONFIG_PATH, Benchmark2Config

    p = argparse.ArgumentParser(description="benchmark2 report JSON → 多表 Excel")
    p.add_argument("--config", default=DEFAULT_CONFIG_PATH, help="yaml 配置路径")
    p.add_argument("--report", default=None, help="report JSON；默认用 config 的 report_path")
    p.add_argument("--evals", default=None, help="evals JSON，用于逐题明细；默认用 config")
    p.add_argument("--out", default=None, help="xlsx 写出路径")
    args = p.parse_args(argv)

    cfg = Benchmark2Config.from_yaml(args.config).resolve_paths()
    report_p = Path(args.report) if args.report else cfg.report_file()
    if args.report and not report_p.is_absolute():
        report_p = resolve_path(report_p)
    evals_p = Path(args.evals) if args.evals else cfg.eval_results_file()
    if args.evals and not evals_p.is_absolute():
        evals_p = resolve_path(evals_p)
    out = Path(args.out) if args.out else cfg.report_excel_file()
    if args.out and not out.is_absolute():
        out = resolve_path(out)

    if not report_p.is_file() and not evals_p.is_file():
        print(
            "未找到 report / evals JSON，无法导出。\n"
            f"  report: {report_p}\n"
            f"  evals:  {evals_p}\n"
            "  请先 run.mode=report，或用 --report / --evals 指定文件。",
            file=sys.stderr,
        )
        return 2

    rp = report_p if report_p.is_file() else None
    ep = evals_p if evals_p.is_file() else None
    if rp is None and ep is not None:
        print(f"[excel] 无 report，改从 evals 生成: {ep}", file=sys.stderr)
    written = export_paths(report_path=rp, evals_path=ep, out_path=out)
    print(f"[saved] {written}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
