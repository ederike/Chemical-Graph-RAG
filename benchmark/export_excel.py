#!/usr/bin/env python3
"""
把 benchmark 的 report / evals JSON 汇总成多表 Excel。

默认读 config.yaml 的 report / eval 路径，写出同名 .xlsx::

    python -m benchmark.export_excel
    python -m benchmark.export_excel \\
        --report benchmark_results/report_complex.json \\
        --evals  benchmark_results/evals_complex.json \\
        --out    benchmark_results/report_complex.xlsx

也可用 run.mode=excel，或在 report / all 结束后自动写出。
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

from .evaluator import (
    RETRIEVE_STAGE_KEYS,
    SYSTEM_HYPERGRAPH,
    SYSTEM_LLM_ONLY,
    QueryEvaluator,
    project_system_row,
)

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:  # pragma: no cover
    raise ImportError("写出 Excel 需要 openpyxl，请 pip install openpyxl") from e


SYSTEM_LABELS = {
    SYSTEM_HYPERGRAPH: "超图",
    SYSTEM_LLM_ONLY: "纯LLM",
}
JUDGE_ORDER = ("正确", "错误", "未知")
RECALL_ORDER = ("hit", "partial", "miss", "unknown")
EXCEL_CELL_MAX = 32767
_INVALID_SHEET_RE = re.compile(r"[\[\]\*?:/\\]")

_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FILL_HG = PatternFill("solid", fgColor="D6EAF8")
_FILL_LO = PatternFill("solid", fgColor="FDEBD0")
_FILL_POS = PatternFill("solid", fgColor="D5F5E3")
_FILL_NEG = PatternFill("solid", fgColor="FADBD8")
_FILL_TIE = PatternFill("solid", fgColor="F4F6F7")
_FILL_ALT = PatternFill("solid", fgColor="F7F9FB")
_FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
_FONT_HEADER = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_FONT_TITLE = Font(name="微软雅黑", bold=True, color="FFFFFF", size=14)
_FONT_CELL = Font(name="微软雅黑", size=10)
_FONT_BOLD = Font(name="微软雅黑", bold=True, size=10)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_R = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# 小工具
# ---------------------------------------------------------------------------


def _sheet_name(name: str, used: Optional[set] = None) -> str:
    s = _INVALID_SHEET_RE.sub(" ", str(name or "Sheet")).strip() or "Sheet"
    s = s[:31]
    if used is None:
        return s
    base, n = s, 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = base[: 31 - len(suffix)] + suffix
    used.add(s)
    return s


def _cell_text(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v)
    if len(s) > EXCEL_CELL_MAX:
        s = s[: EXCEL_CELL_MAX - 1] + "…"
    return s


def _fmt_pct(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _delta_fill(delta: Optional[float], *, higher_is_better: bool) -> Optional[PatternFill]:
    if delta is None:
        return None
    try:
        d = float(delta)
    except (TypeError, ValueError):
        return None
    if d == 0:
        return _FILL_TIE
    better = d > 0 if higher_is_better else d < 0
    return _FILL_POS if better else _FILL_NEG


def _systems(summary: Dict[str, Any]) -> List[str]:
    raw = summary.get("systems")
    out: List[str] = []
    if isinstance(raw, list):
        for k in raw:
            if k in SYSTEM_LABELS:
                out.append(k)
    if not out:
        for k in (SYSTEM_HYPERGRAPH, SYSTEM_LLM_ONLY):
            block = summary.get(k)
            if isinstance(block, dict) and isinstance(block.get("llm_acc"), dict):
                out.append(k)
    return out or [SYSTEM_HYPERGRAPH]


def _sys_label(key: str) -> str:
    return SYSTEM_LABELS.get(key, key)


def _nested(d: Optional[dict], *keys: str, default: Any = None) -> Any:
    cur: Any = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur


def _hop_keys(by_hop: Dict[str, Any]) -> List[str]:
    def key(h: str):
        try:
            return (0, int(h))
        except (TypeError, ValueError):
            return (1, str(h))

    return sorted((by_hop or {}).keys(), key=key)


def _auto_width(ws: Worksheet, *, min_w: float = 8, max_w: float = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0.0
        for cell in col:
            if cell.value is None:
                continue
            text = str(cell.value).replace("\n", " ")
            w = 0.0
            for ch in text[:80]:
                w += 1.7 if ord(ch) > 127 else 1.05
            longest = max(longest, w)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 1.5))


def _apply_number_format(cell, fmt: Optional[str]) -> None:
    if fmt == "pct":
        cell.number_format = "0.00%"
    elif fmt == "int":
        cell.number_format = "#,##0"
    elif fmt == "float2":
        cell.number_format = "0.00"
    elif fmt == "float3":
        cell.number_format = "0.000"
    elif fmt == "float4":
        cell.number_format = "0.0000"


def _write_title(ws: Worksheet, title: str, n_cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, n_cols))
    cell = ws.cell(1, 1, title)
    cell.fill = _FILL_TITLE
    cell.font = _FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def _write_table(
    ws: Worksheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    start_row: int = 1,
    col_formats: Optional[Sequence[Optional[str]]] = None,
    freeze: bool = True,
    autofilter: bool = True,
    zebra: bool = True,
    max_width: float = 48,
) -> int:
    """写一张表，返回下一空行号。"""
    n_cols = len(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _THIN
    ws.row_dimensions[start_row].height = 22

    for i, row in enumerate(rows):
        r = start_row + 1 + i
        alt = zebra and (i % 2 == 1)
        for c in range(1, n_cols + 1):
            raw = row[c - 1] if c - 1 < len(row) else None
            val = _cell_text(raw)
            cell = ws.cell(r, c, val)
            cell.font = _FONT_CELL
            cell.border = _THIN
            fmt = col_formats[c - 1] if col_formats and c - 1 < len(col_formats) else None
            _apply_number_format(cell, fmt)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = _ALIGN_R
            else:
                cell.alignment = _ALIGN_L
            if alt:
                cell.fill = _FILL_ALT

    last = start_row + len(rows)
    if autofilter and rows:
        ws.auto_filter.ref = (
            f"{get_column_letter(1)}{start_row}:{get_column_letter(n_cols)}{last}"
        )
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"
    _auto_width(ws, max_w=max_width)
    ws.sheet_view.showGridLines = False
    return last + 1


def _flatten_meta_value(v: Any) -> Any:
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    return json.dumps(v, ensure_ascii=False)


def _sys_fill(system: str) -> PatternFill:
    if system == SYSTEM_HYPERGRAPH:
        return _FILL_HG
    if system == SYSTEM_LLM_ONLY:
        return _FILL_LO
    return _FILL_ALT


def _join_names(names: Any) -> Optional[str]:
    if not names:
        return None
    if isinstance(names, str):
        return names
    if isinstance(names, (list, tuple)):
        return " / ".join(str(x) for x in names if x)
    return str(names)


# ---------------------------------------------------------------------------
# 各 sheet
# ---------------------------------------------------------------------------


def _add_toc(wb: Workbook, items: Sequence[Tuple[str, str]]) -> None:
    ws = wb.create_sheet(_sheet_name("目录"), 0)
    _write_title(ws, "Benchmark 评测报告", 2)
    ws.cell(2, 1, "点击工作表名跳转。JSON 嵌套已摊平成表格。")
    ws.cell(2, 1).font = Font(name="微软雅黑", italic=True, size=10, color="666666")
    headers = ("工作表", "说明")
    rows = [(name, desc) for name, desc in items]
    _write_table(ws, headers, rows, start_row=4, freeze=False, autofilter=False, zebra=True)
    for i, (sheet, _) in enumerate(items):
        cell = ws.cell(5 + i, 1, sheet)
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate, location=f"'{sheet}'!A1", display=sheet
        )
        cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        cell.alignment = _ALIGN_L
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 72


def _add_meta(
    wb: Workbook, report: Dict[str, Any], eval_data: Optional[Dict[str, Any]]
) -> str:
    name = _sheet_name("运行信息")
    ws = wb.create_sheet(name)
    meta = dict(report.get("meta") or {})
    ev_meta = (eval_data or {}).get("meta") if isinstance(eval_data, dict) else {}
    if isinstance(ev_meta, dict):
        for k, v in ev_meta.items():
            meta.setdefault(k, v)
    skip = {"config", "dataset_meta", "progress"}
    rows: List[List[Any]] = []
    preferred = [
        "created_at",
        "updated_at",
        "eval_done",
        "done",
        "query_mode",
        "judge_model",
        "judge_mode",
        "enable_llm_only",
        "enable_doc_recall",
        "n_results",
        "n_questions",
        "n_total_planned",
        "n_skipped_gen_fail",
        "eval_elapsed_s",
        "elapsed_s",
        "db_path",
        "hop_counts",
        "seed",
        "source_path",
        "config_file",
        "dhmf_config_path",
    ]
    seen = set()
    for k in preferred:
        if k in meta:
            rows.append([k, _flatten_meta_value(meta[k])])
            seen.add(k)
    for k in sorted(meta.keys()):
        if k in seen or k in skip:
            continue
        rows.append([k, _flatten_meta_value(meta[k])])
    cfg = meta.get("config") if isinstance(meta.get("config"), dict) else None
    if cfg:
        rows.append(["— 配置快照 —", ""])
        for k, v in cfg.items():
            rows.append([f"config.{k}", _flatten_meta_value(v)])
    _write_title(ws, "运行信息", 2)
    _write_table(
        ws,
        ("字段", "值"),
        rows,
        start_row=3,
        freeze=True,
        autofilter=False,
        max_width=80,
    )
    return name


def _metric_row_overall(
    sys_blocks: Dict[str, Dict[str, Any]], systems: Sequence[str]
) -> Tuple[List[List[Any]], List[Tuple[str, str, Tuple[str, ...], Optional[bool]]], bool]:
    def g(sys: str, *path, default=None):
        return _nested(sys_blocks.get(sys), *path, default=default)

    specs: List[Tuple[str, str, Tuple[str, ...], Optional[bool]]] = [
        ("题目数 n", "int", ("n_total",), None),
        ("流水线成功", "int", ("pipeline", "n_ok"), True),
        ("问答失败", "int", ("pipeline", "n_query_fail"), False),
        ("评判失败", "int", ("pipeline", "n_judge_fail"), False),
        ("已评判", "int", ("llm_acc", "n_judged"), True),
        ("正确", "int", ("llm_acc", "n_correct"), True),
        ("错误", "int", ("llm_acc", "n_wrong"), False),
        ("未知", "int", ("llm_acc", "counts", "未知"), None),
        ("正确率", "pct", ("llm_acc", "accuracy"), True),
        ("错误率", "pct", ("llm_acc", "error_rate"), False),
        ("均文档召回", "pct", ("doc_recall", "mean_recall"), True),
        ("均问答耗时 s", "float3", ("latency", "mean_query_s"), False),
        ("均检索耗时 s", "float3", ("latency", "mean_retrieve_s"), False),
        ("均墙钟耗时 s", "float3", ("latency", "mean_wall_s"), False),
        ("问答总耗时 s", "float3", ("latency", "sum_query_s"), False),
        ("检索总耗时 s", "float3", ("latency", "sum_retrieve_s"), False),
        ("墙钟总耗时 s", "float3", ("latency", "sum_wall_s"), False),
        ("prompt tokens 合计", "int", ("tokens", "sum_prompt"), None),
        ("completion tokens 合计", "int", ("tokens", "sum_completion"), None),
        ("tokens 合计", "int", ("tokens", "sum_total"), None),
        ("均 prompt tokens", "float2", ("tokens", "mean_prompt"), None),
        ("均 completion tokens", "float2", ("tokens", "mean_completion"), None),
        ("评判 prompt tokens 合计", "int", ("tokens", "sum_judge_prompt"), None),
        ("评判 completion tokens 合计", "int", ("tokens", "sum_judge_completion"), None),
        ("评判 tokens 合计", "int", ("tokens", "sum_judge_total"), None),
    ]

    has_pair = len(systems) >= 2
    rows: List[List[Any]] = []
    for name, fmt, path, better in specs:
        vals = [g(s, *path) for s in systems]
        if fmt == "pct":
            vals = [_fmt_pct(v) for v in vals]
        elif fmt == "int":
            vals = [_int(v) if v is not None else None for v in vals]
        else:
            vals = [_num(v) if v is not None else None for v in vals]
        delta = None
        if has_pair and vals[0] is not None and vals[1] is not None:
            try:
                delta = round(float(vals[0]) - float(vals[1]), 6)
            except (TypeError, ValueError):
                delta = None
        row: List[Any] = [name, *vals]
        if has_pair:
            row.append(delta)
        rows.append(row)
    return rows, specs, has_pair


def _add_overall(wb: Workbook, summary: Dict[str, Any]) -> str:
    sheet_name = _sheet_name("总体指标")
    ws = wb.create_sheet(sheet_name)
    systems = _systems(summary)
    sys_blocks = {s: summary.get(s) or {} for s in systems}
    rows, specs, has_pair = _metric_row_overall(sys_blocks, systems)

    headers = ["指标"] + [_sys_label(s) for s in systems]
    col_formats: List[Optional[str]] = [None] + [None] * len(systems)
    if has_pair:
        headers.append("差值（超图 − 纯LLM）")
        col_formats.append("float4")

    n_cols = len(headers)
    _write_title(ws, "总体指标：超图 vs 纯LLM" if has_pair else "总体指标", n_cols)
    start = 3
    _write_table(
        ws,
        headers,
        rows,
        start_row=start,
        col_formats=col_formats,
        freeze=True,
        autofilter=False,
        zebra=True,
    )
    for i, (_metric, fmt, _path, better) in enumerate(specs):
        r = start + 1 + i
        for c, s in enumerate(systems, 2):
            cell = ws.cell(r, c)
            _apply_number_format(cell, fmt)
            cell.fill = _sys_fill(s)
        if has_pair:
            dcell = ws.cell(r, 2 + len(systems))
            _apply_number_format(dcell, fmt)
            if better is not None:
                fill = _delta_fill(dcell.value, higher_is_better=bool(better))
                if fill is not None:
                    dcell.fill = fill
    return sheet_name


def _parse_crosstab(xt: Dict[str, Any]) -> Dict[Tuple[str, str], int]:
    matrix: Dict[Tuple[str, str], int] = {}
    for k, v in (xt or {}).items():
        key = str(k)
        if not key.startswith("超图") or "_纯LLM" not in key:
            continue
        rest = key[len("超图") :]
        hj, lj = rest.split("_纯LLM", 1)
        try:
            matrix[(hj, lj)] = int(v)
        except (TypeError, ValueError):
            continue
    return matrix


def _add_comparison(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    cmp_ = summary.get("comparison")
    if not isinstance(cmp_, dict) or not cmp_:
        return None
    name = _sheet_name("系统对比")
    ws = wb.create_sheet(name)
    _write_title(ws, "超图 vs 纯LLM 对比", 5)

    win = cmp_.get("win") or {}
    n_paired = _int(cmp_.get("n_paired")) or 0
    n_hg = _int(win.get("hypergraph")) or 0
    n_lo = _int(win.get("llm_only")) or 0
    n_tie = _int(win.get("tie")) or 0

    overview = [
        ["配对题数", n_paired],
        ["对比打分题数", _int(cmp_.get("n_pairwise_judged"))],
        ["双方都正确", _int(cmp_.get("both_correct"))],
        ["仅超图正确", _int(cmp_.get("hypergraph_only"))],
        ["仅纯LLM正确", _int(cmp_.get("llm_only_only"))],
        ["双方都错误", _int(cmp_.get("both_wrong"))],
        ["胜负来源", cmp_.get("win_note") or "胜负优先取裁判 comparison.winner"],
        ["超图胜", n_hg],
        ["纯LLM胜", n_lo],
        ["平局", n_tie],
        ["超图胜率", (n_hg / n_paired) if n_paired else None],
        ["纯LLM胜率", (n_lo / n_paired) if n_paired else None],
        ["平局率", (n_tie / n_paired) if n_paired else None],
    ]
    _write_table(
        ws,
        ("项目", "值"),
        overview,
        start_row=3,
        freeze=False,
        autofilter=False,
    )
    for r, row in enumerate(overview, 4):
        label = str(row[0])
        if "率" in label:
            _apply_number_format(ws.cell(r, 2), "pct")
        elif label == "超图胜" or label == "仅超图正确":
            ws.cell(r, 2).fill = _FILL_POS
        elif label == "纯LLM胜" or label == "仅纯LLM正确":
            ws.cell(r, 2).fill = _FILL_NEG
        elif label in ("平局", "双方都正确", "双方都错误"):
            ws.cell(r, 2).fill = _FILL_TIE

    bar_row = 18
    ws.cell(bar_row, 1, "胜负").font = _FONT_BOLD
    _write_table(
        ws,
        ("结果", "题数"),
        [["超图胜", n_hg], ["纯LLM胜", n_lo], ["平局", n_tie]],
        start_row=bar_row + 1,
        col_formats=[None, "int"],
        freeze=False,
        autofilter=False,
        zebra=False,
    )
    chart = BarChart()
    chart.type = "col"
    chart.title = "对比胜负"
    chart.y_axis.title = "题数"
    data_ref = Reference(ws, min_col=2, min_row=bar_row + 1, max_row=bar_row + 4)
    cats = Reference(ws, min_col=1, min_row=bar_row + 2, max_row=bar_row + 4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.width = 12
    chart.height = 7
    ws.add_chart(chart, "D19")

    xt_start = 25
    ws.cell(xt_start, 1, "评判交叉表（行=超图，列=纯LLM）").font = _FONT_BOLD
    matrix = _parse_crosstab(cmp_.get("judgment_crosstab") or {})
    labels = list(JUDGE_ORDER)
    xt_headers = ["超图 \\ 纯LLM"] + [f"纯LLM·{x}" for x in labels] + ["行合计"]
    xt_rows = []
    col_tot = {lj: 0 for lj in labels}
    for hj in labels:
        row = [f"超图·{hj}"]
        s = 0
        for lj in labels:
            v = matrix.get((hj, lj), 0)
            row.append(v)
            s += v
            col_tot[lj] += v
        row.append(s)
        xt_rows.append(row)
    xt_rows.append(["列合计"] + [col_tot[lj] for lj in labels] + [sum(col_tot.values())])
    _write_table(
        ws,
        xt_headers,
        xt_rows,
        start_row=xt_start + 1,
        col_formats=[None] + ["int"] * (len(labels) + 1),
        freeze=False,
        autofilter=False,
        zebra=False,
    )
    return name


def _add_by_hop(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    systems = _systems(summary)
    hop_maps: Dict[str, Dict[str, Any]] = {}
    hops: List[str] = []
    seen = set()
    for s in systems:
        bh = (summary.get(s) or {}).get("by_hop") or {}
        hop_maps[s] = bh if isinstance(bh, dict) else {}
        for h in _hop_keys(hop_maps[s]):
            if h not in seen:
                hops.append(h)
                seen.add(h)
    if not hops:
        return None

    name = _sheet_name("按跳数")
    ws = wb.create_sheet(name)
    has_pair = len(systems) >= 2
    headers = [
        "跳数",
        "n",
        "正确率_超图",
        "正确率_纯LLM",
        "正确率差",
        "正确_超图",
        "正确_纯LLM",
        "错误_超图",
        "错误_纯LLM",
        "均问答s_超图",
        "均问答s_纯LLM",
        "均检索s_超图",
        "均文档召回_超图",
    ]
    fmts: List[Optional[str]] = [
        None,
        "int",
        "pct",
        "pct",
        "pct",
        "int",
        "int",
        "int",
        "int",
        "float3",
        "float3",
        "float3",
        "pct",
    ]
    if not has_pair:
        headers = [
            "跳数",
            "n",
            "正确",
            "错误",
            "正确率",
            "均问答s",
            "均检索s",
            "均文档召回",
        ]
        fmts = [None, "int", "int", "int", "pct", "float3", "float3", "pct"]

    rows: List[List[Any]] = []
    chart_rows: List[List[Any]] = []
    for hop in hops:
        hg = hop_maps.get(SYSTEM_HYPERGRAPH, {}).get(hop) or {}
        lo = hop_maps.get(SYSTEM_LLM_ONLY, {}).get(hop) or {}
        hg_acc = (hg.get("llm_acc_counts") or {})
        lo_acc = (lo.get("llm_acc_counts") or {})
        n = _int(hg.get("n")) or _int(lo.get("n"))
        acc_h = _fmt_pct(hg.get("accuracy"))
        acc_l = _fmt_pct(lo.get("accuracy"))
        delta = None
        if acc_h is not None and acc_l is not None:
            delta = round(float(acc_h) - float(acc_l), 6)
        if has_pair:
            rows.append(
                [
                    hop,
                    n,
                    acc_h,
                    acc_l,
                    delta,
                    _int(hg_acc.get("正确")),
                    _int(lo_acc.get("正确")),
                    _int(hg_acc.get("错误")),
                    _int(lo_acc.get("错误")),
                    _num(hg.get("mean_query_latency_s")),
                    _num(lo.get("mean_query_latency_s")),
                    _num(hg.get("mean_retrieve_latency_s")),
                    _fmt_pct(hg.get("mean_doc_recall")),
                ]
            )
        else:
            src = hg or lo
            counts = src.get("llm_acc_counts") or {}
            rows.append(
                [
                    hop,
                    n,
                    _int(counts.get("正确")),
                    _int(counts.get("错误")),
                    _fmt_pct(src.get("accuracy")),
                    _num(src.get("mean_query_latency_s")),
                    _num(src.get("mean_retrieve_latency_s")),
                    _fmt_pct(src.get("mean_doc_recall")),
                ]
            )
        chart_rows.append(
            [hop]
            + [
                _fmt_pct((hop_maps.get(s) or {}).get(hop, {}).get("accuracy")) or 0
                for s in systems
            ]
        )

    _write_title(ws, "按跳数切分：超图 vs 纯LLM" if has_pair else "按跳数切分", len(headers))
    start = 3
    _write_table(
        ws,
        headers,
        rows,
        start_row=start,
        col_formats=fmts,
        freeze=True,
        autofilter=True,
        max_width=16,
        zebra=True,
    )
    if has_pair:
        for i in range(len(rows)):
            r = start + 1 + i
            fill = _delta_fill(ws.cell(r, 5).value, higher_is_better=True)
            if fill is not None:
                ws.cell(r, 5).fill = fill
            ws.cell(r, 3).fill = _FILL_HG
            ws.cell(r, 4).fill = _FILL_LO

    chart_start = start + len(rows) + 3
    ws.cell(chart_start, 1, "各跳正确率").font = _FONT_BOLD
    chart_headers = ["跳数"] + [_sys_label(s) for s in systems]
    _write_table(
        ws,
        chart_headers,
        chart_rows,
        start_row=chart_start + 1,
        col_formats=[None] + ["pct"] * len(systems),
        freeze=False,
        autofilter=False,
        zebra=False,
    )
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "各跳正确率"
    chart.y_axis.title = "正确率"
    chart.x_axis.title = "跳数"
    data_ref = Reference(
        ws,
        min_col=2,
        min_row=chart_start + 1,
        max_col=1 + len(systems),
        max_row=chart_start + 1 + len(chart_rows),
    )
    cats = Reference(
        ws, min_col=1, min_row=chart_start + 2, max_row=chart_start + 1 + len(chart_rows)
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, "F" + str(chart_start))
    return name


def _add_doc_recall(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    hg = summary.get(SYSTEM_HYPERGRAPH) or summary
    rec = (hg or {}).get("doc_recall")
    if not isinstance(rec, dict) or not rec:
        rec = summary.get("doc_recall")
    if not isinstance(rec, dict) or not rec:
        return None

    name = _sheet_name("文档召回")
    ws = wb.create_sheet(name)
    _write_title(ws, "文档召回 × 答案准确率", 6)

    vs = rec.get("vs_accuracy") if isinstance(rec.get("vs_accuracy"), dict) else {}
    overview = [
        ["均文档召回", _fmt_pct(rec.get("mean_recall"))],
        ["有召回字段题数", _int(vs.get("n_with_recall"))],
        ["总题数", _int(vs.get("n_total") or summary.get("n_total"))],
    ]
    _write_table(
        ws,
        ("项目", "值"),
        overview,
        start_row=3,
        freeze=False,
        autofilter=False,
    )
    _apply_number_format(ws.cell(4, 2), "pct")

    definition = vs.get("definition") if isinstance(vs.get("definition"), dict) else {}
    def_start = 8
    ws.cell(def_start, 1, "口径").font = _FONT_BOLD
    def_rows = [[k, v] for k, v in definition.items()]
    if def_rows:
        _write_table(
            ws,
            ("桶", "说明"),
            def_rows,
            start_row=def_start + 1,
            freeze=False,
            autofilter=False,
            max_width=80,
        )

    matrix_start = def_start + 2 + max(len(def_rows), 1) + 1
    ws.cell(matrix_start, 1, "召回 × 准确率（行=召回桶）").font = _FONT_BOLD
    cells: Dict[Tuple[str, str], int] = {}
    for item in vs.get("matrix") or []:
        if not isinstance(item, dict):
            continue
        cells[(str(item.get("recall")), str(item.get("accuracy")))] = _int(item.get("n")) or 0
    table = vs.get("table") if isinstance(vs.get("table"), dict) else {}
    alias = {"正确": "correct", "错误": "wrong", "未知": "unknown"}
    for b in RECALL_ORDER:
        for lab in JUDGE_ORDER:
            if (b, lab) in cells:
                continue
            key = f"{b}_{alias.get(lab, lab)}"
            if key in table:
                cells[(b, lab)] = _int(table.get(key)) or 0

    row_tot = vs.get("row_totals") if isinstance(vs.get("row_totals"), dict) else {}
    col_tot = vs.get("col_totals") if isinstance(vs.get("col_totals"), dict) else {}
    mx_headers = ["召回 \\ 准确率"] + list(JUDGE_ORDER) + ["行合计"]
    mx_rows = []
    for b in RECALL_ORDER:
        row = [b]
        s = 0
        for lab in JUDGE_ORDER:
            v = cells.get((b, lab), 0)
            row.append(v)
            s += v
        row.append(_int(row_tot.get(b)) if row_tot.get(b) is not None else s)
        mx_rows.append(row)
    mx_rows.append(
        ["列合计"]
        + [
            _int(col_tot.get(lab))
            if col_tot.get(lab) is not None
            else sum(cells.get((b, lab), 0) for b in RECALL_ORDER)
            for lab in JUDGE_ORDER
        ]
        + [
            _int(vs.get("n_total"))
            or sum(_int(r[-1]) or 0 for r in mx_rows)
        ]
    )
    _write_table(
        ws,
        mx_headers,
        mx_rows,
        start_row=matrix_start + 1,
        col_formats=[None] + ["int"] * (len(JUDGE_ORDER) + 1),
        freeze=False,
        autofilter=False,
        zebra=False,
    )

    br = vs.get("hit_wrong_breakdown") if isinstance(vs.get("hit_wrong_breakdown"), dict) else {}
    if br:
        br_start = matrix_start + 3 + len(mx_rows)
        ws.cell(br_start, 1, "hit × 错误 拆分").font = _FONT_BOLD
        _write_table(
            ws,
            ("项目", "值"),
            [
                ["query 成功", _int(br.get("query_ok"))],
                ["query 失败", _int(br.get("query_fail"))],
                ["说明", br.get("note")],
            ],
            start_row=br_start + 1,
            freeze=False,
            autofilter=False,
            max_width=72,
        )
    return name


def _add_latency_tokens(wb: Workbook, summary: Dict[str, Any]) -> str:
    name = _sheet_name("延迟与Token")
    ws = wb.create_sheet(name)
    _write_title(ws, "检索阶段延迟 + Token", 6)
    systems = _systems(summary)

    lat_headers = ["系统", "阶段", "均值 s", "合计 s"]
    lat_rows: List[List[Any]] = []
    stages = [
        ("问答 query", "mean_query_s", "sum_query_s"),
        ("检索 retrieve", "mean_retrieve_s", "sum_retrieve_s"),
        ("墙钟 wall", "mean_wall_s", "sum_wall_s"),
    ]
    for src, mk, sk in RETRIEVE_STAGE_KEYS:
        label = src.replace("_latency_s", "")
        stages.append((f"检索·{label}", mk, sk))
    for system in systems:
        lat = (summary.get(system) or {}).get("latency") or {}
        for label, mk, sk in stages:
            lat_rows.append(
                [_sys_label(system), label, _num(lat.get(mk)), _num(lat.get(sk))]
            )
    _write_table(
        ws,
        lat_headers,
        lat_rows,
        start_row=3,
        col_formats=[None, None, "float3", "float3"],
        freeze=False,
        autofilter=True,
        zebra=True,
    )

    tok_start = 5 + len(lat_rows)
    ws.cell(tok_start, 1, "Token").font = _FONT_BOLD
    tok_headers = [
        "系统",
        "prompt合计",
        "completion合计",
        "合计",
        "均prompt",
        "均completion",
        "评判prompt合计",
        "评判completion合计",
        "评判合计",
    ]
    tok_rows = []
    for system in systems:
        t = (summary.get(system) or {}).get("tokens") or {}
        tok_rows.append(
            [
                _sys_label(system),
                _int(t.get("sum_prompt")),
                _int(t.get("sum_completion")),
                _int(t.get("sum_total")),
                _num(t.get("mean_prompt")),
                _num(t.get("mean_completion")),
                _int(t.get("sum_judge_prompt")),
                _int(t.get("sum_judge_completion")),
                _int(t.get("sum_judge_total")),
            ]
        )
    _write_table(
        ws,
        tok_headers,
        tok_rows,
        start_row=tok_start + 1,
        col_formats=[None, "int", "int", "int", "float2", "float2", "int", "int", "int"],
        freeze=False,
        autofilter=False,
        zebra=False,
        max_width=18,
    )
    return name


def _result_system_block(result: Dict[str, Any], system: str) -> Dict[str, Any]:
    block = result.get(system)
    if isinstance(block, dict) and block:
        return block
    row = project_system_row(result, system)
    return {
        "answer": result.get("rag_answer")
        if system == SYSTEM_HYPERGRAPH
        else result.get("llm_only_answer"),
        "raw_answer": result.get("rag_raw_answer")
        if system == SYSTEM_HYPERGRAPH
        else None,
        **{
            k: row.get(k)
            for k in (
                "llm_acc",
                "query_status",
                "query_error",
                "judge_status",
                "judge_error",
                "metrics",
            )
        },
        "judge_reason": result.get("judge_reason")
        if system == SYSTEM_HYPERGRAPH
        else (block or {}).get("judge_reason"),
        "retrieval_sources": result.get("retrieval_sources")
        if system == SYSTEM_HYPERGRAPH
        else [],
    }


def _pair_winner_label(result: Dict[str, Any]) -> str:
    w = (result.get("comparison") or {}).get("winner")
    if w == SYSTEM_HYPERGRAPH:
        return "超图"
    if w == SYSTEM_LLM_ONLY:
        return "纯LLM"
    if w == "tie":
        return "平"
    hg = _result_system_block(result, SYSTEM_HYPERGRAPH).get("llm_acc")
    lo = _result_system_block(result, SYSTEM_LLM_ONLY).get("llm_acc")
    if hg == "正确" and lo != "正确":
        return "超图"
    if lo == "正确" and hg != "正确":
        return "纯LLM"
    if hg and lo:
        return "平"
    return ""


def _detail_headers(has_lo: bool) -> Tuple[List[str], List[Optional[str]]]:
    headers = [
        "ID",
        "跳数",
        "问题",
        "标准答案",
        "文档召回",
        "命中gold",
        "gold总数",
        "检索文档数",
        "gold文档",
        "超图评判",
        "超图query状态",
        "超图问答s",
        "超图检索s",
        "超图tokens",
        "超图检索来源数",
        "超图评判原因",
        "超图答案",
        "超图错误",
    ]
    fmts: List[Optional[str]] = [
        None,
        "int",
        None,
        None,
        "pct",
        "int",
        "int",
        "int",
        None,
        None,
        "int",
        "float3",
        "float3",
        "int",
        "int",
        None,
        None,
        None,
    ]
    if has_lo:
        headers += [
            "纯LLM评判",
            "纯LLM query状态",
            "纯LLM问答s",
            "纯LLM tokens",
            "纯LLM评判原因",
            "纯LLM答案",
            "纯LLM错误",
            "胜者",
            "对比理由",
            "评判是否一致",
        ]
        fmts += [None, "int", "float3", "int", None, None, None, None, None, None]
    return headers, fmts


def _detail_row(r: Dict[str, Any], has_lo: bool) -> List[Any]:
    hg = _result_system_block(r, SYSTEM_HYPERGRAPH)
    hm = hg.get("metrics") or {}
    rec = r.get("recall") if isinstance(r.get("recall"), dict) else {}
    hg_err = hg.get("query_error") or hg.get("judge_error")
    row: List[Any] = [
        r.get("id"),
        _int(r.get("hop")),
        r.get("question"),
        r.get("ground_truth_answer") or r.get("expected_answer"),
        _fmt_pct(rec.get("recall")),
        _int(rec.get("n_hit")),
        _int(rec.get("n_expected")),
        _int(rec.get("n_retrieved")),
        _join_names(r.get("source_names") or rec.get("expected_docs")),
        hg.get("llm_acc"),
        _int(hg.get("query_status")),
        _num(hm.get("query_latency_s")),
        _num(hm.get("retrieve_latency_s")),
        _int(hm.get("total_tokens")),
        len(hg.get("retrieval_sources") or []),
        hg.get("judge_reason"),
        hg.get("answer") or hg.get("raw_answer"),
        hg_err,
    ]
    if has_lo:
        lo = _result_system_block(r, SYSTEM_LLM_ONLY)
        lm = lo.get("metrics") or {}
        lo_err = lo.get("query_error") or lo.get("judge_error")
        hj, lj = hg.get("llm_acc"), lo.get("llm_acc")
        same = ""
        if hj or lj:
            same = "是" if hj == lj else "否"
        row += [
            lo.get("llm_acc"),
            _int(lo.get("query_status")),
            _num(lm.get("query_latency_s")),
            _int(lm.get("total_tokens")),
            lo.get("judge_reason"),
            lo.get("answer") or lo.get("raw_answer"),
            lo_err,
            _pair_winner_label(r),
            (r.get("comparison") or {}).get("reason") or "",
            same,
        ]
    return row


def _add_details(
    wb: Workbook,
    results: Sequence[Dict[str, Any]],
    systems: Sequence[str],
) -> Tuple[str, Optional[str], Optional[str]]:
    has_lo = SYSTEM_LLM_ONLY in systems
    headers, fmts = _detail_headers(has_lo)
    all_rows = [_detail_row(r, has_lo) for r in results]

    name = _sheet_name("逐题明细")
    ws = wb.create_sheet(name)
    _write_title(ws, "逐题明细（点选单元格可在编辑栏看全文）", len(headers))
    _write_table(
        ws,
        headers,
        all_rows,
        start_row=3,
        col_formats=fmts,
        freeze=True,
        autofilter=True,
        max_width=28,
        zebra=True,
    )
    ws.freeze_panes = "E4"
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 28

    diff_name = None
    if has_lo:
        winner_idx = headers.index("胜者") if "胜者" in headers else None
        same_idx = headers.index("评判是否一致") if "评判是否一致" in headers else None
        diff_rows = []
        for row in all_rows:
            win = row[winner_idx] if winner_idx is not None else ""
            same = row[same_idx] if same_idx is not None else ""
            if win in ("超图", "纯LLM") or same == "否":
                diff_rows.append(row)
        diff_name = _sheet_name("差异题")
        dws = wb.create_sheet(diff_name)
        _write_title(
            dws,
            f"评判不一致或有胜负的题（{len(diff_rows)} / {len(all_rows)}）",
            len(headers),
        )
        _write_table(
            dws,
            headers,
            diff_rows,
            start_row=3,
            col_formats=fmts,
            freeze=True,
            autofilter=True,
            max_width=28,
            zebra=True,
        )
        dws.freeze_panes = "E4"
        dws.column_dimensions["C"].width = 36
        if winner_idx is not None:
            c = winner_idx + 1
            for i in range(len(diff_rows)):
                cell = dws.cell(4 + i, c)
                if cell.value == "超图":
                    cell.fill = _FILL_POS
                elif cell.value == "纯LLM":
                    cell.fill = _FILL_NEG

    fail_rows: List[List[Any]] = []
    fail_headers = [
        "ID",
        "跳数",
        "系统",
        "问题",
        "query状态",
        "query错误",
        "judge状态",
        "judge错误",
    ]
    for r in results:
        for system in systems:
            b = _result_system_block(r, system)
            qerr = b.get("query_error")
            jerr = b.get("judge_error")
            qs, js = b.get("query_status"), b.get("judge_status")
            if qerr or jerr or qs == 0 or js == 0:
                fail_rows.append(
                    [
                        r.get("id"),
                        r.get("hop"),
                        _sys_label(system),
                        r.get("question"),
                        qs,
                        qerr,
                        js,
                        jerr,
                    ]
                )
    fail_name = _sheet_name("失败列表")
    fws = wb.create_sheet(fail_name)
    _write_title(
        fws,
        f"问答 / 评判失败（{len(fail_rows)} 条）" if fail_rows else "问答 / 评判失败（无）",
        len(fail_headers),
    )
    _write_table(
        fws,
        fail_headers,
        fail_rows,
        start_row=3,
        freeze=True,
        autofilter=bool(fail_rows),
        max_width=40,
    )
    return name, diff_name, fail_name


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------


def write_report_workbook(
    report: Dict[str, Any],
    out_path: Union[str, Path],
    *,
    eval_data: Optional[Dict[str, Any]] = None,
) -> Path:
    """把 report（及可选 evals）写成多 sheet 的 xlsx。"""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    summary = report.get("summary") or {}
    if eval_data is None and isinstance(report.get("results"), list):
        eval_data = report
    results = list((eval_data or {}).get("results") or [])
    systems = _systems(summary)

    wb = Workbook()
    default = wb.active
    default.title = "tmp"

    toc: List[Tuple[str, str]] = []
    toc.append((_add_meta(wb, report, eval_data), "评测配置、模型、题量、耗时"))
    toc.append((_add_overall(wb, summary), "两路总体正确率 / 耗时 / Token / 文档召回"))
    cmp_name = _add_comparison(wb, summary)
    if cmp_name:
        toc.append((cmp_name, "配对胜负、双方对错、评判交叉表"))
    hop_name = _add_by_hop(wb, summary)
    if hop_name:
        toc.append((hop_name, "1/2/3 跳正确率、耗时、文档召回"))
    rec_name = _add_doc_recall(wb, summary)
    if rec_name:
        toc.append((rec_name, "文档召回 × 答案正确率列联表"))
    toc.append((_add_latency_tokens(wb, summary), "检索各阶段耗时和 token 用量"))

    if results:
        d_name, diff_name, fail_name = _add_details(wb, results, systems)
        toc.append((d_name, "每题两路评判、答案、召回、原因"))
        if diff_name:
            toc.append((diff_name, "评判不一致或有胜负的题"))
        if fail_name:
            toc.append((fail_name, "问答或评判失败清单"))
    else:
        ws = wb.create_sheet(_sheet_name("逐题明细"))
        ws["A1"] = "未提供 evals JSON，没有逐题明细。重新导出时加上 --evals。"
        ws["A1"].font = _FONT_CELL
        toc.append((ws.title, "无逐题数据（缺少 evals）"))

    _add_toc(wb, toc)
    if "tmp" in wb.sheetnames:
        del wb["tmp"]

    wb.save(str(out))
    return out


def load_report_and_evals(
    report_path: Optional[Union[str, Path]] = None,
    evals_path: Optional[Union[str, Path]] = None,
) -> Tuple[Dict[str, Any], Optional[Dict[str, Any]]]:
    """读 JSON。若只给 evals，会现场生成 report。"""
    report: Optional[Dict[str, Any]] = None
    evals: Optional[Dict[str, Any]] = None

    def _load(p: Union[str, Path]) -> Dict[str, Any]:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)

    if report_path:
        data = _load(report_path)
        if isinstance(data.get("results"), list):
            evals = data
            report = QueryEvaluator.build_report_document(data, source_path=str(report_path))
        elif isinstance(data.get("summary"), dict):
            report = data
        else:
            raise ValueError(f"无法识别的 JSON（缺少 summary / results）: {report_path}")

    if evals_path:
        evals = _load(evals_path)
        if not isinstance(evals.get("results"), list):
            raise ValueError(f"非法评测结果（缺少 results）: {evals_path}")
        if report is None:
            report = QueryEvaluator.build_report_document(
                evals, source_path=str(evals_path)
            )

    if report is None:
        raise FileNotFoundError("需要 --report 或 --evals（或 config.yaml 里的对应路径）")
    return report, evals


def export_paths(
    *,
    report_path: Optional[Union[str, Path]] = None,
    evals_path: Optional[Union[str, Path]] = None,
    out_path: Optional[Union[str, Path]] = None,
) -> Path:
    report, evals = load_report_and_evals(report_path, evals_path)
    if out_path is None:
        src = Path(report_path or evals_path or "report.json")
        out_path = src.with_suffix(".xlsx")
    return write_report_workbook(report, out_path, eval_data=evals)


def main(argv: Optional[List[str]] = None) -> int:
    import argparse
    import sys

    root = Path(__file__).resolve().parent.parent
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))

    from benchmark.config import DEFAULT_CONFIG_PATH, BenchmarkConfig
    from benchmark.utils import resolve_path

    p = argparse.ArgumentParser(description="benchmark report JSON → 多表 Excel")
    p.add_argument("--config", default=DEFAULT_CONFIG_PATH, help="yaml 配置路径")
    p.add_argument("--report", default=None, help="report JSON；默认用 config 的 report 路径")
    p.add_argument("--evals", default=None, help="evals JSON，用于逐题明细；默认用 config")
    p.add_argument("--out", default=None, help="xlsx 写出路径")
    args = p.parse_args(argv)

    cfg = BenchmarkConfig.from_yaml(args.config).resolve_paths()
    report_p = Path(args.report) if args.report else cfg.report_file()
    if args.report and not report_p.is_absolute():
        report_p = resolve_path(report_p)
    evals_p = Path(args.evals) if args.evals else cfg.eval_results_file()
    if args.evals and not evals_p.is_absolute():
        evals_p = resolve_path(evals_p)
    out = Path(args.out) if args.out else cfg.report_excel_file()
    if args.out and not out.is_absolute():
        out = resolve_path(out)

    if not report_p.is_file() and not evals_p.is_file():
        print(
            "未找到 report / evals JSON，无法导出。\n"
            f"  report: {report_p}\n"
            f"  evals:  {evals_p}\n"
            "  请先 run.mode=report，或用 --report / --evals 指定文件。",
            file=sys.stderr,
        )
        return 2

    rp = report_p if report_p.is_file() else None
    ep = evals_p if evals_p.is_file() else None
    if rp is None and ep is not None:
        print(f"[excel] 无 report，改从 evals 生成: {ep}", file=sys.stderr)
    written = export_paths(report_path=rp, evals_path=ep, out_path=out)
    print(f"[saved] {written}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
