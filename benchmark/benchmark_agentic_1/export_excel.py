"""把 report / evals JSON 摊成多 sheet Excel。"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

from .report import STAGE_LABELS, build_report_document, project_system_row
from .utils import (
    RETRIEVE_STAGE_KEYS,
    SYSTEM_AGENTIC,
    SYSTEM_LLM_ONLY,
    ratio_to_percent,
)

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:  # pragma: no cover
    raise ImportError("写出 Excel 需要 openpyxl，请 pip install openpyxl") from e


SYSTEM_LABELS = {
    SYSTEM_AGENTIC: "Agentic",
    SYSTEM_LLM_ONLY: "纯LLM",
}
JUDGE_ORDER = ("正确", "错误", "未知")
RECALL_ORDER = ("hit", "partial", "miss", "unknown")
EXCEL_CELL_MAX = 32767
_INVALID_SHEET_RE = re.compile(r"[\[\]\*?:/\\]")

_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FILL_AG = PatternFill("solid", fgColor="D6EAF8")
_FILL_LO = PatternFill("solid", fgColor="FDEBD0")
_FILL_POS = PatternFill("solid", fgColor="D5F5E3")
_FILL_NEG = PatternFill("solid", fgColor="FADBD8")
_FILL_TIE = PatternFill("solid", fgColor="F4F6F7")
_FILL_ALT = PatternFill("solid", fgColor="F7F9FB")
_FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
_FONT_HEADER = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_FONT_TITLE = Font(name="微软雅黑", bold=True, color="FFFFFF", size=14)
_FONT_CELL = Font(name="微软雅黑", size=10)
_FONT_BOLD = Font(name="微软雅黑", bold=True, size=10)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_R = Alignment(horizontal="right", vertical="center")


def _sheet_name(name: str, used: Optional[set] = None) -> str:
    s = _INVALID_SHEET_RE.sub(" ", str(name or "Sheet")).strip() or "Sheet"
    s = s[:31]
    if used is None:
        return s
    base, n = s, 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = base[: 31 - len(suffix)] + suffix
    used.add(s)
    return s


def _cell_text(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v)
    if len(s) > EXCEL_CELL_MAX:
        s = s[: EXCEL_CELL_MAX - 1] + "…"
    return s


def _fmt_pct(v: Any) -> Optional[float]:
    return ratio_to_percent(v)


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _delta_fill(delta: Optional[float], *, higher_is_better: bool) -> Optional[PatternFill]:
    if delta is None:
        return None
    try:
        d = float(delta)
    except (TypeError, ValueError):
        return None
    if d == 0:
        return _FILL_TIE
    better = d > 0 if higher_is_better else d < 0
    return _FILL_POS if better else _FILL_NEG


def _systems(summary: Dict[str, Any]) -> List[str]:
    raw = summary.get("systems")
    out: List[str] = []
    if isinstance(raw, list):
        for k in raw:
            if k in SYSTEM_LABELS:
                out.append(k)
    if not out:
        for k in (SYSTEM_AGENTIC, SYSTEM_LLM_ONLY):
            block = summary.get(k)
            if isinstance(block, dict) and isinstance(block.get("llm_acc"), dict):
                out.append(k)
    return out or [SYSTEM_AGENTIC]


def _sys_label(key: str) -> str:
    return SYSTEM_LABELS.get(key, key)


def _nested(d: Optional[dict], *keys: str, default: Any = None) -> Any:
    cur: Any = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur


def _hop_keys(by_hop: Dict[str, Any]) -> List[str]:
    def key(h: str):
        try:
            return (0, int(h))
        except (TypeError, ValueError):
            return (1, str(h))

    return sorted((by_hop or {}).keys(), key=key)


def _auto_width(ws: Worksheet, *, min_w: float = 8, max_w: float = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0.0
        for cell in col:
            if cell.value is None:
                continue
            text = str(cell.value).replace("\n", " ")
            w = 0.0
            for ch in text[:80]:
                w += 1.7 if ord(ch) > 127 else 1.05
            longest = max(longest, w)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 1.5))


def _apply_number_format(cell, fmt: Optional[str]) -> None:
    if fmt == "pct":
        cell.number_format = '0.00"%"'
    elif fmt == "int":
        cell.number_format = "#,##0"
    elif fmt == "float2":
        cell.number_format = "0.00"
    elif fmt == "float3":
        cell.number_format = "0.000"


def _write_title(ws: Worksheet, title: str, n_cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, n_cols))
    cell = ws.cell(1, 1, title)
    cell.fill = _FILL_TITLE
    cell.font = _FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def _write_table(
    ws: Worksheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    start_row: int = 1,
    col_formats: Optional[Sequence[Optional[str]]] = None,
    freeze: bool = True,
    autofilter: bool = True,
    zebra: bool = True,
    max_width: float = 48,
) -> int:
    n_cols = len(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _THIN
    ws.row_dimensions[start_row].height = 22

    for i, row in enumerate(rows):
        r = start_row + 1 + i
        alt = zebra and (i % 2 == 1)
        for c in range(1, n_cols + 1):
            raw = row[c - 1] if c - 1 < len(row) else None
            val = _cell_text(raw)
            cell = ws.cell(r, c, val)
            cell.font = _FONT_CELL
            cell.border = _THIN
            fmt = col_formats[c - 1] if col_formats and c - 1 < len(col_formats) else None
            _apply_number_format(cell, fmt)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = _ALIGN_R
            else:
                cell.alignment = _ALIGN_L
            if alt:
                cell.fill = _FILL_ALT

    last = start_row + len(rows)
    if autofilter and rows:
        ws.auto_filter.ref = (
            f"{get_column_letter(1)}{start_row}:{get_column_letter(n_cols)}{last}"
        )
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"
    _auto_width(ws, max_w=max_width)
    ws.sheet_view.showGridLines = False
    return last + 1


def _flatten_meta_value(v: Any) -> Any:
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    return json.dumps(v, ensure_ascii=False)


def _sys_fill(system: str) -> PatternFill:
    if system == SYSTEM_AGENTIC:
        return _FILL_AG
    if system == SYSTEM_LLM_ONLY:
        return _FILL_LO
    return _FILL_ALT


def _join_names(names: Any) -> Optional[str]:
    if not names:
        return None
    if isinstance(names, str):
        return names
    if isinstance(names, (list, tuple)):
        return " / ".join(str(x) for x in names if x)
    return str(names)


def _add_toc(wb: Workbook, items: Sequence[Tuple[str, str]]) -> None:
    ws = wb.create_sheet(_sheet_name("目录"), 0)
    _write_title(ws, "benchmark_agentic_1 评测报告", 2)
    ws.cell(2, 1, "点击工作表名跳转。")
    ws.cell(2, 1).font = Font(name="微软雅黑", italic=True, size=10, color="666666")
    headers = ("工作表", "说明")
    rows = [(name, desc) for name, desc in items]
    _write_table(ws, headers, rows, start_row=4, freeze=False, autofilter=False, zebra=True)
    for i, (sheet, _) in enumerate(items):
        cell = ws.cell(5 + i, 1, sheet)
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate, location=f"'{sheet}'!A1", display=sheet
        )
        cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        cell.alignment = _ALIGN_L
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 72


def _add_meta(
    wb: Workbook, report: Dict[str, Any], eval_data: Optional[Dict[str, Any]]
) -> str:
    name = _sheet_name("运行信息")
    ws = wb.create_sheet(name)
    meta = dict(report.get("meta") or {})
    ev_meta = (eval_data or {}).get("meta") if isinstance(eval_data, dict) else {}
    if isinstance(ev_meta, dict):
        for k, v in ev_meta.items():
            meta.setdefault(k, v)
    skip = {"config", "dataset_meta", "progress"}
    rows: List[List[Any]] = []
    preferred = [
        "created_at", "updated_at", "eval_done", "done", "query_mode",
        "judge_model", "judge_mode", "enable_llm_only", "enable_doc_recall",
        "n_results", "n_questions", "n_total_planned", "n_skipped_gen_fail",
        "eval_elapsed_s", "elapsed_s", "db_path", "hop_counts", "seed",
        "doc_id_min", "doc_id_max", "source_path", "config_file", "dhmf_config_path",
    ]
    seen = set()
    for k in preferred:
        if k in meta:
            rows.append([k, _flatten_meta_value(meta[k])])
            seen.add(k)
    for k in sorted(meta.keys()):
        if k in seen or k in skip:
            continue
        rows.append([k, _flatten_meta_value(meta[k])])
    cfg = meta.get("config") if isinstance(meta.get("config"), dict) else None
    if cfg:
        rows.append(["— 配置快照 —", ""])
        for k, v in cfg.items():
            rows.append([f"config.{k}", _flatten_meta_value(v)])
    _write_title(ws, "运行信息", 2)
    _write_table(
        ws, ("字段", "值"), rows, start_row=3, freeze=True, autofilter=False, max_width=80,
    )
    return name


def _metric_row_overall(
    sys_blocks: Dict[str, Dict[str, Any]], systems: Sequence[str]
) -> Tuple[List[List[Any]], List[Tuple[str, str, Tuple[str, ...], Optional[bool]]], bool]:
    def g(sys: str, *path, default=None):
        return _nested(sys_blocks.get(sys), *path, default=default)

    specs: List[Tuple[str, str, Tuple[str, ...], Optional[bool]]] = [
        ("题目数 n", "int", ("n_total",), None),
        ("流水线成功", "int", ("pipeline", "n_ok"), True),
        ("问答失败", "int", ("pipeline", "n_query_fail"), False),
        ("评判失败", "int", ("pipeline", "n_judge_fail"), False),
        ("已评判", "int", ("llm_acc", "n_judged"), True),
        ("正确", "int", ("llm_acc", "n_correct"), True),
        ("错误", "int", ("llm_acc", "n_wrong"), False),
        ("未知", "int", ("llm_acc", "counts", "未知"), None),
        ("正确率", "pct", ("llm_acc", "accuracy"), True),
        ("错误率", "pct", ("llm_acc", "error_rate"), False),
        ("均文档召回", "pct", ("doc_recall", "mean_recall"), True),
        ("每问平均延迟 s", "float3", ("latency", "mean_query_s"), False),
        ("每问平均检索合计 s", "float3", ("latency", "mean_retrieve_s"), False),
        ("每问平均墙钟 s", "float3", ("latency", "mean_wall_s"), False),
        ("单次 search 平均 s", "float3", ("latency", "mean_search_s"), False),
        ("每问平均 search 次数", "float2", ("latency", "mean_n_search"), None),
        ("每问平均 turns", "float2", ("latency", "mean_n_turns"), None),
        ("预加载 每问平均 s", "float3", ("latency", "mean_precompute_s"), False),
        ("改写 每问平均 s", "float3", ("latency", "mean_rewrite_s"), False),
        ("chunk路 每问平均 s", "float3", ("latency", "mean_chunk_s"), False),
        ("node路 每问平均 s", "float3", ("latency", "mean_node_s"), False),
        ("keyword路 每问平均 s", "float3", ("latency", "mean_keyword_s"), False),
        ("重排 每问平均 s", "float3", ("latency", "mean_rerank_s"), False),
        ("chunk路 每次search平均 s", "float3", ("latency", "mean_chunk_s_per_search"), False),
        ("node路 每次search平均 s", "float3", ("latency", "mean_node_s_per_search"), False),
        ("keyword路 每次search平均 s", "float3", ("latency", "mean_keyword_s_per_search"), False),
        ("问答总耗时 s", "float3", ("latency", "sum_query_s"), False),
        ("检索总耗时 s", "float3", ("latency", "sum_retrieve_s"), False),
        ("prompt tokens 合计", "int", ("tokens", "sum_prompt"), None),
        ("completion tokens 合计", "int", ("tokens", "sum_completion"), None),
        ("tokens 合计", "int", ("tokens", "sum_total"), None),
        ("均 prompt tokens", "float2", ("tokens", "mean_prompt"), None),
        ("均 completion tokens", "float2", ("tokens", "mean_completion"), None),
    ]

    has_pair = len(systems) >= 2
    rows: List[List[Any]] = []
    for name, fmt, path, better in specs:
        vals = [g(s, *path) for s in systems]
        if fmt == "pct":
            vals = [_fmt_pct(v) for v in vals]
        elif fmt == "int":
            vals = [_int(v) if v is not None else None for v in vals]
        else:
            vals = [_num(v) if v is not None else None for v in vals]
        delta = None
        if has_pair and vals[0] is not None and vals[1] is not None:
            try:
                delta = round(float(vals[0]) - float(vals[1]), 6)
            except (TypeError, ValueError):
                delta = None
        row: List[Any] = [name, *vals]
        if has_pair:
            row.append(delta)
        rows.append(row)
    return rows, specs, has_pair


def _add_overall(wb: Workbook, summary: Dict[str, Any]) -> str:
    sheet_name = _sheet_name("总体指标")
    ws = wb.create_sheet(sheet_name)
    systems = _systems(summary)
    sys_blocks = {s: summary.get(s) or {} for s in systems}
    rows, specs, has_pair = _metric_row_overall(sys_blocks, systems)

    headers = ["指标"] + [_sys_label(s) for s in systems]
    col_formats: List[Optional[str]] = [None] + [None] * len(systems)
    if has_pair:
        headers.append("差值（Agentic − 纯LLM）")
        col_formats.append("float3")

    n_cols = len(headers)
    _write_title(ws, "总体指标：Agentic vs 纯LLM" if has_pair else "总体指标", n_cols)
    start = 3
    _write_table(
        ws, headers, rows, start_row=start, col_formats=col_formats,
        freeze=True, autofilter=False, zebra=True,
    )
    for i, (_metric, fmt, _path, better) in enumerate(specs):
        r = start + 1 + i
        for c, s in enumerate(systems, 2):
            cell = ws.cell(r, c)
            _apply_number_format(cell, fmt)
            cell.fill = _sys_fill(s)
        if has_pair:
            dcell = ws.cell(r, 2 + len(systems))
            _apply_number_format(dcell, fmt)
            if better is not None:
                fill = _delta_fill(dcell.value, higher_is_better=bool(better))
                if fill is not None:
                    dcell.fill = fill
    return sheet_name


def _parse_crosstab(xt: Dict[str, Any]) -> Dict[Tuple[str, str], int]:
    matrix: Dict[Tuple[str, str], int] = {}
    for k, v in (xt or {}).items():
        key = str(k)
        if not key.startswith("Agentic") or "_纯LLM" not in key:
            continue
        rest = key[len("Agentic"):]
        hj, lj = rest.split("_纯LLM", 1)
        try:
            matrix[(hj, lj)] = int(v)
        except (TypeError, ValueError):
            continue
    return matrix


def _add_comparison(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    cmp_ = summary.get("comparison")
    if not isinstance(cmp_, dict) or not cmp_:
        return None
    name = _sheet_name("系统对比")
    ws = wb.create_sheet(name)
    _write_title(ws, "Agentic vs 纯LLM 对比", 5)

    win = cmp_.get("win") or {}
    n_paired = _int(cmp_.get("n_paired")) or 0
    n_ag = _int(win.get("agentic")) or 0
    n_lo = _int(win.get("llm_only")) or 0
    n_tie = _int(win.get("tie")) or 0

    overview = [
        ["配对题数", n_paired],
        ["对比打分题数", _int(cmp_.get("n_pairwise_judged"))],
        ["双方都正确", _int(cmp_.get("both_correct"))],
        ["仅 Agentic 正确", _int(cmp_.get("agentic_only"))],
        ["仅纯LLM正确", _int(cmp_.get("llm_only_only"))],
        ["双方都错误", _int(cmp_.get("both_wrong"))],
        ["胜负来源", cmp_.get("win_note") or ""],
        ["Agentic 胜", n_ag],
        ["纯LLM胜", n_lo],
        ["平局", n_tie],
        ["Agentic 胜率", (n_ag / n_paired) if n_paired else None],
        ["纯LLM胜率", (n_lo / n_paired) if n_paired else None],
        ["平局率", (n_tie / n_paired) if n_paired else None],
    ]
    _write_table(ws, ("项目", "值"), overview, start_row=3, freeze=False, autofilter=False)
    for r, row in enumerate(overview, 4):
        label = str(row[0])
        if "率" in label:
            _apply_number_format(ws.cell(r, 2), "pct")
        elif "Agentic 胜" in label or "仅 Agentic" in label:
            ws.cell(r, 2).fill = _FILL_POS
        elif "纯LLM胜" in label or "仅纯LLM" in label:
            ws.cell(r, 2).fill = _FILL_NEG
        elif label in ("平局", "双方都正确", "双方都错误"):
            ws.cell(r, 2).fill = _FILL_TIE

    bar_row = 18
    ws.cell(bar_row, 1, "胜负").font = _FONT_BOLD
    _write_table(
        ws, ("结果", "题数"),
        [["Agentic 胜", n_ag], ["纯LLM胜", n_lo], ["平局", n_tie]],
        start_row=bar_row + 1, col_formats=[None, "int"],
        freeze=False, autofilter=False, zebra=False,
    )
    chart = BarChart()
    chart.type = "col"
    chart.title = "对比胜负"
    chart.y_axis.title = "题数"
    data_ref = Reference(ws, min_col=2, min_row=bar_row + 1, max_row=bar_row + 4)
    cats = Reference(ws, min_col=1, min_row=bar_row + 2, max_row=bar_row + 4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.width = 12
    chart.height = 7
    ws.add_chart(chart, "D19")

    xt_start = 25
    ws.cell(xt_start, 1, "评判交叉表（行=Agentic，列=纯LLM）").font = _FONT_BOLD
    matrix = _parse_crosstab(cmp_.get("judgment_crosstab") or {})
    labels = list(JUDGE_ORDER)
    xt_headers = ["Agentic \\ 纯LLM"] + [f"纯LLM·{x}" for x in labels] + ["行合计"]
    xt_rows = []
    col_tot = {lj: 0 for lj in labels}
    for hj in labels:
        row = [f"Agentic·{hj}"]
        s = 0
        for lj in labels:
            v = matrix.get((hj, lj), 0)
            row.append(v)
            s += v
            col_tot[lj] += v
        row.append(s)
        xt_rows.append(row)
    xt_rows.append(["列合计"] + [col_tot[lj] for lj in labels] + [sum(col_tot.values())])
    _write_table(
        ws, xt_headers, xt_rows, start_row=xt_start + 1,
        col_formats=[None] + ["int"] * (len(labels) + 1),
        freeze=False, autofilter=False, zebra=False,
    )
    return name


def _add_by_hop(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    systems = _systems(summary)
    hop_maps: Dict[str, Dict[str, Any]] = {}
    hops: List[str] = []
    seen = set()
    for s in systems:
        bh = (summary.get(s) or {}).get("by_hop") or {}
        hop_maps[s] = bh if isinstance(bh, dict) else {}
        for h in _hop_keys(hop_maps[s]):
            if h not in seen:
                hops.append(h)
                seen.add(h)
    if not hops:
        return None

    name = _sheet_name("按跳数")
    ws = wb.create_sheet(name)
    has_pair = len(systems) >= 2
    headers = [
        "跳数", "n",
        "正确率_Agentic", "正确率_纯LLM", "正确率差",
        "正确_Agentic", "正确_纯LLM", "错误_Agentic", "错误_纯LLM",
        "每问平均s_Agentic", "每问平均s_纯LLM",
        "每问检索合计s", "单次search平均s",
        "chunk路每问s", "node路每问s", "keyword路每问s",
        "改写每问s", "重排每问s", "预加载每问s",
        "均文档召回",
    ]
    fmts: List[Optional[str]] = [
        None, "int", "pct", "pct", "pct",
        "int", "int", "int", "int",
        "float3", "float3",
        "float3", "float3",
        "float3", "float3", "float3",
        "float3", "float3", "float3",
        "pct",
    ]
    rows: List[List[Any]] = []
    chart_rows: List[List[Any]] = []
    for hop in hops:
        ag = hop_maps.get(SYSTEM_AGENTIC, {}).get(hop) or {}
        lo = hop_maps.get(SYSTEM_LLM_ONLY, {}).get(hop) or {}
        ag_acc = ag.get("llm_acc_counts") or {}
        lo_acc = lo.get("llm_acc_counts") or {}
        n = _int(ag.get("n")) or _int(lo.get("n"))
        acc_a = _fmt_pct(ag.get("accuracy"))
        acc_l = _fmt_pct(lo.get("accuracy"))
        delta = None
        if acc_a is not None and acc_l is not None:
            delta = round(float(acc_a) - float(acc_l), 6)
        if has_pair:
            rows.append([
                hop, n, acc_a, acc_l, delta,
                _int(ag_acc.get("正确")), _int(lo_acc.get("正确")),
                _int(ag_acc.get("错误")), _int(lo_acc.get("错误")),
                _num(ag.get("mean_query_latency_s")),
                _num(lo.get("mean_query_latency_s")),
                _num(ag.get("mean_retrieve_latency_s")),
                _num(ag.get("mean_search_s")),
                _num(ag.get("mean_chunk_s")),
                _num(ag.get("mean_node_s")),
                _num(ag.get("mean_keyword_s")),
                _num(ag.get("mean_rewrite_s")),
                _num(ag.get("mean_rerank_s")),
                _num(ag.get("mean_precompute_s")),
                _fmt_pct(ag.get("mean_doc_recall")),
            ])
        else:
            src = ag or lo
            counts = src.get("llm_acc_counts") or {}
            rows.append([
                hop, n,
                _fmt_pct(src.get("accuracy")), None, None,
                _int(counts.get("正确")), None,
                _int(counts.get("错误")), None,
                _num(src.get("mean_query_latency_s")), None,
                _num(src.get("mean_retrieve_latency_s")),
                _num(src.get("mean_search_s")),
                _num(src.get("mean_chunk_s")),
                _num(src.get("mean_node_s")),
                _num(src.get("mean_keyword_s")),
                _num(src.get("mean_rewrite_s")),
                _num(src.get("mean_rerank_s")),
                _num(src.get("mean_precompute_s")),
                _fmt_pct(src.get("mean_doc_recall")),
            ])
        chart_rows.append(
            [hop]
            + [
                _fmt_pct((hop_maps.get(s) or {}).get(hop, {}).get("accuracy")) or 0
                for s in systems
            ]
        )

    _write_title(ws, "按跳数切分：Agentic vs 纯LLM" if has_pair else "按跳数切分", len(headers))
    start = 3
    _write_table(
        ws, headers, rows, start_row=start, col_formats=fmts,
        freeze=True, autofilter=True, max_width=16, zebra=True,
    )
    if has_pair:
        for i in range(len(rows)):
            r = start + 1 + i
            fill = _delta_fill(ws.cell(r, 5).value, higher_is_better=True)
            if fill is not None:
                ws.cell(r, 5).fill = fill
            ws.cell(r, 3).fill = _FILL_AG
            ws.cell(r, 4).fill = _FILL_LO

    chart_start = start + len(rows) + 3
    ws.cell(chart_start, 1, "各跳正确率").font = _FONT_BOLD
    chart_headers = ["跳数"] + [_sys_label(s) for s in systems]
    _write_table(
        ws, chart_headers, chart_rows, start_row=chart_start + 1,
        col_formats=[None] + ["pct"] * len(systems),
        freeze=False, autofilter=False, zebra=False,
    )
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "各跳正确率"
    chart.y_axis.title = "正确率"
    chart.x_axis.title = "跳数"
    data_ref = Reference(
        ws, min_col=2, min_row=chart_start + 1,
        max_col=1 + len(systems),
        max_row=chart_start + 1 + len(chart_rows),
    )
    cats = Reference(
        ws, min_col=1, min_row=chart_start + 2,
        max_row=chart_start + 1 + len(chart_rows),
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, "F" + str(chart_start))
    return name


def _add_doc_recall(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    ag = summary.get(SYSTEM_AGENTIC) or summary
    rec = (ag or {}).get("doc_recall")
    if not isinstance(rec, dict) or not rec:
        rec = summary.get("doc_recall")
    if not isinstance(rec, dict) or not rec:
        return None

    name = _sheet_name("文档召回")
    ws = wb.create_sheet(name)
    _write_title(ws, "文档召回 × 答案准确率", 6)
    vs = rec.get("vs_accuracy") if isinstance(rec.get("vs_accuracy"), dict) else {}
    overview = [
        ["均文档召回", _fmt_pct(rec.get("mean_recall"))],
        ["有召回字段题数", _int(vs.get("n_with_recall"))],
        ["总题数", _int(vs.get("n_total") or summary.get("n_total"))],
    ]
    _write_table(ws, ("项目", "值"), overview, start_row=3, freeze=False, autofilter=False)
    _apply_number_format(ws.cell(4, 2), "pct")

    definition = vs.get("definition") if isinstance(vs.get("definition"), dict) else {}
    def_start = 8
    ws.cell(def_start, 1, "口径").font = _FONT_BOLD
    def_rows = [[k, v] for k, v in definition.items()]
    if def_rows:
        _write_table(
            ws, ("桶", "说明"), def_rows, start_row=def_start + 1,
            freeze=False, autofilter=False, max_width=80,
        )

    matrix_start = def_start + 2 + max(len(def_rows), 1) + 1
    ws.cell(matrix_start, 1, "召回 × 准确率").font = _FONT_BOLD
    cells: Dict[Tuple[str, str], int] = {}
    for item in vs.get("matrix") or []:
        if not isinstance(item, dict):
            continue
        cells[(str(item.get("recall")), str(item.get("accuracy")))] = _int(item.get("n")) or 0
    table = vs.get("table") if isinstance(vs.get("table"), dict) else {}
    alias = {"正确": "correct", "错误": "wrong", "未知": "unknown"}
    for b in RECALL_ORDER:
        for lab in JUDGE_ORDER:
            if (b, lab) in cells:
                continue
            key = f"{b}_{alias.get(lab, lab)}"
            if key in table:
                cells[(b, lab)] = _int(table.get(key)) or 0

    row_tot = vs.get("row_totals") if isinstance(vs.get("row_totals"), dict) else {}
    col_tot = vs.get("col_totals") if isinstance(vs.get("col_totals"), dict) else {}
    mx_headers = ["召回 \\ 准确率"] + list(JUDGE_ORDER) + ["行合计"]
    mx_rows = []
    for b in RECALL_ORDER:
        row = [b]
        s = 0
        for lab in JUDGE_ORDER:
            v = cells.get((b, lab), 0)
            row.append(v)
            s += v
        row.append(_int(row_tot.get(b)) if row_tot.get(b) is not None else s)
        mx_rows.append(row)
    mx_rows.append(
        ["列合计"]
        + [
            _int(col_tot.get(lab))
            if col_tot.get(lab) is not None
            else sum(cells.get((b, lab), 0) for b in RECALL_ORDER)
            for lab in JUDGE_ORDER
        ]
        + [_int(vs.get("n_total")) or sum(_int(r[-1]) or 0 for r in mx_rows)]
    )
    _write_table(
        ws, mx_headers, mx_rows, start_row=matrix_start + 1,
        col_formats=[None] + ["int"] * (len(JUDGE_ORDER) + 1),
        freeze=False, autofilter=False, zebra=False,
    )
    return name


def _add_latency(wb: Workbook, summary: Dict[str, Any]) -> str:
    name = _sheet_name("延迟明细")
    ws = wb.create_sheet(name)
    _write_title(ws, "检索分阶段延迟（每问合计 / 每次 search 平均）", 6)
    systems = _systems(summary)

    lat_headers = ["系统", "阶段", "每问平均 s", "合计 s", "每次search平均 s"]
    lat_rows: List[List[Any]] = []
    top = [
        ("每问总延迟", "mean_query_s", "sum_query_s", None),
        ("每问检索合计", "mean_retrieve_s", "sum_retrieve_s", "mean_search_s"),
        ("每问墙钟", "mean_wall_s", "sum_wall_s", None),
    ]
    for system in systems:
        lat = (summary.get(system) or {}).get("latency") or {}
        for label, mk, sk, per in top:
            lat_rows.append([
                _sys_label(system), label,
                _num(lat.get(mk)), _num(lat.get(sk)),
                _num(lat.get(per)) if per else None,
            ])
        if system != SYSTEM_AGENTIC:
            continue
        stages = lat.get("stages") if isinstance(lat.get("stages"), dict) else {}
        for k in RETRIEVE_STAGE_KEYS:
            st = stages.get(k) if isinstance(stages.get(k), dict) else {}
            label = st.get("label") or STAGE_LABELS.get(k, k)
            lat_rows.append([
                _sys_label(system),
                label,
                _num(st.get("mean_per_question_s") or lat.get(f"mean_{k}")),
                _num(st.get("sum_s") or lat.get(f"sum_{k}")),
                _num(st.get("mean_per_search_s") or lat.get(f"mean_{k}_per_search")),
            ])
        lat_rows.append([
            _sys_label(system), "每问 search 次数",
            _num(lat.get("mean_n_search")), _int(lat.get("sum_n_search")), None,
        ])
        lat_rows.append([
            _sys_label(system), "每问 turns",
            _num(lat.get("mean_n_turns")), _int(lat.get("sum_n_turns")), None,
        ])

    _write_table(
        ws, lat_headers, lat_rows, start_row=3,
        col_formats=[None, None, "float3", "float3", "float3"],
        freeze=False, autofilter=True, zebra=True,
    )

    tok_start = 5 + len(lat_rows)
    ws.cell(tok_start, 1, "Token").font = _FONT_BOLD
    tok_headers = [
        "系统", "prompt合计", "completion合计", "合计",
        "均prompt", "均completion",
        "评判prompt合计", "评判completion合计", "评判合计",
    ]
    tok_rows = []
    for system in systems:
        t = (summary.get(system) or {}).get("tokens") or {}
        tok_rows.append([
            _sys_label(system),
            _int(t.get("sum_prompt")), _int(t.get("sum_completion")),
            _int(t.get("sum_total")),
            _num(t.get("mean_prompt")), _num(t.get("mean_completion")),
            _int(t.get("sum_judge_prompt")),
            _int(t.get("sum_judge_completion")),
            _int(t.get("sum_judge_total")),
        ])
    _write_table(
        ws, tok_headers, tok_rows, start_row=tok_start + 1,
        col_formats=[None, "int", "int", "int", "float2", "float2", "int", "int", "int"],
        freeze=False, autofilter=False, zebra=False, max_width=18,
    )
    return name


def _result_system_block(result: Dict[str, Any], system: str) -> Dict[str, Any]:
    block = result.get(system)
    if isinstance(block, dict) and block:
        return block
    row = project_system_row(result, system)
    return {
        "answer": None,
        "raw_answer": None,
        **{k: row.get(k) for k in (
            "llm_acc", "query_status", "query_error",
            "judge_status", "judge_error", "metrics",
        )},
        "judge_reason": (block or {}).get("judge_reason") if isinstance(block, dict) else "",
        "retrieval_sources": [],
    }


def _pair_winner_label(result: Dict[str, Any]) -> str:
    w = (result.get("comparison") or {}).get("winner")
    if w == SYSTEM_AGENTIC:
        return "Agentic"
    if w == SYSTEM_LLM_ONLY:
        return "纯LLM"
    if w == "tie":
        return "平"
    ag = _result_system_block(result, SYSTEM_AGENTIC).get("llm_acc")
    lo = _result_system_block(result, SYSTEM_LLM_ONLY).get("llm_acc")
    if ag == "正确" and lo != "正确":
        return "Agentic"
    if lo == "正确" and ag != "正确":
        return "纯LLM"
    if ag and lo:
        return "平"
    return ""


def _detail_headers(has_lo: bool) -> Tuple[List[str], List[Optional[str]]]:
    headers = [
        "ID", "跳数", "问题", "标准答案",
        "文档召回", "命中gold", "gold总数", "检索文档数", "gold文档",
        "Agentic评判", "Agentic query状态",
        "Agentic每问s", "检索合计s", "单次search平均s",
        "n_search", "n_turns",
        "预加载s", "改写s", "chunk路s", "node路s", "keyword路s", "重排s",
        "Agentic tokens", "Agentic检索来源数",
        "Agentic评判原因", "Agentic答案", "Agentic错误",
    ]
    fmts: List[Optional[str]] = [
        None, "int", None, None,
        "pct", "int", "int", "int", None,
        None, "int",
        "float3", "float3", "float3",
        "int", "int",
        "float3", "float3", "float3", "float3", "float3", "float3",
        "int", "int",
        None, None, None,
    ]
    if has_lo:
        headers += [
            "纯LLM评判", "纯LLM query状态", "纯LLM问答s", "纯LLM tokens",
            "纯LLM评判原因", "纯LLM答案", "纯LLM错误",
            "胜者", "对比理由", "评判是否一致",
        ]
        fmts += [None, "int", "float3", "int", None, None, None, None, None, None]
    return headers, fmts


def _detail_row(r: Dict[str, Any], has_lo: bool) -> List[Any]:
    ag = _result_system_block(r, SYSTEM_AGENTIC)
    hm = ag.get("metrics") or {}
    rec = r.get("recall") if isinstance(r.get("recall"), dict) else {}
    ag_err = ag.get("query_error") or ag.get("judge_error")
    row: List[Any] = [
        r.get("id"),
        _int(r.get("hop")),
        r.get("question"),
        r.get("ground_truth_answer") or r.get("expected_answer"),
        _fmt_pct(rec.get("recall")),
        _int(rec.get("n_hit")),
        _int(rec.get("n_expected")),
        _int(rec.get("n_retrieved")),
        _join_names(r.get("source_names") or rec.get("expected_docs")),
        ag.get("llm_acc"),
        _int(ag.get("query_status")),
        _num(hm.get("query_latency_s")),
        _num(hm.get("retrieve_latency_s")),
        _num(hm.get("mean_search_s")),
        _int(hm.get("n_search")),
        _int(hm.get("n_turns")),
        _num(hm.get("precompute_s") or hm.get("precompute_latency_s")),
        _num(hm.get("rewrite_s") or hm.get("rewrite_latency_s")),
        _num(hm.get("chunk_s") or hm.get("chunk_latency_s")),
        _num(hm.get("node_s") or hm.get("node_latency_s")),
        _num(hm.get("keyword_s") or hm.get("keyword_latency_s")),
        _num(hm.get("rerank_s") or hm.get("rerank_latency_s")),
        _int(hm.get("total_tokens")),
        len(ag.get("retrieval_sources") or []),
        ag.get("judge_reason"),
        ag.get("answer") or ag.get("raw_answer"),
        ag_err,
    ]
    if has_lo:
        lo = _result_system_block(r, SYSTEM_LLM_ONLY)
        lm = lo.get("metrics") or {}
        lo_err = lo.get("query_error") or lo.get("judge_error")
        hj, lj = ag.get("llm_acc"), lo.get("llm_acc")
        same = ""
        if hj or lj:
            same = "是" if hj == lj else "否"
        row += [
            lo.get("llm_acc"),
            _int(lo.get("query_status")),
            _num(lm.get("query_latency_s")),
            _int(lm.get("total_tokens")),
            lo.get("judge_reason"),
            lo.get("answer") or lo.get("raw_answer"),
            lo_err,
            _pair_winner_label(r),
            (r.get("comparison") or {}).get("reason") or "",
            same,
        ]
    return row


def _add_details(
    wb: Workbook,
    results: Sequence[Dict[str, Any]],
    systems: Sequence[str],
) -> Tuple[str, Optional[str], Optional[str]]:
    has_lo = SYSTEM_LLM_ONLY in systems
    headers, fmts = _detail_headers(has_lo)
    all_rows = [_detail_row(r, has_lo) for r in results]

    name = _sheet_name("逐题明细")
    ws = wb.create_sheet(name)
    _write_title(ws, "逐题明细（点选单元格可在编辑栏看全文）", len(headers))
    _write_table(
        ws, headers, all_rows, start_row=3, col_formats=fmts,
        freeze=True, autofilter=True, max_width=22, zebra=True,
    )
    ws.freeze_panes = "E4"
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 28

    diff_name = None
    if has_lo:
        winner_idx = headers.index("胜者") if "胜者" in headers else None
        same_idx = headers.index("评判是否一致") if "评判是否一致" in headers else None
        diff_rows = []
        for row in all_rows:
            win = row[winner_idx] if winner_idx is not None else ""
            same = row[same_idx] if same_idx is not None else ""
            if win in ("Agentic", "纯LLM") or same == "否":
                diff_rows.append(row)
        diff_name = _sheet_name("差异题")
        dws = wb.create_sheet(diff_name)
        _write_title(
            dws,
            f"评判不一致或有胜负的题（{len(diff_rows)} / {len(all_rows)}）",
            len(headers),
        )
        _write_table(
            dws, headers, diff_rows, start_row=3, col_formats=fmts,
            freeze=True, autofilter=True, max_width=22, zebra=True,
        )
        dws.freeze_panes = "E4"
        dws.column_dimensions["C"].width = 36
        if winner_idx is not None:
            c = winner_idx + 1
            for i in range(len(diff_rows)):
                cell = dws.cell(4 + i, c)
                if cell.value == "Agentic":
                    cell.fill = _FILL_POS
                elif cell.value == "纯LLM":
                    cell.fill = _FILL_NEG

    fail_rows: List[List[Any]] = []
    fail_headers = [
        "ID", "跳数", "系统", "问题", "query状态", "query错误", "judge状态", "judge错误",
    ]
    for r in results:
        for system in systems:
            b = _result_system_block(r, system)
            qerr = b.get("query_error")
            jerr = b.get("judge_error")
            qs, js = b.get("query_status"), b.get("judge_status")
            if qerr or jerr or qs == 0 or js == 0:
                fail_rows.append([
                    r.get("id"), r.get("hop"), _sys_label(system),
                    r.get("question"), qs, qerr, js, jerr,
                ])
    fail_name = _sheet_name("失败列表")
    fws = wb.create_sheet(fail_name)
    _write_title(
        fws,
        f"问答 / 评判失败（{len(fail_rows)} 条）" if fail_rows else "问答 / 评判失败（无）",
        len(fail_headers),
    )
    _write_table(
        fws, fail_headers, fail_rows, start_row=3,
        freeze=True, autofilter=bool(fail_rows), max_width=40,
    )
    return name, diff_name, fail_name


def write_report_workbook(
    report: Dict[str, Any],
    out_path: Union[str, Path],
    *,
    eval_data: Optional[Dict[str, Any]] = None,
) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    summary = report.get("summary") or {}
    if eval_data is None and isinstance(report.get("results"), list):
        eval_data = report
    results = list((eval_data or {}).get("results") or [])
    systems = _systems(summary)

    wb = Workbook()
    default = wb.active
    default.title = "tmp"

    toc: List[Tuple[str, str]] = []
    toc.append((_add_meta(wb, report, eval_data), "评测配置、模型、题量、耗时"))
    toc.append((_add_overall(wb, summary), "两路总体正确率 / 耗时 / Token / 文档召回"))
    cmp_name = _add_comparison(wb, summary)
    if cmp_name:
        toc.append((cmp_name, "配对胜负、双方对错、评判交叉表"))
    hop_name = _add_by_hop(wb, summary)
    if hop_name:
        toc.append((hop_name, "按跳正确率、三路检索延迟、改写/重排/预加载"))
    rec_name = _add_doc_recall(wb, summary)
    if rec_name:
        toc.append((rec_name, "文档召回 × 答案正确率列联表"))
    toc.append((_add_latency(wb, summary), "检索各阶段耗时和 token 用量"))

    if results:
        d_name, diff_name, fail_name = _add_details(wb, results, systems)
        toc.append((d_name, "每题两路评判、答案、召回、分阶段延迟"))
        if diff_name:
            toc.append((diff_name, "评判不一致或有胜负的题"))
        if fail_name:
            toc.append((fail_name, "问答或评判失败清单"))
    else:
        ws = wb.create_sheet(_sheet_name("逐题明细"))
        ws["A1"] = "未提供 evals JSON，没有逐题明细。"
        ws["A1"].font = _FONT_CELL
        toc.append((ws.title, "无逐题数据（缺少 evals）"))

    _add_toc(wb, toc)
    if "tmp" in wb.sheetnames:
        del wb["tmp"]

    wb.save(str(out))
    return out
