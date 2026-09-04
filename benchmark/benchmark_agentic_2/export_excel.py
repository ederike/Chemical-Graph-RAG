"""把 report + evals 摊成多 sheet Excel。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from .utils import SYSTEM_AGENTIC, SYSTEM_LLM_ONLY, resolve_path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:  # pragma: no cover
    raise ImportError("export_excel 需要 openpyxl") from e


def _sheet_name(name: str, used: Optional[set] = None) -> str:
    bad = set(r"[]:*?/\\")
    s = "".join("_" if c in bad else c for c in (name or "Sheet"))[:31] or "Sheet"
    used = used if used is not None else set()
    base, i = s, 1
    while s in used:
        suffix = f"_{i}"
        s = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(s)
    return s


def _cell(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    return s[:32000]


def _num(v: Any) -> Optional[float]:
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def _auto_width(ws, *, min_w: float = 8, max_w: float = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _write_table(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]], start_row: int = 1) -> int:
    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E2F3")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = header_font
        cell.fill = fill
    r = start_row + 1
    for row in rows:
        for c, v in enumerate(row, 1):
            ws.cell(r, c, _cell(v))
        r += 1
    _auto_width(ws)
    return r


def _systems(summary: Dict[str, Any]) -> List[str]:
    systems = summary.get("systems") or {}
    order = summary.get("system_order") or [SYSTEM_AGENTIC, SYSTEM_LLM_ONLY]
    return [k for k in order if k in systems] or list(systems.keys())


def _sys_label(key: str) -> str:
    return {"agentic": "Agentic", "llm_only": "纯LLM"}.get(key, key)


def _add_meta(wb: Workbook, report: Dict[str, Any], eval_data: Optional[Dict[str, Any]]) -> str:
    ws = wb.create_sheet(_sheet_name("运行信息"))
    meta = report.get("meta") or {}
    eval_meta = (eval_data or {}).get("meta") or meta.get("eval_meta") or {}
    rows = [
        ("created_at", report.get("created_at")),
        ("n_results", report.get("n_results")),
        ("category_fields", ", ".join(meta.get("category_fields") or [])),
    ]
    cfg = eval_meta.get("config") or {}
    for k, v in cfg.items():
        rows.append((f"config.{k}", v))
    ds = eval_meta.get("dataset") or {}
    for k, v in ds.items():
        if k == "headers":
            continue
        rows.append((f"dataset.{k}", v if not isinstance(v, (list, dict)) else str(v)[:500]))
    _write_table(ws, ["字段", "值"], rows)
    return ws.title


def _add_overall(wb: Workbook, summary: Dict[str, Any]) -> str:
    ws = wb.create_sheet(_sheet_name("汇总"))
    headers = ["系统", "题数", "正确率", "部分正确率", "错误率", "均分",
               "mean_query_s", "mean_wall_s", "mean_retrieve_s",
               "mean_n_turns", "mean_n_search", "mean_search_s", "mean_turn_s",
               "mean_embed_s", "mean_rerank_s", "query_fail", "judge_fail"]
    rows = []
    for sys in _systems(summary):
        s = (summary.get("systems") or {}).get(sys) or {}
        acc = s.get("llm_acc") or {}
        lat = s.get("latency") or {}
        pipe = s.get("pipeline") or {}
        rows.append([
            _sys_label(sys),
            s.get("n_total"),
            acc.get("accuracy"),
            acc.get("partial_rate"),
            acc.get("error_rate"),
            acc.get("mean_score"),
            lat.get("mean_query_s"),
            lat.get("mean_wall_s"),
            lat.get("mean_retrieve_s"),
            lat.get("mean_n_turns"),
            lat.get("mean_n_search"),
            lat.get("mean_search_s"),
            lat.get("mean_turn_s"),
            lat.get("mean_embed_s"),
            lat.get("mean_rerank_s"),
            pipe.get("n_query_fail"),
            pipe.get("n_judge_fail"),
        ])
    _write_table(ws, headers, rows)

    cmp_ = summary.get("comparison") or {}
    pw = cmp_.get("pairwise") or {}
    if pw:
        ws.append([])
        ws.append(["对比胜负", "值"])
        for k in ("n", "agentic_win", "llm_only_win", "tie",
                  "agentic_win_rate", "llm_only_win_rate", "tie_rate"):
            ws.append([k, pw.get(k)])
        ws.append(["mean_score_delta_ag-llm", cmp_.get("mean_score_delta_agentic_minus_llm")])
    return ws.title


def _add_latency(wb: Workbook, summary: Dict[str, Any]) -> str:
    ws = wb.create_sheet(_sheet_name("延迟分解"))
    ag = ((summary.get("systems") or {}).get(SYSTEM_AGENTIC) or {}).get("latency") or {}
    stages = ag.get("stages") or {}
    headers = ["阶段", "mean_per_question_s", "mean_per_search_s", "sum_s", "n"]
    rows = []
    for k, block in stages.items():
        rows.append([
            block.get("label") or k,
            block.get("mean_per_question_s"),
            block.get("mean_per_search_s"),
            block.get("sum_s"),
            block.get("n"),
        ])
    # top-level agentic metrics
    rows.append([])
    rows.append(["mean_n_turns", ag.get("mean_n_turns"), None, ag.get("sum_n_turns"), None])
    rows.append(["mean_n_search", ag.get("mean_n_search"), None, ag.get("sum_n_search"), None])
    rows.append(["mean_search_s", ag.get("mean_search_s"), None, None, None])
    rows.append(["mean_turn_s", ag.get("mean_turn_s"), None, None, None])
    _write_table(ws, headers, rows)
    return ws.title


def _add_comparison(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    cmp_ = summary.get("comparison")
    if not cmp_:
        return None
    ws = wb.create_sheet(_sheet_name("对比胜负"))
    rows = [
        ("paired", cmp_.get("n_paired")),
        ("both_correct", cmp_.get("both_correct")),
        ("agentic_only_correct", cmp_.get("agentic_only_correct")),
        ("llm_only_only_correct", cmp_.get("llm_only_only_correct")),
        ("neither_correct", cmp_.get("neither_correct")),
    ]
    pw = cmp_.get("pairwise") or {}
    for k, v in pw.items():
        rows.append((f"pairwise.{k}", v))
    xt = cmp_.get("crosstab") or {}
    for k, v in xt.items():
        rows.append((f"crosstab.{k}", v))
    _write_table(ws, ["指标", "值"], rows)
    return ws.title


def _add_category(wb: Workbook, summary: Dict[str, Any]) -> Optional[str]:
    # flatten by_category for agentic
    ag = (summary.get("systems") or {}).get(SYSTEM_AGENTIC) or {}
    by_cat = ag.get("by_category") or {}
    if not by_cat:
        return None
    ws = wb.create_sheet(_sheet_name("分类统计"))
    headers = ["字段", "标签", "n", "正确率", "部分正确率", "均分",
               "mean_query_s", "mean_wall_s", "mean_n_turns", "mean_n_search"]
    rows = []
    for field, labels in by_cat.items():
        for label, g in (labels or {}).items():
            acc = g.get("llm_acc") or {}
            rows.append([
                field, label, g.get("n"),
                acc.get("accuracy"), acc.get("partial_rate"), g.get("mean_score"),
                g.get("mean_query_latency_s"), g.get("mean_wall_latency_s"),
                g.get("mean_n_turns"), g.get("mean_n_search"),
            ])
    _write_table(ws, headers, rows)
    return ws.title


def _result_block(result: Dict[str, Any], system: str) -> Dict[str, Any]:
    block = result.get(system)
    return block if isinstance(block, dict) else {}


def _add_details(wb: Workbook, eval_data: Optional[Dict[str, Any]]) -> Optional[str]:
    if not eval_data:
        return None
    results = list(eval_data.get("results") or [])
    if not results:
        return None
    # discover category fields
    cat_fields: List[str] = []
    for r in results:
        for k in (r.get("categories") or {}).keys():
            if k not in cat_fields:
                cat_fields.append(k)

    headers = [
        "dataset_id", "id", "hop", "question",
        *cat_fields,
        "agentic_answer", "agentic_score", "agentic_judgment",
        "agentic_query_s", "agentic_wall_s", "agentic_retrieve_s",
        "n_turns", "n_search", "mean_search_s",
        "embed_s", "rerank_s",
        "llm_answer", "llm_score", "llm_judgment", "llm_query_s",
        "winner", "compare_reason",
        "recall",
        "agentic_error", "llm_error",
    ]
    rows = []
    for r in results:
        ag = _result_block(r, SYSTEM_AGENTIC)
        lo = _result_block(r, SYSTEM_LLM_ONLY)
        am = ag.get("metrics") or {}
        lm = lo.get("metrics") or {}
        cats = r.get("categories") or {}
        cmp_ = r.get("comparison") or {}
        rec = r.get("recall") or {}
        row = [
            r.get("dataset_id"), r.get("id"), r.get("hop"), r.get("question"),
            *[cats.get(f) for f in cat_fields],
            ag.get("answer"), ag.get("score"), ag.get("llm_acc"),
            am.get("query_latency_s"), am.get("wall_latency_s"), am.get("retrieve_latency_s"),
            am.get("n_turns"), am.get("n_search"), am.get("mean_search_s"),
            am.get("embed_latency_s") or am.get("embed_s"),
            am.get("rerank_latency_s") or am.get("rerank_s"),
            lo.get("answer"), lo.get("score"), lo.get("llm_acc"), lm.get("query_latency_s"),
            cmp_.get("winner"), cmp_.get("reason"),
            (rec.get("recall") if isinstance(rec, dict) else None),
            ag.get("query_error") or ag.get("judge_error"),
            lo.get("query_error") or lo.get("judge_error"),
        ]
        rows.append(row)
    ws = wb.create_sheet(_sheet_name("逐题明细"))
    _write_table(ws, headers, rows)
    return ws.title


def _add_per_dataset(wb: Workbook, report: Dict[str, Any]) -> Optional[str]:
    items = report.get("per_dataset") or []
    if not items:
        return None
    ws = wb.create_sheet(_sheet_name("分测试集"))
    headers = ["dataset_id", "系统", "n", "accuracy", "mean_score",
               "mean_wall_s", "mean_n_turns", "mean_n_search",
               "ag_win", "llm_win", "tie"]
    rows = []
    for item in items:
        did = item.get("dataset_id")
        summary = item.get("summary") or {}
        pw = ((summary.get("comparison") or {}).get("pairwise") or {})
        for sys in _systems(summary):
            s = (summary.get("systems") or {}).get(sys) or {}
            acc = s.get("llm_acc") or {}
            lat = s.get("latency") or {}
            rows.append([
                did, _sys_label(sys), s.get("n_total"),
                acc.get("accuracy"), acc.get("mean_score"),
                lat.get("mean_wall_s"), lat.get("mean_n_turns"), lat.get("mean_n_search"),
                pw.get("agentic_win") if sys == SYSTEM_AGENTIC else None,
                pw.get("llm_only_win") if sys == SYSTEM_AGENTIC else None,
                pw.get("tie") if sys == SYSTEM_AGENTIC else None,
            ])
    _write_table(ws, headers, rows)
    return ws.title


def write_report_workbook(
    report: Dict[str, Any],
    path: Any,
    *,
    eval_data: Optional[Dict[str, Any]] = None,
) -> Path:
    out = Path(path)
    if not out.is_absolute():
        out = resolve_path(out)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    summary = report.get("summary") or {}
    _add_meta(wb, report, eval_data)
    _add_overall(wb, summary)
    _add_per_dataset(wb, report)
    _add_comparison(wb, summary)
    _add_latency(wb, summary)
    _add_category(wb, summary)
    _add_details(wb, eval_data)

    wb.save(str(out))
    return out
