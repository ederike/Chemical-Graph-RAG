"""
从 Excel 读取测试集，并结构化构建统计 JSON。

支持两种表结构（自动识别）：
  - 测试集2：问题文本 / 预期答案要点 / 评分维度 + 分布统计
  - 测试集1：问题描述 / 标准答案要点 / 评估要点 + 统计概览
"""

from __future__ import annotations

import re
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

from .utils import resolve_path

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise ImportError("读取 Excel 需要 openpyxl，请 pip install openpyxl") from e


# 问题表字段别名 → 规范名
_FIELD_ALIASES = {
    "id": ("序号", "id", "ID", "题号"),
    "question": ("问题文本", "问题描述", "问题", "query"),
    "expected_answer": (
        "预期答案要点",
        "标准答案要点",
        "标准答案",
        "答案要点",
        "ground_truth",
    ),
    "score_dimensions": ("评分维度", "评估要点", "评分要点"),
    "reasoning_path": (
        "参考推理链/解题路径",
        "预期检索路径",
        "推理链",
        "解题路径",
    ),
    "knowledge_source": ("所需知识来源", "知识来源"),
    "note": ("业务场景备注", "备注", "场景备注"),
}

# 统计页优先使用的分类字段（按表型）
SET2_CATEGORY_FIELDS = (
    "身份分类",
    "难度等级",
    "能力维度",
    "问题层面",
    "推理跳数",
)
SET2_EXTRA_FIELDS = ("提问者身份", "所需知识来源")

SET1_CATEGORY_FIELDS = (
    "难度等级",
    "查询类型",
    "问题维度",
    "推理跳数",
    "知识层",
)
SET1_EXTRA_FIELDS = ("材料类别", "产品类别", "应用场景")

_QUESTION_SHEET_HINTS = ("测试集", "测试问题", "问题集", "questions")
_STATS_SHEET_HINTS = ("分布统计", "统计概览", "统计", "stats", "distribution")
_DESIGN_SHEET_HINTS = ("设计说明", "说明", "design")

_DIM_SPLIT_RE = re.compile(r"[；;、|/]+")
_RATIO_NUM_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v != v:  # NaN
            return ""
        if v.is_integer():
            return str(int(v))
        return str(v).rstrip("0").rstrip(".")
    return str(v).strip()


def cell_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v != v:
            return None
        return int(v)
    s = cell_str(v).replace(",", "").replace("%", "")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def parse_ratio(
    v: Any,
    *,
    count: Optional[int] = None,
    total: Optional[int] = None,
) -> Optional[float]:
    """把 0.294 / 29.4% / '29.4%' 统一成 0~1 比例。"""
    if v is None or v == "":
        if count is not None and total:
            return round(count / float(total), 6)
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        x = float(v)
        if x != x:
            return None
        if x > 1.0 + 1e-9:
            x = x / 100.0
        return round(x, 6)
    s = cell_str(v)
    if not s:
        return None
    has_pct = "%" in s
    m = _RATIO_NUM_RE.search(s.replace(",", ""))
    if not m:
        return None
    x = float(m.group(0))
    if has_pct or x > 1.0 + 1e-9:
        x = x / 100.0
    return round(x, 6)


def split_dimensions(text: Any) -> List[str]:
    raw = cell_str(text)
    if not raw:
        return []
    parts = [p.strip() for p in _DIM_SPLIT_RE.split(raw)]
    return [p for p in parts if p]


def _norm_header(name: Any) -> str:
    return cell_str(name).replace("\n", "").replace(" ", "")


def _match_alias(header: str, aliases: Sequence[str]) -> bool:
    h = _norm_header(header)
    if not h:
        return False
    hl = h.lower()
    for a in aliases:
        a_n = _norm_header(a)
        if h == a_n or hl == a_n.lower():
            return True
    return False


def detect_schema(headers: Sequence[str]) -> str:
    """set2 / set1 / generic。"""
    hs = [_norm_header(h) for h in headers]
    if any(h in ("问题文本",) for h in hs) or any(
        h in ("预期答案要点", "评分维度") for h in hs
    ):
        return "set2"
    if any(h in ("问题描述",) for h in hs) or any(
        h in ("标准答案要点", "评估要点") for h in hs
    ):
        return "set1"
    return "generic"


def category_fields_for(schema: str, headers: Sequence[str]) -> List[str]:
    if schema == "set2":
        preferred = list(SET2_CATEGORY_FIELDS) + list(SET2_EXTRA_FIELDS)
    elif schema == "set1":
        preferred = list(SET1_CATEGORY_FIELDS) + list(SET1_EXTRA_FIELDS)
    else:
        preferred = []
    known = set(_norm_header(h) for group in _FIELD_ALIASES.values() for h in group)
    known.update({"序号", "id", "ID"})
    out: List[str] = []
    header_map = {_norm_header(h): h for h in headers if _norm_header(h)}
    for name in preferred:
        raw = header_map.get(_norm_header(name))
        if raw and raw not in out:
            out.append(raw)
    for h in headers:
        hn = _norm_header(h)
        if not hn or hn in known:
            continue
        if h not in out:
            out.append(h)
    return out


def pick_sheet(
    sheetnames: Sequence[str],
    preferred: Optional[str],
    hints: Sequence[str],
    *,
    exclude: Optional[Sequence[str]] = None,
) -> Optional[str]:
    if preferred:
        if preferred in sheetnames:
            return preferred
        low = {s.lower(): s for s in sheetnames}
        if preferred.lower() in low:
            return low[preferred.lower()]
        raise ValueError(f"工作表不存在: {preferred!r}  可选: {list(sheetnames)}")
    excl = set(exclude or [])
    for s in sheetnames:
        if s in excl:
            continue
        sl = s.lower()
        for h in hints:
            if h.lower() in sl:
                return s
    return None


def _sheet_rows(ws) -> List[Tuple[Any, ...]]:
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def parse_design_sheet(ws) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = cell_str(row[0] if len(row) > 0 else None)
        val = cell_str(row[1] if len(row) > 1 else None)
        if key:
            out[key] = val
    return out


def _looks_like_dist_header(cells: Sequence[Any]) -> Optional[str]:
    """
    (身份分类, 数量, 占比) → '身份分类'
    (类别, 数量, 占比) → '__generic__'
    """
    nonempty = [cell_str(c) for c in cells if cell_str(c)]
    if len(nonempty) < 2:
        return None
    a, b = nonempty[0], nonempty[1]
    third = nonempty[2] if len(nonempty) > 2 else ""
    if b in ("数量", "计数", "count", "n") and (
        not third or third in ("占比", "比例", "ratio", "percent", "%")
    ):
        if a in ("类别", "分类", "label", "项"):
            return "__generic__"
        return a
    return None


def _looks_like_section_title(cells: Sequence[Any]) -> Optional[str]:
    """按难度等级分布 / 按查询类型分布 → 难度等级 / 查询类型。"""
    nonempty = [cell_str(c) for c in cells if cell_str(c)]
    if len(nonempty) != 1:
        return None
    title = nonempty[0]
    m = re.match(r"^按(.+?)分布$", title)
    if m:
        return m.group(1)
    return None


def parse_stats_sheet(ws) -> Dict[str, Any]:
    """
    解析统计页中的「按 XX 分布」表格。

    返回::
        {
          "title": str | None,
          "n_questions_declared": int | None,
          "distributions": {field: {field, total, items:[{label,count,ratio}]}}
        }
    """
    rows = _sheet_rows(ws)
    title = None
    n_declared = None
    distributions: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()

    pending_section: Optional[str] = None
    current_field: Optional[str] = None
    current_items: List[Dict[str, Any]] = []

    def _flush():
        nonlocal current_field, current_items
        if current_field and current_items:
            total = sum(it["count"] for it in current_items if it.get("count") is not None)
            for it in current_items:
                if it.get("ratio") is None and it.get("count") is not None and total:
                    it["ratio"] = round(it["count"] / float(total), 6)
            distributions[current_field] = {
                "field": current_field,
                "source": "sheet",
                "total": total,
                "items": current_items,
            }
        current_field = None
        current_items = []

    for idx, row in enumerate(rows):
        cells = list(row) if row else []
        nonempty = [cell_str(c) for c in cells if cell_str(c)]
        if not nonempty:
            _flush()
            pending_section = None
            continue

        # 标题行（仅第一处单单元格长文本）
        if title is None and idx == 0 and len(nonempty) == 1 and len(nonempty[0]) > 8:
            title = nonempty[0]
            continue

        # 总问题数, 1177
        if nonempty[0] in ("总问题数", "问题总数", "n_questions") and len(cells) >= 2:
            n_declared = cell_int(cells[1])
            continue

        section = _looks_like_section_title(cells)
        if section:
            _flush()
            pending_section = section
            continue

        header_field = _looks_like_dist_header(cells)
        if header_field is not None:
            _flush()
            if header_field == "__generic__":
                current_field = pending_section or "未命名分类"
            else:
                current_field = header_field
            pending_section = None
            current_items = []
            continue

        if current_field:
            label = cell_str(cells[0] if cells else None)
            count = cell_int(cells[1] if len(cells) > 1 else None)
            ratio = parse_ratio(
                cells[2] if len(cells) > 2 else None,
                count=count,
            )
            if not label or count is None:
                _flush()
                continue
            current_items.append(
                {
                    "label": label,
                    "count": count,
                    "ratio": ratio,
                }
            )

    _flush()
    return {
        "title": title,
        "n_questions_declared": n_declared,
        "distributions": dict(distributions),
    }


def _build_header_index(headers: Sequence[str]) -> Dict[str, int]:
    """规范字段名 → 列下标。"""
    idx: Dict[str, int] = {}
    for i, h in enumerate(headers):
        hn = _norm_header(h)
        if not hn:
            continue
        for canon, aliases in _FIELD_ALIASES.items():
            if canon in idx:
                continue
            if _match_alias(hn, aliases):
                idx[canon] = i
    return idx


def _make_qid(raw_id: Any, row_i: int) -> str:
    s = cell_str(raw_id)
    if not s:
        return f"Q{row_i:04d}"
    if s.upper().startswith("Q") and s[1:].isdigit():
        return f"Q{int(s[1:]):04d}"
    if s.isdigit():
        return f"Q{int(s):04d}"
    return s


def parse_questions_sheet(
    ws,
    *,
    schema: Optional[str] = None,
) -> Tuple[str, List[str], List[str], List[Dict[str, Any]]]:
    rows = _sheet_rows(ws)
    if not rows:
        raise ValueError("问题表为空")
    headers = [cell_str(c) for c in rows[0]]
    if not any(headers):
        raise ValueError("问题表表头为空")
    schema = schema or detect_schema(headers)
    col = _build_header_index(headers)
    if "question" not in col:
        raise ValueError(
            f"问题表缺少问题文本列。表头={headers} "
            f"需要其一: {_FIELD_ALIASES['question']}"
        )
    cat_fields = category_fields_for(schema, headers)
    questions: List[Dict[str, Any]] = []
    for row_i, row in enumerate(rows[1:], start=2):
        cells = list(row) if row else []

        def _get(canon: str) -> str:
            i = col.get(canon)
            if i is None or i >= len(cells):
                return ""
            return cell_str(cells[i])

        question = _get("question")
        if not question:
            continue
        raw_id = _get("id")
        score_raw = _get("score_dimensions")
        cats: Dict[str, str] = {}
        header_pos = {h: i for i, h in enumerate(headers)}
        for fname in cat_fields:
            i = header_pos.get(fname)
            if i is None or i >= len(cells):
                continue
            cats[fname] = cell_str(cells[i]) or "NA"

        qid = _make_qid(raw_id, row_i - 1)
        item = {
            "id": qid,
            "row": row_i,
            "raw_id": raw_id,
            "question": question,
            "expected_answer": _get("expected_answer"),
            "score_dimensions": split_dimensions(score_raw),
            "score_dimensions_raw": score_raw,
            "reasoning_path": _get("reasoning_path"),
            "knowledge_source": _get("knowledge_source"),
            "note": _get("note"),
            "categories": cats,
        }
        questions.append(item)
    return schema, headers, cat_fields, questions


def compute_distributions(
    questions: Sequence[Dict[str, Any]],
    fields: Sequence[str],
) -> Dict[str, Any]:
    n = len(questions)
    out: Dict[str, Any] = {}
    for field in fields:
        counter: Counter = Counter()
        for q in questions:
            cats = q.get("categories") or {}
            label = cell_str(cats.get(field)) or "NA"
            counter[label] += 1
        items = []
        for label, count in counter.most_common():
            items.append(
                {
                    "label": label,
                    "count": int(count),
                    "ratio": round(count / float(n), 6) if n else None,
                }
            )
        out[field] = {
            "field": field,
            "source": "computed",
            "total": n,
            "n_labels": len(items),
            "items": items,
        }
    return out


def build_stats_document(
    *,
    excel_path: Union[str, Path],
    questions_sheet: Optional[str] = None,
    stats_sheet: Optional[str] = None,
    design_sheet: Optional[str] = None,
) -> Dict[str, Any]:
    """读 Excel，产出结构化统计 JSON（含题目分类计数，不含题目正文）。"""
    path = resolve_path(excel_path)
    if not path.is_file():
        raise FileNotFoundError(f"测试集 Excel 不存在: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        names = list(wb.sheetnames)
        q_name = pick_sheet(names, questions_sheet, _QUESTION_SHEET_HINTS)
        if q_name is None:
            # 回退：第一张非统计/说明表
            skip = set()
            s_try = pick_sheet(names, None, _STATS_SHEET_HINTS)
            d_try = pick_sheet(names, None, _DESIGN_SHEET_HINTS)
            if s_try:
                skip.add(s_try)
            if d_try:
                skip.add(d_try)
            q_name = next((s for s in names if s not in skip), names[0])

        s_name = pick_sheet(
            names, stats_sheet, _STATS_SHEET_HINTS, exclude=[q_name]
        )
        d_name = pick_sheet(
            names,
            design_sheet,
            _DESIGN_SHEET_HINTS,
            exclude=[q_name, s_name] if s_name else [q_name],
        )

        schema, headers, cat_fields, questions = parse_questions_sheet(wb[q_name])
        sheet_stats = parse_stats_sheet(wb[s_name]) if s_name else {
            "title": None,
            "n_questions_declared": None,
            "distributions": {},
        }
        design = parse_design_sheet(wb[d_name]) if d_name else {}
    finally:
        wb.close()

    n = len(questions)
    computed = compute_distributions(questions, cat_fields)

    # 统计页字段优先作为「主分类」；没有则用 schema 默认
    sheet_fields = list((sheet_stats.get("distributions") or {}).keys())
    if schema == "set2":
        primary = [f for f in SET2_CATEGORY_FIELDS if f in cat_fields or f in sheet_fields]
    elif schema == "set1":
        primary = [f for f in SET1_CATEGORY_FIELDS if f in cat_fields or f in sheet_fields]
    else:
        primary = list(sheet_fields) or list(cat_fields)
    # 补上统计页有、默认列表没有的字段
    for f in sheet_fields:
        if f not in primary:
            primary.append(f)

    extra = [f for f in cat_fields if f not in primary]

    dim_counter: Counter = Counter()
    for q in questions:
        for d in q.get("score_dimensions") or []:
            dim_counter[d] += 1
    score_dimension_dist = [
        {
            "label": lab,
            "count": int(c),
            "ratio": round(c / float(n), 6) if n else None,
        }
        for lab, c in dim_counter.most_common()
    ]

    return {
        "schema_version": 1,
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "source_excel": str(path),
            "source_excel_name": path.name,
            "schema": schema,
            "questions_sheet": q_name,
            "stats_sheet": s_name,
            "design_sheet": d_name,
            "n_questions": n,
            "n_questions_declared": sheet_stats.get("n_questions_declared"),
            "title": sheet_stats.get("title") or design.get("测试集名称") or path.stem,
            "headers": headers,
        },
        "design": design,
        "category_fields": {
            "primary": primary,
            "extra": extra,
            "all": list(cat_fields),
        },
        "distributions": sheet_stats.get("distributions") or {},
        "computed_distributions": computed,
        "score_dimensions": {
            "n_unique": len(score_dimension_dist),
            "items": score_dimension_dist,
        },
    }


def load_excel_dataset(
    *,
    excel_path: Union[str, Path],
    questions_sheet: Optional[str] = None,
    stats_sheet: Optional[str] = None,
    design_sheet: Optional[str] = None,
    n_limit: Optional[int] = None,
    id_list: Optional[Iterable[str]] = None,
) -> Dict[str, Any]:
    """读 Excel，返回 {meta, stats, questions}，供评测使用。"""
    path = resolve_path(excel_path)
    if not path.is_file():
        raise FileNotFoundError(f"测试集 Excel 不存在: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        names = list(wb.sheetnames)
        q_name = pick_sheet(names, questions_sheet, _QUESTION_SHEET_HINTS)
        if q_name is None:
            skip = set()
            s_try = pick_sheet(names, None, _STATS_SHEET_HINTS)
            d_try = pick_sheet(names, None, _DESIGN_SHEET_HINTS)
            if s_try:
                skip.add(s_try)
            if d_try:
                skip.add(d_try)
            q_name = next((s for s in names if s not in skip), names[0])
        s_name = pick_sheet(names, stats_sheet, _STATS_SHEET_HINTS, exclude=[q_name])
        d_name = pick_sheet(
            names,
            design_sheet,
            _DESIGN_SHEET_HINTS,
            exclude=[q_name, s_name] if s_name else [q_name],
        )
        schema, headers, cat_fields, questions = parse_questions_sheet(wb[q_name])
        # 统计文档走统一入口（再开一次只读即可；题目已在内存）
    finally:
        wb.close()

    stats = build_stats_document(
        excel_path=path,
        questions_sheet=questions_sheet or q_name,
        stats_sheet=stats_sheet or s_name,
        design_sheet=design_sheet or d_name,
    )

    if id_list:
        want = set()
        for x in id_list:
            s = cell_str(x)
            if not s:
                continue
            want.add(s)
            want.add(_make_qid(s, 0))
            if s.isdigit():
                want.add(str(int(s)))
                want.add(f"Q{int(s):04d}")
        questions = [
            q
            for q in questions
            if q.get("id") in want or cell_str(q.get("raw_id")) in want
        ]
    if n_limit is not None and int(n_limit) >= 0:
        questions = list(questions)[: int(n_limit)]

    return {
        "meta": {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "source_excel": str(path),
            "schema": schema,
            "questions_sheet": q_name,
            "stats_sheet": s_name,
            "n_questions": len(questions),
            "n_questions_full": stats["meta"]["n_questions"],
            "n_limit": n_limit,
            "headers": headers,
            "category_fields": cat_fields,
        },
        "stats": stats,
        "questions": questions,
    }


def build_stats_cli(excel_path: Union[str, Path], out_path: Union[str, Path]) -> Path:
    """命令行入口：Excel → 统计 JSON。"""
    import json

    stats = build_stats_document(excel_path=excel_path)
    out = Path(out_path)
    if not out.is_absolute():
        out = resolve_path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps(stats, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out
